import openpyxl
from itertools import combinations


def score_high_and_often(rolls):
    """Highest number rolled * count of that number"""
    highest = max(rolls)
    count = rolls.count(highest)
    return highest * count


def score_summation(rolls):
    """Sum of all six dice"""
    return sum(rolls)


def score_highs_and_lows(rolls):
    """Highest * lowest * (highest - lowest)"""
    highest = max(rolls)
    lowest = min(rolls)
    return highest * lowest * (highest - lowest)


def score_only_two_numbers(rolls):
    """If all 6 rolls use exactly two distinct numbers, score = 30"""
    if len(set(rolls)) == 2:
        return 30
    return 0


def score_all_the_numbers(rolls):
    """If rolls are 1-2-3-4-5-6 in any order, score = 40"""
    if sorted(rolls) == [1, 2, 3, 4, 5, 6]:
        return 40
    return 0


def score_ordered_subset_of_four(rolls):
    """If the rolls (in order) contain a subsequence of 4 consecutive
    increasing or decreasing integers, score = 50.
    E.g. [2,3,4,6,4,5] contains subsequence 2,3,4,5 at positions 0,1,2,5."""
    for combo in combinations(range(len(rolls)), 4):
        sub = [rolls[i] for i in combo]
        increasing = all(sub[j+1] == sub[j] + 1 for j in range(3))
        decreasing = all(sub[j+1] == sub[j] - 1 for j in range(3))
        if increasing or decreasing:
            return 50
    return 0


CATEGORY_NAMES = [
    'high_and_often', 'summation', 'highs_and_lows',
    'only_two_numbers', 'all_the_numbers', 'ordered_subset_of_four'
]
CATEGORY_FUNCS = [
    score_high_and_often, score_summation, score_highs_and_lows,
    score_only_two_numbers, score_all_the_numbers, score_ordered_subset_of_four
]


def compute_all_scores(rolls):
    """Compute score for each category for a given turn's rolls."""
    return {name: func(rolls) for name, func in zip(CATEGORY_NAMES, CATEGORY_FUNCS)}


def best_game_score(turns_in_game):
    """Find highest combined score from turns where no category is reused.
    Each turn is scored in exactly one category, and the two turns must use
    different categories. Game score = sum of the two chosen category scores.
    For 1 turn: pick the best single category."""
    scores_list = [compute_all_scores(t['rolls']) for t in turns_in_game]
    
    if len(scores_list) == 1:
        return max(scores_list[0].values())
    
    best = 0
    for c1 in CATEGORY_NAMES:
        for c2 in CATEGORY_NAMES:
            if c1 != c2:
                total = scores_list[0][c1] + scores_list[1][c2]
                if total > best:
                    best = total
    return best


def parse_dice_data(filepath):
    """Parse Excel workbook and return (turns_list, games_dict)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[-1]]
    
    data_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    
    header_idx = None
    for i, row in enumerate(data_rows):
        for cell in row:
            if cell == 'Turn number':
                header_idx = i
                break
        if header_idx is not None:
            break
    
    header = data_rows[header_idx]
    col_map = {}
    for j, val in enumerate(header):
        if val == 'Turn number':
            col_map['turn'] = j
        elif val == 'Game number':
            col_map['game'] = j
        elif val and str(val).startswith('Roll'):
            col_map.setdefault('rolls', []).append(j)
    
    turns = []
    games = {}
    for i in range(header_idx + 1, len(data_rows)):
        row = data_rows[i]
        turn_num = row[col_map['turn']]
        game_num = row[col_map['game']]
        if turn_num is not None and game_num is not None:
            rolls = [int(row[j]) for j in col_map['rolls']]
            t = {'turn_num': int(turn_num), 'game_num': int(game_num), 'rolls': rolls}
            turns.append(t)
            games.setdefault(int(game_num), []).append(t)
    
    return turns, games


def compute_all_game_scores(games):
    """Compute optimal game score for each game."""
    game_scores = {}
    for gn in sorted(games.keys()):
        game_turns = sorted(games[gn], key=lambda x: x['turn_num'])
        game_scores[gn] = best_game_score(game_turns)
    return game_scores


def match_players(game_scores):
    """Match odd vs even games. Odd = P1, Even = P2.
    Returns (p1_wins, p2_wins, ties)."""
    max_game = max(game_scores.keys())
    p1_wins = 0
    p2_wins = 0
    ties = 0
    for g in range(1, max_game, 2):
        g_odd = g
        g_even = g + 1
        if g_odd in game_scores and g_even in game_scores:
            s1 = game_scores[g_odd]
            s2 = game_scores[g_even]
            if s1 > s2:
                p1_wins += 1
            elif s2 > s1:
                p2_wins += 1
            else:
                ties += 1
    return p1_wins, p2_wins, ties
