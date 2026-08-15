import openpyxl
from copy import copy


def load_workbook(path):
    """Load workbook preserving everything."""
    return openpyxl.load_workbook(path)


def find_missing_cells(wb):
    """Find all cells containing '???' across all sheets."""
    missing = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == '???':
                    missing.append((sname, cell.coordinate, cell.row, cell.column))
    return missing


def read_sheet_data(ws):
    """Read a worksheet into a dict of {(row,col): value}."""
    data = {}
    for row in ws.iter_rows():
        for cell in row:
            data[(cell.row, cell.column)] = cell.value
    return data


def get_row_label(ws, row):
    """Get the label in column A for a given row."""
    return ws.cell(row=row, column=1).value


def get_col_label(ws, col):
    """Get the label in row 3 (header row) for a given column."""
    return ws.cell(row=3, column=col).value


def sum_row_except(ws, row, exclude_col, min_col=2, max_col=None):
    """Sum numeric values in a row, excluding a specific column."""
    if max_col is None:
        max_col = ws.max_column
    total = 0
    for c in range(min_col, max_col + 1):
        if c == exclude_col:
            continue
        v = ws.cell(row=row, column=c).value
        if isinstance(v, (int, float)):
            total += v
    return total


def compute_yoy_change(current, previous):
    """Compute year-over-year percentage change."""
    if previous == 0:
        return 0
    return round(100 * (current - previous) / previous, 2)


def compute_share(value, total):
    """Compute percentage share."""
    if total == 0:
        return 0
    return round(100 * value / total, 2)


def compute_cagr(start_val, end_val, n_periods):
    """Compute compound annual growth rate."""
    if start_val <= 0:
        return 0
    return round(100 * ((end_val / start_val) ** (1 / n_periods) - 1), 2)


def compute_average(values):
    """Compute average, rounded to 1 decimal place."""
    if not values:
        return 0
    return round(sum(values) / len(values), 1)


def recover_budget_values(wb):
    """Recover missing budget values from the Budget by Directorate sheet.
    Uses Total = sum of components relationship.
    Returns dict of {(row, col): recovered_value}."""
    ws = wb['Budget by Directorate']
    recovered = {}
    total_col = ws.max_column  # K = Total column
    
    # Iterative solving: some values depend on others
    # Pass 1: Solve rows where Total is known but one component is missing
    # Pass 2: Solve rows where all components are known but Total is missing
    for _ in range(3):  # Multiple passes for dependencies
        for row in range(4, ws.max_row + 1):
            missing_in_row = []
            for col in range(2, total_col + 1):
                v = ws.cell(row=row, column=col).value
                if v == '???' and (row, col) not in recovered:
                    missing_in_row.append(col)
            
            if len(missing_in_row) == 1:
                miss_col = missing_in_row[0]
                if miss_col == total_col:
                    # Total is missing - sum all components
                    total = 0
                    for c in range(2, total_col):
                        v = ws.cell(row=row, column=c).value
                        if (row, c) in recovered:
                            v = recovered[(row, c)]
                        if isinstance(v, (int, float)):
                            total += v
                    recovered[(row, total_col)] = total
                else:
                    # A component is missing - derive from total minus others
                    total_val = ws.cell(row=row, column=total_col).value
                    if (row, total_col) in recovered:
                        total_val = recovered[(row, total_col)]
                    if isinstance(total_val, (int, float)):
                        others_sum = 0
                        for c in range(2, total_col):
                            if c == miss_col:
                                continue
                            v = ws.cell(row=row, column=c).value
                            if (row, c) in recovered:
                                v = recovered[(row, c)]
                            if isinstance(v, (int, float)):
                                others_sum += v
                        recovered[(row, miss_col)] = total_val - others_sum
            
            elif len(missing_in_row) == 2:
                # Two missing: try to recover one from cross-sheet data
                # Use YoY changes to recover component, then derive Total
                pass
    
    return recovered


def recover_from_yoy(wb, budget_recovered):
    """Use YoY changes sheet to recover budget values that can't be solved by totals alone.
    Also recover missing YoY values."""
    ws_budget = wb['Budget by Directorate']
    ws_yoy = wb['YoY Changes (%)']
    recovered_budget = dict(budget_recovered)
    recovered_yoy = {}
    
    # Map fiscal years to rows in budget sheet
    budget_year_rows = {}
    for row in range(4, ws_budget.max_row + 1):
        yr = ws_budget.cell(row=row, column=1).value
        if isinstance(yr, (int, float)):
            budget_year_rows[int(yr)] = row
    
    # Map fiscal years to rows in YoY sheet
    yoy_year_rows = {}
    for row in range(4, ws_yoy.max_row + 1):
        yr = ws_yoy.cell(row=row, column=1).value
        if isinstance(yr, (int, float)):
            yoy_year_rows[int(yr)] = row
    
    # For each YoY cell that is '???', compute from budget values
    for row in range(4, ws_yoy.max_row + 1):
        for col in range(2, ws_yoy.max_column + 1):
            if ws_yoy.cell(row=row, column=col).value == '???':
                yr = int(ws_yoy.cell(row=row, column=1).value)
                prev_yr = yr - 1
                if prev_yr in budget_year_rows and yr in budget_year_rows:
                    brow_curr = budget_year_rows[yr]
                    brow_prev = budget_year_rows[prev_yr]
                    
                    curr_val = ws_budget.cell(row=brow_curr, column=col).value
                    if (brow_curr, col) in recovered_budget:
                        curr_val = recovered_budget[(brow_curr, col)]
                    
                    prev_val = ws_budget.cell(row=brow_prev, column=col).value
                    if (brow_prev, col) in recovered_budget:
                        prev_val = recovered_budget[(brow_prev, col)]
                    
                    if isinstance(curr_val, (int, float)) and isinstance(prev_val, (int, float)):
                        yoy_val = compute_yoy_change(curr_val, prev_val)
                        recovered_yoy[(row, col)] = yoy_val
    
    # For budget cells still missing, try to recover from YoY
    for yr in sorted(budget_year_rows.keys()):
        brow = budget_year_rows[yr]
        prev_yr = yr - 1
        if prev_yr not in budget_year_rows:
            continue
        brow_prev = budget_year_rows[prev_yr]
        
        for col in range(2, ws_budget.max_column):
            curr_val = ws_budget.cell(row=brow, column=col).value
            if (brow, col) in recovered_budget:
                curr_val = recovered_budget[(brow, col)]
            
            if curr_val == '???':
                # Try to get from YoY
                if yr in yoy_year_rows:
                    yoy_row = yoy_year_rows[yr]
                    yoy_val = ws_yoy.cell(row=yoy_row, column=col).value
                    if (yoy_row, col) in recovered_yoy:
                        yoy_val = recovered_yoy[(yoy_row, col)]
                    
                    prev_val = ws_budget.cell(row=brow_prev, column=col).value
                    if (brow_prev, col) in recovered_budget:
                        prev_val = recovered_budget[(brow_prev, col)]
                    
                    if isinstance(yoy_val, (int, float)) and isinstance(prev_val, (int, float)):
                        new_val = round(prev_val * (1 + yoy_val / 100))
                        recovered_budget[(brow, col)] = new_val
    
    return recovered_budget, recovered_yoy


def recover_shares(wb, budget_recovered):
    """Recover missing share values from budget data."""
    ws_budget = wb['Budget by Directorate']
    ws_shares = wb['Directorate Shares (%)']
    recovered_shares = {}
    total_col_budget = ws_budget.max_column
    
    budget_year_rows = {}
    for row in range(4, ws_budget.max_row + 1):
        yr = ws_budget.cell(row=row, column=1).value
        if isinstance(yr, (int, float)):
            budget_year_rows[int(yr)] = row
    
    for row in range(4, ws_shares.max_row + 1):
        for col in range(2, ws_shares.max_column + 1):
            if ws_shares.cell(row=row, column=col).value == '???':
                yr = int(ws_shares.cell(row=row, column=1).value)
                if yr in budget_year_rows:
                    brow = budget_year_rows[yr]
                    comp_val = ws_budget.cell(row=brow, column=col).value
                    if (brow, col) in budget_recovered:
                        comp_val = budget_recovered[(brow, col)]
                    
                    total_val = ws_budget.cell(row=brow, column=total_col_budget).value
                    if (brow, total_col_budget) in budget_recovered:
                        total_val = budget_recovered[(brow, total_col_budget)]
                    
                    if isinstance(comp_val, (int, float)) and isinstance(total_val, (int, float)):
                        share = compute_share(comp_val, total_val)
                        recovered_shares[(row, col)] = share
    
    return recovered_shares


def recover_growth(wb, budget_recovered):
    """Recover missing growth analysis values."""
    ws_budget = wb['Budget by Directorate']
    ws_growth = wb['Growth Analysis']
    recovered_growth = {}
    
    budget_year_rows = {}
    for row in range(4, ws_budget.max_row + 1):
        yr = ws_budget.cell(row=row, column=1).value
        if isinstance(yr, (int, float)):
            budget_year_rows[int(yr)] = row
    
    # Determine the year range from the sheet title or labels
    # Title says "5-Year Growth Analysis (FY2019-2024)"
    start_year = 2019
    end_year = 2024
    n_periods = end_year - start_year  # 5
    
    for row in range(4, ws_growth.max_row + 1):
        metric = ws_growth.cell(row=row, column=1).value
        for col in range(2, ws_growth.max_column + 1):
            if ws_growth.cell(row=row, column=col).value == '???':
                if metric and '5-Year CAGR' in str(metric):
                    # Need start and end values
                    start_row = None
                    end_row = None
                    for r in range(4, ws_growth.max_row + 1):
                        m = ws_growth.cell(row=r, column=1).value
                        if m and 'FY2019' in str(m):
                            start_row = r
                        elif m and 'FY2024' in str(m):
                            end_row = r
                    
                    start_val = ws_growth.cell(row=start_row, column=col).value
                    if (start_row, col) in recovered_growth:
                        start_val = recovered_growth[(start_row, col)]
                    # Also check budget recovered
                    if start_val == '???':
                        if start_year in budget_year_rows:
                            brow = budget_year_rows[start_year]
                            bv = ws_budget.cell(row=brow, column=col).value
                            if (brow, col) in budget_recovered:
                                bv = budget_recovered[(brow, col)]
                            if isinstance(bv, (int, float)):
                                start_val = bv
                                recovered_growth[(start_row, col)] = bv
                    
                    end_val = ws_growth.cell(row=end_row, column=col).value
                    if (end_row, col) in recovered_growth:
                        end_val = recovered_growth[(end_row, col)]
                    
                    if isinstance(start_val, (int, float)) and isinstance(end_val, (int, float)):
                        cagr = compute_cagr(start_val, end_val, n_periods)
                        recovered_growth[(row, col)] = cagr
                
                elif metric and 'FY2019' in str(metric):
                    # FY2019 budget value
                    if start_year in budget_year_rows:
                        brow = budget_year_rows[start_year]
                        bv = ws_budget.cell(row=brow, column=col).value
                        if (brow, col) in budget_recovered:
                            bv = budget_recovered[(brow, col)]
                        if isinstance(bv, (int, float)):
                            recovered_growth[(row, col)] = bv
                
                elif metric and 'FY2024' in str(metric):
                    if end_year in budget_year_rows:
                        brow = budget_year_rows[end_year]
                        bv = ws_budget.cell(row=brow, column=col).value
                        if (brow, col) in budget_recovered:
                            bv = budget_recovered[(brow, col)]
                        if isinstance(bv, (int, float)):
                            recovered_growth[(row, col)] = bv
                
                elif metric and '5-Year Change' in str(metric):
                    start_row_g = None
                    end_row_g = None
                    for r in range(4, ws_growth.max_row + 1):
                        m = ws_growth.cell(row=r, column=1).value
                        if m and 'FY2019' in str(m):
                            start_row_g = r
                        elif m and 'FY2024' in str(m):
                            end_row_g = r
                    
                    sv = ws_growth.cell(row=start_row_g, column=col).value
                    if (start_row_g, col) in recovered_growth:
                        sv = recovered_growth[(start_row_g, col)]
                    ev = ws_growth.cell(row=end_row_g, column=col).value
                    if (end_row_g, col) in recovered_growth:
                        ev = recovered_growth[(end_row_g, col)]
                    
                    if isinstance(sv, (int, float)) and isinstance(ev, (int, float)):
                        recovered_growth[(row, col)] = ev - sv
                
                elif metric and 'Avg Annual' in str(metric):
                    # Average of all years from start to end
                    values = []
                    for yr in range(start_year, end_year + 1):
                        if yr in budget_year_rows:
                            brow = budget_year_rows[yr]
                            bv = ws_budget.cell(row=brow, column=col).value
                            if (brow, col) in budget_recovered:
                                bv = budget_recovered[(brow, col)]
                            if isinstance(bv, (int, float)):
                                values.append(bv)
                    if values:
                        recovered_growth[(row, col)] = compute_average(values)
    
    return recovered_growth


def apply_recoveries(wb, sheet_name, recovered):
    """Apply recovered values to a worksheet."""
    ws = wb[sheet_name]
    for (row, col), value in recovered.items():
        ws.cell(row=row, column=col).value = value


def run_recovery(input_path, output_path):
    """End-to-end recovery of missing values."""
    wb = load_workbook(input_path)
    
    # Step 1: Initial budget recovery (totals)
    budget_recovered = recover_budget_values(wb)
    
    # Step 2: Use YoY to recover remaining budget values, and recover YoY values
    budget_recovered, yoy_recovered = recover_from_yoy(wb, budget_recovered)
    
    # Step 3: Re-run budget recovery with new values applied
    # Apply budget values temporarily to solve remaining
    for (row, col), val in budget_recovered.items():
        wb['Budget by Directorate'].cell(row=row, column=col).value = val
    
    # Now recover totals for rows that had 2 unknowns (now resolved)
    budget_recovered2 = recover_budget_values(wb)
    budget_recovered.update(budget_recovered2)
    for (row, col), val in budget_recovered2.items():
        wb['Budget by Directorate'].cell(row=row, column=col).value = val
    
    # Re-compute YoY with all budget values
    _, yoy_recovered2 = recover_from_yoy(wb, budget_recovered)
    yoy_recovered.update(yoy_recovered2)
    
    # Step 4: Recover shares
    shares_recovered = recover_shares(wb, budget_recovered)
    
    # Step 5: Recover growth analysis
    growth_recovered = recover_growth(wb, budget_recovered)
    
    # Reload original workbook to apply all changes cleanly
    wb = load_workbook(input_path)
    apply_recoveries(wb, 'Budget by Directorate', budget_recovered)
    apply_recoveries(wb, 'YoY Changes (%)', yoy_recovered)
    apply_recoveries(wb, 'Directorate Shares (%)', shares_recovered)
    apply_recoveries(wb, 'Growth Analysis', growth_recovered)
    
    wb.save(output_path)
    return budget_recovered, yoy_recovered, shares_recovered, growth_recovered


def validate_recovery(output_path):
    """Validate that all '???' have been replaced and values are consistent."""
    wb = load_workbook(output_path)
    issues = []
    
    # Check no '???' remain
    missing = find_missing_cells(wb)
    if missing:
        issues.append(f"Still have {len(missing)} '???' cells: {missing}")
    
    # Validate Budget totals
    ws = wb['Budget by Directorate']
    total_col = ws.max_column
    for row in range(4, ws.max_row + 1):
        row_sum = 0
        for col in range(2, total_col):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)):
                row_sum += v
        total_v = ws.cell(row=row, column=total_col).value
        if isinstance(total_v, (int, float)) and abs(row_sum - total_v) > 0.01:
            issues.append(f"Budget row {row}: sum={row_sum} != total={total_v}")
    
    # Validate YoY values
    ws_yoy = wb['YoY Changes (%)']
    budget_year_rows = {}
    for row in range(4, ws.max_row + 1):
        yr = ws.cell(row=row, column=1).value
        if isinstance(yr, (int, float)):
            budget_year_rows[int(yr)] = row
    
    for row in range(4, ws_yoy.max_row + 1):
        yr = int(ws_yoy.cell(row=row, column=1).value)
        prev_yr = yr - 1
        if prev_yr in budget_year_rows and yr in budget_year_rows:
            for col in range(2, ws_yoy.max_column + 1):
                yoy_v = ws_yoy.cell(row=row, column=col).value
                curr = ws.cell(row=budget_year_rows[yr], column=col).value
                prev = ws.cell(row=budget_year_rows[prev_yr], column=col).value
                if all(isinstance(x, (int, float)) for x in [yoy_v, curr, prev]) and prev != 0:
                    expected = round(100 * (curr - prev) / prev, 2)
                    if abs(yoy_v - expected) > 0.015:
                        issues.append(f"YoY row {row} col {col}: {yoy_v} != expected {expected}")
    
    if issues:
        print("VALIDATION ISSUES:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("VALIDATION PASSED: All values consistent.")
    
    return issues


def recalculate_with_engine(output_path):
    """Recalculate workbook using LibreOffice if available."""
    import subprocess
    import shutil
    import os
    
    lo_path = shutil.which('libreoffice') or shutil.which('soffice')
    if lo_path:
        tmp_dir = '/tmp/lo_recalc'
        os.makedirs(tmp_dir, exist_ok=True)
        result = subprocess.run(
            [lo_path, '--headless', '--calc', '--convert-to',
             'xlsx:Calc MS Excel 2007 XML', '--outdir', tmp_dir, output_path],
            capture_output=True, text=True, timeout=30
        )
        basename = os.path.basename(output_path)
        tmp_out = os.path.join(tmp_dir, basename)
        if os.path.exists(tmp_out):
            shutil.copy2(tmp_out, output_path)
            os.remove(tmp_out)
            print(f"Recalculated with LibreOffice: {output_path}")
            return True
        else:
            print(f"LibreOffice conversion failed: {result.stderr}")
            return False
    else:
        print("No LibreOffice found - skipping recalculation")
        return False


def run_full_pipeline(input_path, output_path):
    """Complete end-to-end pipeline: recover, save, recalculate, validate."""
    budget_rec, yoy_rec, shares_rec, growth_rec = run_recovery(input_path, output_path)
    print(f"Recovered: {len(budget_rec)} budget, {len(yoy_rec)} YoY, {len(shares_rec)} shares, {len(growth_rec)} growth")
    
    # Recalculate with spreadsheet engine
    recalculate_with_engine(output_path)
    
    # Validate
    issues = validate_recovery(output_path)
    return issues
