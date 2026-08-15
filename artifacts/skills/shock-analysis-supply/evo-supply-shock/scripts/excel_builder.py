"""Build the supply-side shock Excel model."""
import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))
from data_collection import collect_pwt_data, collect_weo_data, collect_ecb_cfc_data


def build_supply_model(template_path, output_path):
    """Main entry point: populate the Excel template with data and formulas."""
    print('Collecting PWT data...')
    pwt_data = collect_pwt_data()
    print(f'  Got {len(pwt_data)} years of rnna data')
    
    print('Collecting WEO data...')
    gdp_levels, growth_rates = collect_weo_data()
    print(f'  Got {len(gdp_levels)} years of GDP levels, {len(growth_rates)} years of growth rates')
    
    print('Collecting ECB CFC data...')
    cfc_data = collect_ecb_cfc_data()
    print(f'  Got {len(cfc_data)} years of CFC data')
    
    wb = openpyxl.load_workbook(template_path)
    
    populate_pwt_sheet(wb, pwt_data)
    populate_weo_sheet(wb, gdp_levels, growth_rates)
    populate_cfc_sheet(wb, cfc_data)
    populate_production_sheet(wb)
    
    wb.save(output_path)
    print(f'Saved to {output_path}')
    return output_path


def populate_pwt_sheet(wb, pwt_data):
    """Fill PWT sheet column B with rnna data (millions of national currency)."""
    ws = wb['PWT']
    for row in range(2, 36):
        year = ws.cell(row=row, column=1).value
        if year and year in pwt_data:
            ws.cell(row=row, column=2).value = pwt_data[year]
    print('  PWT sheet populated')


def populate_weo_sheet(wb, gdp_levels, growth_rates):
    """Fill WEO_Data sheet.
    Column C: Real GDP level (billions of national currency)
    Column D: Real GDP growth rate (percent)
    Rows 8-35: years 2000-2027
    Rows 36-51: years 2028-2043 (extended)
    """
    ws = wb['WEO_Data']
    
    # Fill historical data (2000-2023) and WEO forecasts (2024-2027)
    for row in range(8, 36):
        year = ws.cell(row=row, column=2).value
        if year is None:
            continue
        year = int(year)
        
        if year in gdp_levels:
            ws.cell(row=row, column=3).value = gdp_levels[year]
        
        if year in growth_rates:
            ws.cell(row=row, column=4).value = growth_rates[year]
    
    # For 2024-2027: use formula for GDP level based on growth rate
    for row in range(32, 36):  # 2024-2027
        year = ws.cell(row=row, column=2).value
        if year and int(year) >= 2024:
            ws.cell(row=row, column=3).value = f'=C{row-1}*(1+D{row}/100)'
    
    # For 2028-2043 (rows 36-51): extend with constant growth rate from 2027
    for row in range(36, 52):
        year = ws.cell(row=row, column=2).value
        if year is None:
            continue
        ws.cell(row=row, column=4).value = f'=$D$35'
        ws.cell(row=row, column=3).value = f'=C{row-1}*(1+D{row}/100)'
    
    print('  WEO_Data sheet populated')


def populate_cfc_sheet(wb, cfc_data):
    """Fill CFC data sheet.
    Column C: CFC values (Lari, current prices)
    Column D: Capital Stock (linked from PWT, millions of Lari)
    Column E: Depreciation rate = C / (D * 1000000)
    """
    ws = wb['CFC data']
    
    for row in range(2, 30):
        year = ws.cell(row=row, column=2).value
        if year is None:
            continue
        year = int(year)
        
        if year in cfc_data:
            ws.cell(row=row, column=3).value = cfc_data[year]
        
        # Capital Stock from PWT (year -> PWT row)
        pwt_row = year - 1990 + 2
        if 2 <= pwt_row <= 35:
            ws.cell(row=row, column=4).value = f'=PWT!B{pwt_row}'
        
        # Depreciation rate: CFC (Lari) / (Capital Stock in millions * 1e6)
        # = CFC / (D * 1000000)
        ws.cell(row=row, column=5).value = f'=IF(AND(C{row}<>"",D{row}<>""),C{row}/(D{row}*1000000),"")'
    
    print('  CFC data sheet populated')


def populate_production_sheet(wb):
    """Fill Production sheet with all formulas."""
    ws = wb['Production']
    
    # B3: annual depreciation rate = average of most recent 8 years (2016-2023)
    # CFC data rows: 1996=row2, so 2016=row22, 2023=row29
    ws.cell(row=3, column=2).value = '=AVERAGE(\'CFC data\'!E22:\'CFC data\'!E29)'
    
    # ========== HP Filter area (rows 6-27, years 2002-2023) ==========
    for i in range(22):  # 0-21 for years 2002-2023
        row = 6 + i
        year = 2002 + i
        
        # D: K from PWT (millions of Lari)
        pwt_row = year - 1990 + 2
        ws.cell(row=row, column=4).value = f'=PWT!B{pwt_row}'
        
        # E: Real GDP from WEO_Data (billions -> multiply by 1000 for millions)
        weo_row = year - 2000 + 8
        ws.cell(row=row, column=5).value = f'=WEO_Data!C{weo_row}*1000'
        
        # F: LnK = LN(D)
        ws.cell(row=row, column=6).value = f'=LN(D{row})'
        
        # G: LnY = LN(E)
        ws.cell(row=row, column=7).value = f'=LN(E{row})'
        
        # K: LnZ = LnY - alpha * LnK
        ws.cell(row=row, column=11).value = f'=G{row}-$B$2*F{row}'
        
        # L: LnZ_HP (placeholder = LnZ initially)
        ws.cell(row=row, column=12).value = f'=K{row}'
        
        # N: LnA-Trend(Check) = LnZ - LnZ_HP (should be 0 initially)
        ws.cell(row=row, column=14).value = f'=K{row}-L{row}'
    
    # M: Second-order difference (interior points only)
    for row in range(7, 27):  # rows 7-26
        ws.cell(row=row, column=13).value = f'=L{row+1}-2*L{row}+L{row-1}'
    
    # P5: Objective = SUM((LnZ - LnZ_HP)^2) + 100 * SUM(second_order_diff^2)
    ws.cell(row=5, column=16).value = '=SUMPRODUCT((K6:K27-L6:L27)^2)+100*SUMPRODUCT(M7:M26,M7:M26)'
    
    # ========== Production function area (rows 36-75, years 2002-2041) ==========
    
    for i in range(40):  # 0-39 for years 2002-2041
        row = 36 + i
        year = 2002 + i
        
        # ----- E: K (capital stock, millions of Lari) -----
        if year <= 2023:
            # From PWT
            pwt_row = year - 1990 + 2
            ws.cell(row=row, column=5).value = f'=PWT!B{pwt_row}'
        else:
            # Extended: K = avg_K/Y * Y
            # avg K/Y of most recent 9 years (2015-2023) = rows 49-57
            ws.cell(row=row, column=5).value = f'=AVERAGE(D49:D57)*F{row}'
        
        # ----- F: Y (Real GDP, millions of Lari) -----
        weo_row = year - 2000 + 8
        if weo_row <= 51:  # WEO data goes to row 51 (2043)
            ws.cell(row=row, column=6).value = f'=WEO_Data!C{weo_row}*1000'
        
        # ----- D: K/Y ratio (only 2002-2023) -----
        if year <= 2023:
            ws.cell(row=row, column=4).value = f'=E{row}/F{row}'
        
        # ----- G: LnZ trend -----
        if year <= 2023:
            # Link from HP filter (L6:L27)
            hp_row = year - 2002 + 6
            ws.cell(row=row, column=7).value = f'=L{hp_row}'
        else:
            # Extend using TREND function
            ws.cell(row=row, column=7).value = f'=TREND(G36:G57,C36:C57,C{row})'
        
        # ----- H: Ystar_base = EXP(LnZ_trend) * K^alpha -----
        ws.cell(row=row, column=8).value = f'=EXP(G{row})*E{row}^$B$2'
        
        # ----- I: Investment (from Investment sheet) -----
        if 2026 <= year <= 2033:
            inv_row = year - 2026 + 2
            ws.cell(row=row, column=9).value = f'=Investment!C{inv_row}'
        else:
            ws.cell(row=row, column=9).value = 0
        
        # ----- N: Projected GDP -----
        if weo_row <= 51:
            ws.cell(row=row, column=14).value = f'=WEO_Data!C{weo_row}*1000'
    
    # ----- J: deltaK = depreciation_rate * K_with_{t-1} -----
    # First row (2002): no delta
    ws.cell(row=36, column=10).value = 0
    for row in range(37, 76):
        ws.cell(row=row, column=10).value = f'=$B$3*K{row-1}'
    
    # ----- K: K_With -----
    # K_with_2002 = K_2002 (baseline)
    ws.cell(row=36, column=11).value = f'=E36'
    for row in range(37, 76):
        # K_with_t = K_with_{t-1} * (1 - delta) + Investment_t
        ws.cell(row=row, column=11).value = f'=K{row-1}*(1-$B$3)+I{row}'
    
    # ----- L: Ystar_with = EXP(LnZ_trend) * K_with^alpha -----
    for row in range(36, 76):
        ws.cell(row=row, column=12).value = f'=EXP(G{row})*K{row}^$B$2'
    
    # ----- M: Uplift = Ystar_with - Ystar_base -----
    for row in range(36, 76):
        ws.cell(row=row, column=13).value = f'=L{row}-H{row}'
    
    # ----- O: Projected GDP Growth -----
    for row in range(37, 76):
        ws.cell(row=row, column=15).value = f'=(N{row}-N{row-1})/N{row-1}*100'
    
    # ----- P: Baseline GDP Growth -----
    for row in range(37, 76):
        ws.cell(row=row, column=16).value = f'=(H{row}-H{row-1})/H{row-1}*100'
    
    print('  Production sheet populated')


if __name__ == '__main__':
    template = sys.argv[1] if len(sys.argv) > 1 else '/root/test-supply.xlsx'
    output = sys.argv[2] if len(sys.argv) > 2 else template
    build_supply_model(template, output)
