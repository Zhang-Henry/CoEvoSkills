"""Run HP filter optimization using LibreOffice Solver."""
import subprocess
import os
import time

def create_solver_macro():
    """Create a LibreOffice Basic macro to run the HP filter Solver."""
    macro_code = '''
import uno
from com.sun.star.beans import PropertyValue

def run_hp_solver():
    """Run Solver on the Production sheet to minimize HP filter objective."""
    doc = XSCRIPTCONTEXT.getDocument()
    sheet = doc.Sheets.getByName("Production")
    
    # Get the solver
    solver = doc.createInstance("com.sun.star.sheet.Solver")
    
    # Objective cell: P5 (row 4, col 15 in 0-indexed)
    obj_cell = sheet.getCellByPosition(15, 4)  # P5
    solver.Objective = sheet.getCellRangeByName("P5").getCellAddress()
    solver.Maximize = False  # Minimize
    
    # Decision variables: L6:L27 (col 11, rows 5-26 in 0-indexed)
    from com.sun.star.sheet import CellRangeAddress
    var_range = sheet.getCellRangeByName("L6:L27")
    
    # Set up variables as individual cell addresses
    variables = []
    for row in range(5, 27):  # 0-indexed rows 5-26 = L6:L27
        cell = sheet.getCellByPosition(11, row)
        variables.append(cell.getCellAddress())
    
    solver.Variables = tuple(variables)
    solver.Constraints = ()  # No constraints
    
    # Solve
    solver.solve()
    
    # Check if solved
    if solver.Success:
        # Apply solution
        solution = solver.Solution
        for i, val in enumerate(solution):
            sheet.getCellByPosition(11, 5 + i).setValue(val)
    
    # Save
    doc.store()
    return None
'''
    return macro_code


def run_solver_via_libreoffice(xlsx_path):
    """Run the HP filter solver using LibreOffice macro."""
    abs_path = os.path.abspath(xlsx_path)
    
    # Create a Python macro file for LibreOffice
    macro_script = f'''
import uno
import sys
import os

def solve_hp_filter():
    # Connect to running LibreOffice
    localContext = uno.getComponentContext()
    resolver = localContext.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", localContext)
    
    ctx = resolver.resolve(
        "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
    smgr = ctx.ServiceManager
    desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
    
    # Open the file
    url = uno.systemPathToFileUrl("{abs_path}")
    doc = desktop.loadComponentFromURL(url, "_blank", 0, ())
    
    sheet = doc.Sheets.getByName("Production")
    
    # Use NLPSolver or CoinMP solver
    solver = doc.createInstance("com.sun.star.sheet.Solver")
    
    # Set objective: P5
    solver.Objective = sheet.getCellRangeByName("P5").getCellAddress()
    solver.Maximize = False
    
    # Set variables: L6:L27
    variables = []
    for row in range(5, 27):
        cell = sheet.getCellByPosition(11, row)
        variables.append(cell.getCellAddress())
    solver.Variables = tuple(variables)
    solver.Constraints = ()
    
    solver.solve()
    
    if solver.Success:
        solution = solver.Solution
        for i, val in enumerate(solution):
            sheet.getCellByPosition(11, 5 + i).setValue(val)
    
    doc.store()
    doc.close(True)

if __name__ == "__main__":
    solve_hp_filter()
'''
    
    # Alternative approach: use a LibreOffice Basic macro via command line
    # Create a macro that runs on document open
    basic_macro = '''Sub SolveHPFilter()
    Dim oDoc As Object
    Dim oSheet As Object
    Dim oSolver As Object
    Dim oObjective As Object
    Dim aVariables() As New com.sun.star.table.CellAddress
    Dim i As Long
    
    oDoc = ThisComponent
    oSheet = oDoc.Sheets.getByName("Production")
    
    ' Set up solver
    oSolver = createUnoService("com.sun.star.sheet.Solver")
    
    ' Objective cell P5 (col=15, row=4 in 0-based)
    Dim oObjAddr As New com.sun.star.table.CellAddress
    oObjAddr.Sheet = oSheet.getRangeAddress().Sheet
    oObjAddr.Column = 15
    oObjAddr.Row = 4
    
    oSolver.Objective = oObjAddr
    oSolver.Maximize = False
    
    ' Variables L6:L27 (col=11, rows 5-26 in 0-based)
    ReDim aVariables(21) As com.sun.star.table.CellAddress
    For i = 0 To 21
        aVariables(i).Sheet = oSheet.getRangeAddress().Sheet
        aVariables(i).Column = 11
        aVariables(i).Row = 5 + i
    Next i
    
    oSolver.Variables = aVariables()
    oSolver.Constraints = Array()
    
    oSolver.Solve()
    
    If oSolver.Success Then
        Dim aSolution() As Double
        aSolution = oSolver.Solution
        For i = 0 To 21
            oSheet.getCellByPosition(11, 5 + i).setValue(aSolution(i))
        Next i
    End If
    
    oDoc.store()
End Sub'''
    
    return basic_macro


def solve_hp_filter_numerically(xlsx_path):
    """Solve HP filter numerically using Python (scipy) and write results back."""
    import openpyxl
    import numpy as np
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Production']
    
    # Read LnZ values from column K (rows 6-27)
    lnz = []
    for row in range(6, 28):
        val = ws.cell(row=row, column=11).value
        if val is not None:
            lnz.append(float(val))
        else:
            lnz.append(0.0)
    
    wb.close()
    
    if all(v == 0 for v in lnz):
        print('Warning: LnZ values are all zero. Need to recalculate first.')
        return None
    
    lnz = np.array(lnz)
    n = len(lnz)
    lam = 100  # HP filter lambda for annual data
    
    # Solve HP filter: minimize sum((y-tau)^2) + lambda * sum((tau[t+1]-2*tau[t]+tau[t-1])^2)
    # This has a closed-form solution: tau = (I + lambda*K'K)^{-1} * y
    # where K is the second-difference matrix
    
    I = np.eye(n)
    K = np.zeros((n-2, n))
    for i in range(n-2):
        K[i, i] = 1
        K[i, i+1] = -2
        K[i, i+2] = 1
    
    # Solve
    A = I + lam * K.T @ K
    tau = np.linalg.solve(A, lnz)
    
    return tau


def apply_hp_solution(xlsx_path, tau_values):
    """Write HP filter solution back to the Excel file."""
    import openpyxl
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Production']
    
    # Write tau values to column L (rows 6-27)
    # Remove the formula and write values directly
    for i, val in enumerate(tau_values):
        ws.cell(row=6+i, column=12).value = val
    
    wb.save(xlsx_path)
    print(f'HP filter solution written to {xlsx_path}')


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else '/root/test-supply.xlsx'
    tau = solve_hp_filter_numerically(path)
    if tau is not None:
        apply_hp_solution(path, tau)
