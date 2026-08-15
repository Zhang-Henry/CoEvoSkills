"""Validate pedestrian-counting Excel output."""
import os, glob


def validate_output(output_path, video_dir, sheet_name, col_filename, col_count):
    """Check the Excel file. Returns (is_valid, issues)."""
    from openpyxl import load_workbook
    issues = []
    if not os.path.exists(output_path):
        return False, ["File not found"]
    wb = load_workbook(output_path)
    if len(wb.sheetnames) != 1:
        issues.append(f"Expected 1 sheet, found {len(wb.sheetnames)}")
    if sheet_name not in wb.sheetnames:
        issues.append(f"Sheet '{sheet_name}' missing")
        wb.close(); return False, issues
    ws = wb[sheet_name]
    if ws['A1'].value != col_filename:
        issues.append("Col A header mismatch")
    if ws['B1'].value != col_count:
        issues.append("Col B header mismatch")
    if ws['C1'].value is not None:
        issues.append("Extra column C")
    exts = ['*.mp4','*.avi','*.mov','*.mkv','*.MP4','*.AVI','*.MOV','*.MKV']
    expected = {os.path.basename(f)
                for ext in exts for f in glob.glob(os.path.join(video_dir, ext))}
    rows, r = [], 2
    while True:
        fn = ws[f'A{r}'].value
        num = ws[f'B{r}'].value
        if fn is None and num is None: break
        rows.append((fn, num)); r += 1
    if len(rows) != len(expected):
        issues.append(f"Row count {len(rows)} != video count {len(expected)}")
    found = set()
    for fn, num in rows:
        if fn not in expected:
            issues.append(f"Unexpected file '{fn}'")
        else:
            found.add(fn)
        if not isinstance(num, (int, float)) or num < 0:
            issues.append(f"Bad count for '{fn}'")
    if expected - found:
        issues.append(f"Missing: {expected - found}")
    if ws[f'A{r}'].value is not None:
        issues.append("Extra trailing rows")
    wb.close()
    ok = not issues
    print("Validation", "PASSED" if ok else "FAILED")
    for fn, num in sorted(rows):
        print(f"  {fn}: {num}")
    for iss in issues:
        print(f"  ! {iss}")
    return ok, issues
