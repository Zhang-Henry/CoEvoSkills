"""
Pedestrian counting in surveillance videos via YOLOv8 + multi-object tracking.

No fixed policy constants. All parameters are either YOLO/tracker defaults
or derived from each video's measured properties.
"""
import cv2
import os
import glob
from collections import defaultdict
import numpy as np


def get_video_files(video_dir):
    """Discover video files in a directory, sorted alphabetically."""
    extensions = ['*.mp4', '*.avi', '*.mov', '*.mkv',
                  '*.MP4', '*.AVI', '*.MOV', '*.MKV']
    files = []
    for ext in extensions:
        files.extend(glob.glob(os.path.join(video_dir, ext)))
    return sorted(files)


def get_video_metadata(video_path):
    """Return dict with fps, frame_count, width, height, duration."""
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise IOError(f"Cannot open video: {video_path}")
    fps = cap.get(cv2.CAP_PROP_FPS)
    fc = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    cap.release()
    return {'fps': fps, 'frame_count': fc, 'width': w, 'height': h,
            'duration': fc / fps if fps > 0 else 0}


def run_tracking(video_path, model, sample_every):
    """
    Run detection+tracking on a video.
    Uses the model's own default confidence and the ultralytics default tracker.
    Returns {track_id: [(frame, cx, cy, w, h, conf, cls), ...]}.
    """
    cap = cv2.VideoCapture(video_path)
    ti = defaultdict(list)
    fc = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if fc % sample_every == 0:
            results = model.track(frame, persist=True, verbose=False)
            for r in results:
                if r.boxes.id is not None:
                    for box, cid, tid, cf in zip(
                        r.boxes.xyxy, r.boxes.cls, r.boxes.id, r.boxes.conf):
                        x1, y1, x2, y2 = box.tolist()
                        ti[int(tid)].append((
                            fc, (x1+x2)/2, (y1+y2)/2,
                            x2-x1, y2-y1, float(cf), int(cid)))
        fc += 1
    cap.release()
    return dict(ti)


def filter_person_tracks(track_info, min_detections):
    """Keep COCO-person (cls 0) tracks with >= min_detections."""
    out = {}
    for tid, entries in track_info.items():
        pe = [e for e in entries if e[6] == 0]
        if len(pe) >= min_detections:
            out[tid] = pe
    return out


def exclude_cyclists(track_info, person_tracks):
    """
    Return set of person-track IDs co-located with bicycle (COCO cls 1).
    A person is excluded if they overlap a bicycle in the majority of
    their co-occurring frames.
    """
    bike = defaultdict(list)
    for entries in track_info.values():
        for e in entries:
            if e[6] == 1:
                bike[e[0]].append((e[1], e[2], e[3], e[4]))
    if not bike:
        return set()
    out = set()
    for ptid, pe in person_tracks.items():
        hits = checks = 0
        for p in pe:
            if p[0] in bike:
                checks += 1
                for bx, by, bw, bh in bike[p[0]]:
                    if np.hypot(p[1]-bx, p[2]-by) < max(p[3], p[4], bw, bh):
                        hits += 1; break
        # majority = more than half
        if checks > 0 and hits > checks // 2:
            out.add(ptid)
    return out


def count_pedestrians_in_video(video_path, model):
    """
    Count unique pedestrians in one video.
    
    The sampling interval is derived from the video's fps:
    we process roughly every (fps/10)th frame, ensuring ~10 observations
    per second regardless of the source frame rate.
    
    The minimum-detection threshold equals the number of observations
    expected in half a second at the derived sampling rate.
    """
    meta = get_video_metadata(video_path)
    fps = meta['fps']
    
    # Derive sample_every from fps: aim for ~10 processed frames/sec
    sample_every = max(1, round(fps / 10))
    # Derive min_detections: person must appear for >= 0.5 seconds
    min_detections = max(2, round(fps * 0.5 / sample_every))
    
    print(f"  {fps:.0f}fps {meta['frame_count']}f | se={sample_every} md={min_detections}")
    
    ti = run_tracking(video_path, model, sample_every)
    pt = filter_person_tracks(ti, min_detections)
    cy = exclude_cyclists(ti, pt)
    if cy:
        print(f"  Excluded {len(cy)} cyclist(s)")
    return len(pt) - len(cy)


def write_results_to_excel(results, output_path, sheet_name, col_filename, col_count):
    """
    Write (filename, count) pairs to Excel.
    Caller supplies sheet_name, col_filename, col_count from the task spec.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws['A1'] = col_filename
    ws['B1'] = col_count
    for i, (fn, cnt) in enumerate(
            sorted(results, key=lambda x: x[0]), start=2):
        ws[f'A{i}'] = fn
        ws[f'B{i}'] = int(cnt)
    wb.save(output_path)
