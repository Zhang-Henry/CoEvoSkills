import re
import os
from collections import Counter
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract


def preprocess_image(img):
    """Apply preprocessing to improve OCR quality."""
    results = []
    gray = img.convert('L')
    results.append(gray)
    
    enhancer = ImageEnhance.Contrast(gray)
    enhanced = enhancer.enhance(2.0)
    enhanced = enhanced.filter(ImageFilter.SHARPEN)
    results.append(enhanced)
    
    threshold = gray.point(lambda x: 255 if x > 140 else 0)
    results.append(threshold)
    
    w, h = img.size
    if w < 800:
        scale = 2
        scaled = gray.resize((w * scale, h * scale), Image.LANCZOS)
        enhancer2 = ImageEnhance.Contrast(scaled)
        scaled = enhancer2.enhance(1.5)
        results.append(scaled)
    
    return results


def ocr_image(image_path):
    """Run OCR with multiple preprocessing attempts."""
    img = Image.open(image_path)
    preprocessed = preprocess_image(img)
    
    texts = []
    for pp_img in preprocessed:
        for psm in [6, 4, 3]:
            try:
                text = pytesseract.image_to_string(pp_img, config=f'--psm {psm}')
                texts.append(text)
            except:
                pass
    
    try:
        text = pytesseract.image_to_string(img)
        texts.append(text)
    except:
        pass
    
    return texts


def extract_date(texts):
    """Extract date from OCR texts. Returns ISO format or None."""
    date_patterns = [
        r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})',
        r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{2})(?!\d)',
    ]
    
    date_counter = Counter()
    
    for text in texts:
        for line in text.split('\n'):
            for pattern in date_patterns:
                for match in re.finditer(pattern, line):
                    d, m, y = match.group(1), match.group(2), match.group(3)
                    try:
                        day, month, year = int(d), int(m), int(y)
                        if year < 100:
                            year += 2000
                        if 1 <= month <= 12 and 1 <= day <= 31 and 2000 <= year <= 2030:
                            date_counter[f"{year:04d}-{month:02d}-{day:02d}"] += 1
                        elif month > 12 and 1 <= day <= 12 and 1 <= month <= 31 and 2000 <= year <= 2030:
                            date_counter[f"{year:04d}-{day:02d}-{month:02d}"] += 1
                    except (ValueError, IndexError):
                        pass
    
    if date_counter:
        return date_counter.most_common(1)[0][0]
    return None


def is_header_line(line_upper):
    """Check if a line is a table header."""
    header_patterns = [
        r'\bQTY\b.*\b(TOTAL|AMOUNT)\b',
        r'\bITEM\b.*\b(TOTAL|AMOUNT)\b',
        r'\bDESC\w*\b.*\b(AMOUNT|TOTAL)\b',
        r'\bPRICE\b.*\b(AMOUNT|TOTAL)\b',
        r'\bU[/.]?P\w*\b.*\b(TOTAL|AMOUNT)\b',
    ]
    for pat in header_patterns:
        if re.search(pat, line_upper):
            return True
    if re.search(r'\b(AMOUNT|TOTAL)\s*(TAX|TX)?\s*$', line_upper) and len(line_upper.split()) >= 2:
        col_words = ['QTY', 'QUANTITY', 'ITEM', 'DESCRIPTION', 'DESC', 'PRICE', 'DISC',
                     'UNIT', 'CODE', 'NO', 'RSP', 'TAX', 'TX']
        words = line_upper.split()
        col_count = sum(1 for w in words if any(cw in w for cw in col_words))
        if col_count >= 1:
            return True
    return False


def is_excluded_line(line_upper):
    """Check if a line should be excluded."""
    include_patterns = [
        r'TOTAL.*INCL.*GST', r'TOTAL.*INCLUDING.*GST', r'TOTAL.*INC.*GST',
        r'TOTAL\s+RM\s+INCL', r'TOTAL\s+RM\s+INCLUDING',
        r'TOTAL\s+AFTER', r'TOTAL\s+SALES',
        r'GRAND\s*TOTAL', r'NETT?\s*TOTAL', r'NET\s*TOTAL',
        r'TOTAL\s*RM', r'TOTAL\s*RN',
        r'TOTAL\s+AMOUNT',
        r'TOTAL\s+DUE', r'AMOUNT\s+DUE', r'BALANCE\s+DUE',
        r'TOTAL\s+ROUNDED',
        r'TOTAL\s+INCLUSIVE',
        r'TOTAL\s*\(RM\)',
        r'TOTAL\s*:\s*RM',
        r'^\s*TOTAL\s*[:\-=]',
    ]
    for pat in include_patterns:
        if re.search(pat, line_upper):
            return False
    
    exclusion_keywords = [
        'SUBTOTAL', 'SUB TOTAL', 'SUB-TOTAL',
        'TOTAL GST', 'GST', 'TAX', 'SST',
        'DISCOUNT', 'CHANGE', 'CASH TENDERED', 'CASH',
        'ROUNDING', 'ROUND',
        'TOTAL EXCL', 'EXCLUDING',
        'TOTAL QTY', 'TOTAL QUANTITY',
        'ITEM COUNT', 'TOTAL SAVING', 'AVERAGE',
    ]
    for kw in exclusion_keywords:
        if kw in line_upper:
            return True
    return False


def find_amount_in_line(line):
    """Find monetary amount in a line. Returns float or None."""
    cleaned = re.sub(r'[Rr][MmNn]\s*', ' ', line)
    cleaned = re.sub(r'MYR\s*', ' ', cleaned, flags=re.IGNORECASE)
    
    # Fix OCR artifacts: spaces around decimal points
    cleaned = re.sub(r'(\d+)\s*\.\s*(\d+)', r'\1.\2', cleaned)
    
    # Fix OCR artifacts: colon instead of period
    cleaned = re.sub(r'(\d+):(\d{2})(?!\d)', r'\1.\2', cleaned)
    
    # First try: numbers with decimal point
    matches = re.findall(r'([\d,]+\.\d{1,2})(?!\d)', cleaned)
    if matches:
        amt_str = matches[-1].replace(',', '')
        try:
            return float(amt_str)
        except ValueError:
            pass
    
    # Second try: space as decimal separator
    space_matches = re.findall(r'(\d+)\s(\d{2})(?!\d)', cleaned)
    if space_matches:
        whole, frac = space_matches[-1]
        try:
            return float(f"{whole}.{frac}")
        except ValueError:
            pass
    
    return None


def find_amount_next_line(lines, idx):
    """Fallback: look at next line for amount."""
    if idx + 1 < len(lines):
        next_line = lines[idx + 1].strip()
        if next_line:
            return find_amount_in_line(next_line)
    return None


def has_rounding_before(lines, idx):
    """Check if there's a Rounding line within 1-3 lines before idx."""
    for j in range(max(0, idx - 3), idx):
        lu = lines[j].upper().strip()
        if re.search(r'ROUND', lu):
            return True
    return False


def extract_total_from_text(text):
    """Extract total from a single OCR text.
    Returns (priority, amount) or None."""
    lines = text.split('\n')
    
    keyword_groups = [
        (1, [r'GRAND\s*TOTA']),
        (2, [r'TOTAL\s+R[MN]\s+(?!INCL|INCLUDING)',
             r'TOTAL\s*:\s*R[MN]\s+(?!INCL|INCLUDING)']),
        (3, [r'TOTAL\s+R[MN]\s+INCL', r'TOTAL\s+R[MN]\s+INCLUDING',
             r'TOTAL\s*:\s*R[MN]\s+INCL', r'TOTAL\s*:\s*R[MN]\s+INCLUDING']),
        (4, [r'TOTAL\s+AMOUNT']),
        (5, [r'TOTAL\s+DUE', r'AMOUNT\s+DUE', r'BALANCE\s+DUE',
             r'NETT?\s+TOTAL', r'NET\s+TOTAL',
             r'TOTAL\s+AFTER\s+ADJ', r'TOTAL\s+AFTER\s+ROUND',
             r'TOTAL\s+ROUNDED',
             r'TOTAL\s+SALES\s+INCL',
             r'TOTAL\s+INCL',
             r'TOTAL\s+INCLUSIVE',
             r'TOTAL\s*\(RM\)',
             r'(?:^|[^A-Z])TOTAL(?:\s*[:\-=]\s*|\s+)(?!QTY|QUANTITY|ITEM|SAVING|GST|TAX|R[MN])',
             ]),
    ]
    
    for priority, kw_patterns in keyword_groups:
        # Collect ALL matches at this priority
        matches_at_priority = []
        
        for i, line in enumerate(lines):
            line_upper = line.upper().strip()
            if not line_upper:
                continue
            if is_header_line(line_upper):
                continue
            if is_excluded_line(line_upper):
                continue
            
            for kw_pattern in kw_patterns:
                if re.search(kw_pattern, line_upper):
                    amt = find_amount_in_line(line)
                    if amt is not None and amt > 0:
                        is_post_rounding = has_rounding_before(lines, i)
                        matches_at_priority.append((amt, is_post_rounding, i))
                    else:
                        amt = find_amount_next_line(lines, i)
                        if amt is not None and amt > 0:
                            is_post_rounding = has_rounding_before(lines, i)
                            matches_at_priority.append((amt, is_post_rounding, i))
                    break
        
        if matches_at_priority:
            # Prefer post-rounding matches
            post_rounding = [m for m in matches_at_priority if m[1]]
            if post_rounding:
                return (priority, post_rounding[0][0])
            else:
                return (priority, matches_at_priority[0][0])
    
    return None


def extract_total_amount(texts):
    """Extract total from all OCR texts using voting."""
    results_by_priority = {}
    
    for text in texts:
        result = extract_total_from_text(text)
        if result is not None:
            priority, amount = result
            if priority not in results_by_priority:
                results_by_priority[priority] = []
            results_by_priority[priority].append(amount)
    
    if not results_by_priority:
        return None
    
    best_priority = min(results_by_priority.keys())
    amounts = results_by_priority[best_priority]
    
    amt_counter = Counter()
    for a in amounts:
        amt_counter[round(a, 2)] += 1
    
    best_amount = amt_counter.most_common(1)[0][0]
    return f"{best_amount:.2f}"


def process_receipt(image_path):
    """Process a single receipt image."""
    texts = ocr_image(image_path)
    date = extract_date(texts)
    total = extract_total_amount(texts)
    return date, total


def process_all_receipts(image_dir, output_path):
    """Process all receipt images and write to Excel."""
    import openpyxl
    
    image_files = sorted([f for f in os.listdir(image_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
    
    results = []
    for fn in image_files:
        filepath = os.path.join(image_dir, fn)
        print(f"Processing {fn}...")
        date, total = process_receipt(filepath)
        print(f"  Date: {date}, Total: {total}")
        results.append((fn, date, total))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(['filename', 'date', 'total_amount'])
    
    for fn, date, total in results:
        ws.append([fn, date, total])
    
    wb.save(output_path)
    print(f"\nResults saved to {output_path}")
    return results


def validate_output(output_path):
    """Validate the output Excel file."""
    import openpyxl
    
    wb = openpyxl.load_workbook(output_path)
    assert 'results' in wb.sheetnames
    assert len(wb.sheetnames) == 1
    
    ws = wb['results']
    headers = [ws.cell(1, c).value for c in range(1, 4)]
    assert headers == ['filename', 'date', 'total_amount']
    
    for row in range(2, ws.max_row + 1):
        fn = ws.cell(row, 1).value
        date = ws.cell(row, 2).value
        total = ws.cell(row, 3).value
        print(f"  {fn}: date={date}, total={total}")
        if date is not None:
            assert re.match(r'^\d{4}-\d{2}-\d{2}$', str(date))
        if total is not None:
            assert re.match(r'^\d+\.\d{2}$', str(total))
    
    print("Validation passed!")
    return True
