"""
Utility functions for proteomics workbook task.
Fills expression lookup formulas, statistics formulas, and fold change formulas
in the Task sheet of a protein_expression.xlsx workbook.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import subprocess
import os
import shutil
import tempfile


def discover_task_layout(ws_task):
    """
    Discover the layout of the Task sheet at runtime.
    Returns dict with all relevant ranges and mappings.
    """
    layout = {}
    
    # Find sample names in row 10 (columns C onwards)
    sample_names = []
    sample_cols = []
    for col in range(3, ws_task.max_column + 1):
        v = ws_task.cell(row=10, column=col).value
        if v is not None:
            sample_names.append(str(v))
            sample_cols.append(col)
    layout['sample_names'] = sample_names
    layout['sample_cols'] = sample_cols
    layout['num_samples'] = len(sample_names)
    
    # Find group assignments in row 9
    control_cols = []
    treated_cols = []
    for col in sample_cols:
        group = ws_task.cell(row=9, column=col).value
        if group and str(group).strip().lower() == 'control':
            control_cols.append(col)
        elif group and str(group).strip().lower() == 'treated':
            treated_cols.append(col)
    layout['control_cols'] = control_cols
    layout['treated_cols'] = treated_cols
    
    # Find protein IDs in column A, rows 11-20
    protein_ids = []
    gene_symbols = []
    protein_rows = []
    for row in range(11, 21):
        pid = ws_task.cell(row=row, column=1).value
        gsym = ws_task.cell(row=row, column=2).value
        if pid is not None:
            protein_ids.append(str(pid))
            gene_symbols.append(str(gsym) if gsym else '')
            protein_rows.append(row)
    layout['protein_ids'] = protein_ids
    layout['gene_symbols'] = gene_symbols
    layout['protein_rows'] = protein_rows
    layout['num_proteins'] = len(protein_ids)
    
    # Statistics rows
    layout['control_mean_row'] = 24
    layout['control_std_row'] = 25
    layout['treated_mean_row'] = 26
    layout['treated_std_row'] = 27
    # Statistics columns B-K (one per protein)
    layout['stats_cols'] = list(range(2, 2 + len(protein_ids)))
    
    # Fold change section
    layout['fc_start_row'] = 32
    layout['fc_protein_col'] = 1  # A
    layout['fc_gene_col'] = 2    # B
    layout['fc_fold_change_col'] = 3  # C
    layout['fc_log2fc_col'] = 4  # D
    
    return layout


def discover_data_layout(ws_data):
    """
    Discover the layout of the Data sheet.
    Returns dict with protein row index and sample column index.
    """
    layout = {}
    
    # Build protein ID -> row mapping
    protein_rows = {}
    for row in range(2, ws_data.max_row + 1):
        pid = ws_data.cell(row=row, column=1).value
        if pid:
            protein_rows[str(pid)] = row
    layout['protein_rows'] = protein_rows
    
    # Build sample name -> column mapping
    sample_cols = {}
    for col in range(2, ws_data.max_column + 1):
        sname = ws_data.cell(row=1, column=col).value
        if sname:
            sample_cols[str(sname)] = col
    layout['sample_cols'] = sample_cols
    
    layout['sheet_name'] = ws_data.title
    layout['max_row'] = ws_data.max_row
    layout['max_col'] = ws_data.max_column
    
    return layout


def write_expression_formulas(ws_task, task_layout, data_layout):
    """
    Step 1: Write INDEX-MATCH formulas in C11:L20 to look up expression values
    from the Data sheet.
    """
    data_sheet = data_layout['sheet_name']
    data_max_row = data_layout['max_row']
    data_max_col = data_layout['max_col']
    last_col_letter = get_column_letter(data_max_col)
    
    for i, prow in enumerate(task_layout['protein_rows']):
        protein_cell = f'$A${prow}'
        
        for j, scol in enumerate(task_layout['sample_cols']):
            sample_cell = f'{get_column_letter(scol)}$10'
            
            formula = (
                f"=INDEX({data_sheet}!$A$1:${last_col_letter}${data_max_row},"
                f"MATCH({protein_cell},{data_sheet}!$A$1:$A${data_max_row},0),"
                f"MATCH({sample_cell},{data_sheet}!$1:$1,0))"
            )
            
            ws_task.cell(row=prow, column=scol).value = formula


def write_statistics_formulas(ws_task, task_layout):
    """
    Step 2: Write AVERAGE and STDEV formulas for control and treated groups.
    Statistics go in rows 24-27, columns B-K (one column per protein).
    """
    for i, prow in enumerate(task_layout['protein_rows']):
        stats_col = task_layout['stats_cols'][i]
        
        control_refs = [f'{get_column_letter(ccol)}{prow}' for ccol in task_layout['control_cols']]
        treated_refs = [f'{get_column_letter(tcol)}{prow}' for tcol in task_layout['treated_cols']]
        
        control_range = ','.join(control_refs)
        treated_range = ','.join(treated_refs)
        
        ws_task.cell(row=task_layout['control_mean_row'], column=stats_col).value = \
            f'=AVERAGE({control_range})'
        ws_task.cell(row=task_layout['control_std_row'], column=stats_col).value = \
            f'=STDEV({control_range})'
        ws_task.cell(row=task_layout['treated_mean_row'], column=stats_col).value = \
            f'=AVERAGE({treated_range})'
        ws_task.cell(row=task_layout['treated_std_row'], column=stats_col).value = \
            f'=STDEV({treated_range})'


def write_fold_change_formulas(ws_task, task_layout):
    """
    Step 3: Write fold change formulas in rows 32-41, columns A-D.
    - A: Protein_ID (reference to A11:A20)
    - B: Gene_Symbol (reference to B11:B20)
    - C: Fold Change = 2^(Log2 FC)
    - D: Log2 FC = Treated Mean - Control Mean
    """
    for i in range(task_layout['num_proteins']):
        fc_row = task_layout['fc_start_row'] + i
        prow = task_layout['protein_rows'][i]
        stats_col = task_layout['stats_cols'][i]
        stats_col_letter = get_column_letter(stats_col)
        
        # A: Protein_ID
        ws_task.cell(row=fc_row, column=task_layout['fc_protein_col']).value = f'=A{prow}'
        
        # B: Gene_Symbol
        ws_task.cell(row=fc_row, column=task_layout['fc_gene_col']).value = f'=B{prow}'
        
        # D: Log2 FC = Treated Mean - Control Mean
        treated_mean_ref = f'{stats_col_letter}{task_layout["treated_mean_row"]}'
        control_mean_ref = f'{stats_col_letter}{task_layout["control_mean_row"]}'
        ws_task.cell(row=fc_row, column=task_layout['fc_log2fc_col']).value = \
            f'={treated_mean_ref}-{control_mean_ref}'
        
        # C: Fold Change = 2^(Log2 FC)
        log2fc_ref = f'{get_column_letter(task_layout["fc_log2fc_col"])}{fc_row}'
        ws_task.cell(row=fc_row, column=task_layout['fc_fold_change_col']).value = \
            f'=POWER(2,{log2fc_ref})'


def recalculate_with_libreoffice(filepath):
    """
    Use LibreOffice to recalculate all formulas and save.
    Copies to temp dir, converts, and copies back to avoid in-place issues.
    """
    abs_path = os.path.abspath(filepath)
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Copy input to temp
        tmp_input = os.path.join(tmpdir, 'input.xlsx')
        tmp_output_dir = os.path.join(tmpdir, 'output')
        os.makedirs(tmp_output_dir)
        shutil.copy2(abs_path, tmp_input)
        
        # Convert with LibreOffice
        cmd = [
            'libreoffice', '--headless', '--calc', '--norestore',
            f'-env:UserInstallation=file://{tmpdir}/lo_user',
            '--convert-to', 'xlsx',
            '--outdir', tmp_output_dir,
            tmp_input
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode != 0:
            print(f'LibreOffice error: {result.stderr}')
            return False
        
        # Copy result back
        tmp_output = os.path.join(tmp_output_dir, 'input.xlsx')
        if os.path.exists(tmp_output):
            shutil.copy2(tmp_output, abs_path)
            return True
        else:
            print(f'LibreOffice output not found at {tmp_output}')
            print(f'Files in output dir: {os.listdir(tmp_output_dir)}')
            return False


def validate_workbook(filepath):
    """
    Validate the workbook by opening in data_only mode and checking
    that all expected cells have numeric values (no errors).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ts = wb['Task']
    
    errors = []
    error_values = {'#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#NUM!', '#N/A', '#NULL!'}
    
    # Check expression values C11:L20
    for row in range(11, 21):
        for col in range(3, 13):
            v = ts.cell(row=row, column=col).value
            cell_ref = f'{get_column_letter(col)}{row}'
            if v is None:
                errors.append(f'{cell_ref}: empty (expected numeric)')
            elif isinstance(v, str) and v in error_values:
                errors.append(f'{cell_ref}: formula error {v}')
            elif not isinstance(v, (int, float)):
                errors.append(f'{cell_ref}: not numeric ({type(v).__name__}: {v})')
    
    # Check statistics B24:K27
    for row in range(24, 28):
        for col in range(2, 12):
            v = ts.cell(row=row, column=col).value
            cell_ref = f'{get_column_letter(col)}{row}'
            if v is None:
                errors.append(f'{cell_ref}: empty (expected numeric)')
            elif isinstance(v, str) and v in error_values:
                errors.append(f'{cell_ref}: formula error {v}')
            elif not isinstance(v, (int, float)):
                errors.append(f'{cell_ref}: not numeric ({type(v).__name__}: {v})')
    
    # Check fold change C32:D41
    for row in range(32, 42):
        for col in range(3, 5):
            v = ts.cell(row=row, column=col).value
            cell_ref = f'{get_column_letter(col)}{row}'
            if v is None:
                errors.append(f'{cell_ref}: empty (expected numeric)')
            elif isinstance(v, str) and v in error_values:
                errors.append(f'{cell_ref}: formula error {v}')
            elif not isinstance(v, (int, float)):
                errors.append(f'{cell_ref}: not numeric ({type(v).__name__}: {v})')
    
    # Check fold change protein IDs A32:B41
    for row in range(32, 42):
        for col in range(1, 3):
            v = ts.cell(row=row, column=col).value
            cell_ref = f'{get_column_letter(col)}{row}'
            if v is None:
                errors.append(f'{cell_ref}: empty (expected protein/gene)')
    
    wb.close()
    
    if errors:
        print(f'Validation FAILED with {len(errors)} errors:')
        for e in errors[:20]:
            print(f'  {e}')
        return False
    else:
        print('Validation PASSED: all cells have numeric values')
        return True


def run_end_to_end(input_path, output_path=None):
    """
    End-to-end entry point: reads the workbook, writes all formulas,
    recalculates with LibreOffice, and validates.
    
    Args:
        input_path: Path to protein_expression.xlsx
        output_path: Path to save result (defaults to overwriting input)
    """
    if output_path is None:
        output_path = input_path
    
    # Load workbook preserving formatting
    wb = openpyxl.load_workbook(input_path)
    ws_task = wb['Task']
    ws_data = wb['Data']
    
    # Discover layouts
    task_layout = discover_task_layout(ws_task)
    data_layout = discover_data_layout(ws_data)
    
    print(f'Found {task_layout["num_proteins"]} proteins, {task_layout["num_samples"]} samples')
    print(f'Control columns: {[get_column_letter(c) for c in task_layout["control_cols"]]}')
    print(f'Treated columns: {[get_column_letter(c) for c in task_layout["treated_cols"]]}')
    
    # Step 1: Expression lookup formulas
    print('\nStep 1: Writing expression lookup formulas...')
    write_expression_formulas(ws_task, task_layout, data_layout)
    
    # Step 2: Statistics formulas
    print('Step 2: Writing statistics formulas...')
    write_statistics_formulas(ws_task, task_layout)
    
    # Step 3: Fold change formulas
    print('Step 3: Writing fold change formulas...')
    write_fold_change_formulas(ws_task, task_layout)
    
    # Save
    print(f'\nSaving to {output_path}...')
    wb.save(output_path)
    wb.close()
    
    # Recalculate with LibreOffice
    print('Recalculating with LibreOffice...')
    success = recalculate_with_libreoffice(output_path)
    if not success:
        print('WARNING: LibreOffice recalculation may have failed')
    
    # Validate
    print('\nValidating...')
    return validate_workbook(output_path)

