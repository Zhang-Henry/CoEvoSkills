"""
Utility functions for demand-side macroeconomic shock analysis.
Builds an Excel workbook with WEO data, Supply-Use Tables, and
investment shock scenarios for a small open economy.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os


def fetch_imf_weo_data(country_code='GEO'):
    """Fetch IMF WEO data from the datamapper API."""
    import urllib.request
    import json
    results = {}
    for ind in ['NGDP_RPCH', 'NGDPD', 'PCPIPCH']:
        try:
            url = f'https://www.imf.org/external/datamapper/api/v1/{ind}/{country_code}'
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64)',
                'Accept': 'application/json'
            })
            resp = urllib.request.urlopen(req, timeout=15)
            data = json.loads(resp.read().decode('utf-8'))
            if 'values' in data and ind in data['values']:
                geo = data['values'][ind].get(country_code, {})
                results[ind] = {int(k): v for k, v in geo.items()}
        except Exception as e:
            print(f'Warning: Could not fetch {ind}: {e}')
    return results


def download_file(url, output_path):
    """Download a file from URL."""
    import urllib.request
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    resp = urllib.request.urlopen(req, timeout=30)
    with open(output_path, 'wb') as f:
        f.write(resp.read())
    return output_path


def download_geostat_gdp_files(output_dir='/tmp'):
    """Download GDP data files from geostat.ge."""
    files = {
        'gdp_current': 'https://geostat.ge/media/81052/03_GDP-at-Current-Prices.xlsx',
        'gdp_constant': 'https://geostat.ge/media/81054/04_GDP-at-constant-prices.xlsx',
        'gdp_growth': 'https://geostat.ge/media/80933/06_Real-GDP-Growth.xlsx',
    }
    paths = {}
    for key, url in files.items():
        path = os.path.join(output_dir, f'{key}.xlsx')
        try:
            download_file(url, path)
            paths[key] = path
        except Exception as e:
            print(f'Warning: Could not download {key}: {e}')
    return paths


def extract_geostat_annual_gdp(gdp_file_path, row_label_search):
    """Extract annual GDP values from a geostat file. Returns {year: value}."""
    wb = openpyxl.load_workbook(gdp_file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    target_row = None
    for r in range(1, ws.max_row + 1):
        b_val = ws.cell(r, 2).value
        if b_val and row_label_search.lower() in str(b_val).lower():
            target_row = r
            break
    if target_row is None:
        return {}
    result = {}
    for c in range(3, ws.max_column + 1):
        h = ws.cell(2, c).value
        v = ws.cell(target_row, c).value
        if h and v is not None:
            h_str = str(h).strip()
            if h_str.isdigit() and len(h_str) == 4:
                result[int(h_str)] = float(v)
    return result


def find_latest_sut_url():
    """Find the latest SUT file URL from geostat.ge."""
    import urllib.request
    import re
    try:
        req = urllib.request.Request(
            'https://www.geostat.ge/en/modules/categories/632',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        resp = urllib.request.urlopen(req, timeout=15)
        html = resp.read().decode('utf-8', errors='replace')
        links = re.findall(r'href=["\']([^"\']*SUT[^"\']*\.xlsx)["\']', html, re.IGNORECASE)
        if links:
            return links[-1]  # Latest
    except Exception as e:
        print(f'Warning: Could not find SUT URL: {e}')
    return None


def copy_sut_sheet(source_wb, source_sheet_name, target_wb, target_sheet_name):
    """Copy a sheet from source workbook to target workbook."""
    src_ws = source_wb[source_sheet_name]
    if target_sheet_name in target_wb.sheetnames:
        del target_wb[target_sheet_name]
    tgt_ws = target_wb.create_sheet(target_sheet_name)
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                 min_col=1, max_col=src_ws.max_column):
        for cell in row:
            new_cell = tgt_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for col_idx in range(1, src_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in src_ws.column_dimensions:
            tgt_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
    for merge in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merge))
    return tgt_ws


def bell_shape_allocation(n_years):
    """Generate bell-shaped allocation profile summing to 1.
    Uses a standard ramp-up/ramp-down pattern.
    For 8 years: [5, 10, 15, 20, 20, 15, 10, 5] / 100
    """
    if n_years <= 0:
        return []
    if n_years == 1:
        return [1.0]
    if n_years == 8:
        return [0.05, 0.10, 0.15, 0.20, 0.20, 0.15, 0.10, 0.05]
    # Generic bell shape for other lengths
    mid = n_years / 2.0
    raw = []
    for i in range(n_years):
        dist = abs(i + 0.5 - mid)
        val = mid - dist
        raw.append(val)
    total = sum(raw)
    return [v / total for v in raw]


def build_weo_data_sheet(wb, weo_data, geostat_data, start_year=2020, end_year=2043):
    """Build the WEO_Data sheet.
    
    Columns:
    A: Year
    B: Real GDP Growth (%)
    C: Nominal GDP (mil GEL) 
    D: Real GDP (mil GEL)
    E: GDP Deflator (Index, base year implied)
    F: GDP Deflator Growth (%)
    """
    if 'WEO_Data' in wb.sheetnames:
        ws = wb['WEO_Data']
    else:
        ws = wb.create_sheet('WEO_Data', 0)
    
    # Headers
    ws.cell(1, 1, 'Year')
    ws.cell(1, 2, 'Real GDP Growth (%)')
    ws.cell(1, 3, 'Nominal GDP (mil GEL)')
    ws.cell(1, 4, 'Real GDP (mil GEL)')
    ws.cell(1, 5, 'GDP Deflator (Index)')
    ws.cell(1, 6, 'GDP Deflator Growth (%)')
    
    real_growth_weo = weo_data.get('NGDP_RPCH', {})
    nominal_gdp_geo = geostat_data.get('nominal_gdp', {})
    real_gdp_geo = geostat_data.get('real_gdp', {})
    deflator_growth_geo = geostat_data.get('deflator_growth', {})
    
    # Compute deflator index: nominal/real * 100
    deflator_index = {}
    for y in sorted(set(nominal_gdp_geo.keys()) & set(real_gdp_geo.keys())):
        if real_gdp_geo[y] != 0:
            deflator_index[y] = (nominal_gdp_geo[y] / real_gdp_geo[y]) * 100
    
    observed_years = sorted(nominal_gdp_geo.keys())
    last_obs = max(observed_years) if observed_years else 2024
    
    # Average deflator growth for recent 4 years (as percentage points)
    # deflator_growth_geo values are index (prev year=100), so growth = value - 100
    deflator_growths = []
    for y in range(last_obs - 3, last_obs + 1):
        if y in deflator_growth_geo:
            deflator_growths.append(deflator_growth_geo[y] - 100.0)
    avg_defl_growth = sum(deflator_growths) / len(deflator_growths) if deflator_growths else 3.0
    
    # Find row for 2027 to reference its growth rate
    row_2027 = 2 + (2027 - start_year)
    
    row = 2
    for y in range(start_year, end_year + 1):
        ws.cell(row, 1, y)
        
        # Col B: Real GDP Growth
        if y in real_growth_weo:
            ws.cell(row, 2, real_growth_weo[y])
        elif y > 2027:
            ws.cell(row, 2, f'=B{row_2027}')
        
        # Col D: Real GDP
        if y in real_gdp_geo:
            ws.cell(row, 4, real_gdp_geo[y])
        elif y > last_obs:
            ws.cell(row, 4, f'=D{row-1}*(1+B{row}/100)')
        
        # Col F: Deflator Growth
        if y in deflator_growth_geo:
            ws.cell(row, 6, deflator_growth_geo[y] - 100.0)
        elif y > last_obs:
            # Use average of recent 4 years as fixed anchor
            # Use AVERAGE formula for recent 4 years
            last_obs_row_weo = 2 + (last_obs - start_year)
            avg_start_row = last_obs_row_weo - 3
            ws.cell(row, 6, f"=AVERAGE(F${avg_start_row}:F${last_obs_row_weo})")
        
        # Col E: Deflator Index
        if y in deflator_index:
            ws.cell(row, 5, deflator_index[y])
        elif y > last_obs:
            ws.cell(row, 5, f'=E{row-1}*(1+F{row}/100)')
        
        # Col C: Nominal GDP
        if y in nominal_gdp_geo:
            ws.cell(row, 3, nominal_gdp_geo[y])
        elif y > last_obs:
            ws.cell(row, 3, f'=D{row}*E{row}/100')
        
        row += 1
    
    return ws


def build_sut_calc_sheet(wb, supply_sheet_name, use_sheet_name):
    """Build the SUT Calc sheet.
    
    Calculates import content share for construction using SUT data.
    The task requires columns C-H and import content share in C46.
    
    Layout:
    Row 1: Title
    Row 2: Headers
    Rows 3-40: Product data (38 products)
    Row 41: blank
    Row 42: Total
    Row 44: blank  
    Row 46: Import content share result in C46
    
    Columns:
    A: Product code
    B: Description  
    C: Intermediate consumption by Construction (from USE)
    D: Total Supply (from SUPPLY)
    E: Imports (from SUPPLY)
    F: Import share = E/D
    G: Weighted import = C*F
    H: Share of construction input = C/total_C
    """
    if 'SUT Calc' in wb.sheetnames:
        ws = wb['SUT Calc']
    else:
        ws = wb.create_sheet('SUT Calc')
    
    supply_ws = wb[supply_sheet_name]
    use_ws = wb[use_sheet_name]
    
    # Find construction column in USE (NACE 41-43)
    constr_col = None
    for c in range(1, use_ws.max_column + 1):
        h = use_ws.cell(2, c).value
        if h and '41' in str(h):
            constr_col = c
            break
    constr_col_letter = get_column_letter(constr_col) if constr_col else 'T'
    
    # Find import column (P7) in SUPPLY
    import_col = None
    for c in range(1, supply_ws.max_column + 1):
        h = supply_ws.cell(2, c).value
        if h and str(h).strip() == 'P7':
            import_col = c
            break
    import_col_letter = get_column_letter(import_col) if import_col else 'AR'
    
    # Find total supply column (TotPP) in SUPPLY
    totsup_col = None
    for c in range(1, supply_ws.max_column + 1):
        h = supply_ws.cell(2, c).value
        if h and str(h).strip() in ('TotPP', 'Tot'):
            totsup_col = c
            break
    totsup_letter = get_column_letter(totsup_col) if totsup_col else 'AS'
    
    # Headers
    ws.cell(1, 1, 'SUT Calculation for Import Content Share')
    ws.cell(2, 1, 'Product Code')
    ws.cell(2, 2, 'Description')
    ws.cell(2, 3, 'Construction Intermediate Use')
    ws.cell(2, 4, 'Total Supply')
    ws.cell(2, 5, 'Imports')
    ws.cell(2, 6, 'Import Share')
    ws.cell(2, 7, 'Weighted Import')
    ws.cell(2, 8, 'Share of Constr. Input')
    
    # SUT data rows: 4 to 41 (38 products)
    sut_first = 4
    sut_last = 41
    
    # Map to calc rows 3-40
    calc_first = 3
    total_calc_row = calc_first + (sut_last - sut_first + 1)  # row 41
    
    for i, sut_r in enumerate(range(sut_first, sut_last + 1)):
        calc_r = calc_first + i
        
        # A: product code
        ws.cell(calc_r, 1, f"='{use_sheet_name}'!A{sut_r}")
        # B: description
        ws.cell(calc_r, 2, f"='{use_sheet_name}'!B{sut_r}")
        # C: construction intermediate use
        ws.cell(calc_r, 3, f"='{use_sheet_name}'!{constr_col_letter}{sut_r}")
        # D: total supply
        ws.cell(calc_r, 4, f"='{supply_sheet_name}'!{totsup_letter}{sut_r}")
        # E: imports
        ws.cell(calc_r, 5, f"='{supply_sheet_name}'!{import_col_letter}{sut_r}")
        # F: import share = imports / total supply
        ws.cell(calc_r, 6, f'=IF(D{calc_r}=0,0,E{calc_r}/D{calc_r})')
        # G: weighted import = construction use * import share
        ws.cell(calc_r, 7, f'=C{calc_r}*F{calc_r}')
        # H: share of construction input
        ws.cell(calc_r, 8, f'=IF(C${total_calc_row}=0,0,C{calc_r}/C${total_calc_row})')
    
    # Total row (row 41)
    ws.cell(total_calc_row, 1, 'Total')
    ws.cell(total_calc_row, 2, 'Total')
    ws.cell(total_calc_row, 3, f'=SUM(C{calc_first}:C{total_calc_row-1})')
    ws.cell(total_calc_row, 4, f'=SUM(D{calc_first}:D{total_calc_row-1})')
    ws.cell(total_calc_row, 5, f'=SUM(E{calc_first}:E{total_calc_row-1})')
    ws.cell(total_calc_row, 7, f'=SUM(G{calc_first}:G{total_calc_row-1})')
    
    # Import content share in C46
    ws.cell(44, 1, 'Estimated Import Content Share')
    ws.cell(46, 1, 'Import Content Share')
    ws.cell(46, 3, f'=G{total_calc_row}/C{total_calc_row}')
    
    return ws, "'SUT Calc'!C46"


def build_na_scenario(wb, start_row, scenario_num, import_content_cell,
                      weo_sheet='WEO_Data',
                      total_inv_usd=6.5, exchange_rate=2.746,
                      demand_multiplier=0.8, project_years=8,
                      import_content_override=None,
                      multiplier_override=None):
    """Build one scenario table on the NA sheet.
    
    Template layout (relative to start_row):
    Row 0: Headers line 1 (B=Year, C=Real GDP, D=Project Allocation, ...)
    Row 1: Headers line 2 (G=I-M, H=Multiplier, I=% of GDP)
    Rows 2-24: Data (years 2020-2042, 23 years)
    Row 27: C=Assumptions
    Row 29: C=total investment, D=value
    Row 30: C=import content share, D=value
    Row 31: C=demand multiplier, D=value
    Row 32: C=Project allocation, D=bell shape
    """
    ws = wb['NA']
    
    # Row offsets from template: row 1 = headers, row 3 = first data year
    # Assumptions at row 28, values at rows 30-33
    h1 = start_row       # header row 1
    h2 = start_row + 1   # header row 2
    first_data = start_row + 2  # first data row (year 2020)
    last_data = start_row + 24  # last data row (year 2042)
    
    assumptions_label = start_row + 27  # 'Assumptions'
    inv_row = start_row + 29    # total investment
    imp_row = start_row + 30    # import content share
    mult_row = start_row + 31   # demand multiplier
    alloc_row = start_row + 32  # Project allocation
    
    # Headers
    if scenario_num == 1:
        ws.cell(h1, 2, 'Year')
    else:
        scenario_names = {
            2: 'Scenario 2: Multiplier = 1',
            3: 'Scenario 3: Import Content = 0.5'
        }
        ws.cell(h1, 2, scenario_names.get(scenario_num, f'Scenario {scenario_num}'))
    
    ws.cell(h1, 3, 'Real GDP')
    ws.cell(h1, 4, 'Project Allocation (%)')
    ws.cell(h1, 5, 'Project Investment')
    ws.cell(h1, 6, 'Imports')
    ws.cell(h1, 7, 'GDP impact')
    ws.cell(h1, 10, 'Baseline GDP Growth')
    ws.cell(h1, 11, 'Projected Growth')
    ws.cell(h2, 7, 'I-M')
    ws.cell(h2, 8, 'Multiplier')
    ws.cell(h2, 9, '% of GDP')
    
    # Assumptions
    ws.cell(assumptions_label, 3, 'Assumptions')
    ws.cell(inv_row, 3, 'total investment')
    ws.cell(imp_row, 3, 'import content share')
    ws.cell(mult_row, 3, 'demand multiplier')
    ws.cell(alloc_row, 3, 'Project allocation')
    
    # Total investment in mil GEL
    total_inv_gel_mil = total_inv_usd * exchange_rate * 1000
    ws.cell(inv_row, 4, total_inv_gel_mil)
    
    # Import content share
    if import_content_override is not None:
        ws.cell(imp_row, 4, import_content_override)
    else:
        ws.cell(imp_row, 4, f'={import_content_cell}')
    
    # Demand multiplier
    actual_mult = multiplier_override if multiplier_override is not None else demand_multiplier
    ws.cell(mult_row, 4, actual_mult)
    
    # Project allocation label
    ws.cell(alloc_row, 4, 'bell shape')
    
    # Bell-shaped allocation
    allocation = bell_shape_allocation(project_years)
    
    # Find WEO data year-to-row mapping
    weo_ws = wb[weo_sheet]
    weo_map = {}  # year -> weo row
    for r in range(2, weo_ws.max_row + 1):
        y = weo_ws.cell(r, 1).value
        if y:
            weo_map[int(y)] = r
    
    # Fill data rows
    project_start = 2026
    project_end = project_start + project_years - 1
    
    for i in range(23):  # 23 years: 2020-2042
        data_row = first_data + i
        year = 2020 + i
        
        # B: Year
        ws.cell(data_row, 2, year)
        
        # C: Real GDP (link from WEO_Data column D)
        if year in weo_map:
            ws.cell(data_row, 3, f"='{weo_sheet}'!D{weo_map[year]}")
        
        # D: Project Allocation (%)
        if project_start <= year <= project_end:
            alloc_idx = year - project_start
            ws.cell(data_row, 4, allocation[alloc_idx])
        else:
            ws.cell(data_row, 4, 0)
        
        # E: Project Investment (nominal) = total_inv * allocation
        # But we need to use GDP deflator as investment deflator
        # The total investment is in current (nominal) terms
        # To get real investment: nominal_inv / (deflator/100)
        # Actually, the investment is given as a fixed USD amount
        # Convert to GEL at fixed rate, then the allocation gives nominal spending per year
        # For real comparison with real GDP, deflate using GDP deflator
        # E = total_investment * allocation_pct (this is nominal)
        weo_row = 2 + i
        ws.cell(data_row, 5, f"=D${inv_row}*D{data_row}*100/'{weo_sheet}'!E{weo_row}")
        
        # F: Imports = investment * import_content_share
        ws.cell(data_row, 6, f'=E{data_row}*D${imp_row}')
        
        # G: GDP impact I-M = Investment - Imports
        ws.cell(data_row, 7, f'=E{data_row}-F{data_row}')
        
        # H: Multiplier effect = (I-M) * demand_multiplier
        ws.cell(data_row, 8, f'=G{data_row}*D${mult_row}')
        
        # I: % of GDP = multiplier_effect / real_GDP
        ws.cell(data_row, 9, f'=IF(C{data_row}=0,0,H{data_row}/C{data_row})')
        
        # J: Baseline GDP Growth (link from WEO_Data column B)
        if year in weo_map:
            ws.cell(data_row, 10, f"='{weo_sheet}'!B{weo_map[year]}")
        
        # K: Projected Growth = baseline + % of GDP * 100
        ws.cell(data_row, 11, f'=J{data_row}+I{data_row}*100')
    
    return alloc_row  # return last row used


def build_complete_workbook(output_path, sut_path=None, sut_url=None,
                            gdp_files=None, weo_data=None):
    """End-to-end: build the complete test_demand.xlsx workbook."""
    
    # Load template
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
    
    # Ensure NA sheet exists
    if 'NA' not in wb.sheetnames:
        wb.create_sheet('NA')
    
    # Download SUT
    if sut_path is None:
        if sut_url is None:
            sut_url = find_latest_sut_url()
        if sut_url:
            sut_path = '/tmp/sut_latest.xlsx'
            download_file(sut_url, sut_path)
    
    # Download GDP files
    if gdp_files is None:
        gdp_files = download_geostat_gdp_files()
    
    # Extract geostat data
    geostat_data = {}
    if 'gdp_current' in gdp_files:
        geostat_data['nominal_gdp'] = extract_geostat_annual_gdp(
            gdp_files['gdp_current'], '(=) GDP at market prices')
    if 'gdp_constant' in gdp_files:
        geostat_data['real_gdp'] = extract_geostat_annual_gdp(
            gdp_files['gdp_constant'], '(=) GDP at market prices')
    if 'gdp_growth' in gdp_files:
        geostat_data['deflator_growth'] = extract_geostat_annual_gdp(
            gdp_files['gdp_growth'], 'GDP deflator')
    
    # Fetch WEO data
    if weo_data is None:
        weo_data = fetch_imf_weo_data('GEO')
    
    # Build WEO_Data sheet
    build_weo_data_sheet(wb, weo_data, geostat_data)
    
    # Copy SUT sheets
    supply_sheet_name = None
    use_sheet_name = None
    if sut_path:
        sut_wb = openpyxl.load_workbook(sut_path, data_only=True)
        for sn in sut_wb.sheetnames:
            if 'SUPPLY' in sn.upper() and '38' in sn:
                supply_sheet_name = sn
            if 'USE' in sn.upper() and '38' in sn:
                use_sheet_name = sn
        if supply_sheet_name and use_sheet_name:
            copy_sut_sheet(sut_wb, supply_sheet_name, wb, supply_sheet_name)
            copy_sut_sheet(sut_wb, use_sheet_name, wb, use_sheet_name)
    
    # Build SUT Calc
    import_content_cell = "'SUT Calc'!C46"
    if supply_sheet_name and use_sheet_name:
        _, import_content_cell = build_sut_calc_sheet(
            wb, supply_sheet_name, use_sheet_name)
    
    # Build NA scenarios
    # Scenario 1: start at row 1 (matching template)
    last_row_s1 = build_na_scenario(
        wb, start_row=1, scenario_num=1,
        import_content_cell=import_content_cell)
    
    # Scenario 2: multiplier = 1, start below scenario 1
    s2_start = last_row_s1 + 3
    last_row_s2 = build_na_scenario(
        wb, start_row=s2_start, scenario_num=2,
        import_content_cell=import_content_cell,
        multiplier_override=1.0)
    
    # Scenario 3: import content = 0.5, start below scenario 2
    s3_start = last_row_s2 + 3
    last_row_s3 = build_na_scenario(
        wb, start_row=s3_start, scenario_num=3,
        import_content_cell=import_content_cell,
        import_content_override=0.5)
    
    # Remove default Sheet if present
    if 'Sheet' in wb.sheetnames:
        sheet = wb['Sheet']
        if sheet.max_row <= 1 and sheet.max_column <= 1:
            del wb['Sheet']
    
    wb.save(output_path)
    print(f'Workbook saved to {output_path}')
    return output_path


def validate_workbook(path):
    """Validate the completed workbook."""
    wb = openpyxl.load_workbook(path)
    issues = []
    
    required = ['WEO_Data', 'NA', 'SUT Calc']
    for s in required:
        if s not in wb.sheetnames:
            issues.append(f'Missing sheet: {s}')
    
    has_supply = any('SUPPLY' in s.upper() and '38' in s for s in wb.sheetnames)
    has_use = any('USE' in s.upper() and '38' in s for s in wb.sheetnames)
    if not has_supply:
        issues.append('Missing SUPPLY (38-38) sheet')
    if not has_use:
        issues.append('Missing USE (38-38) sheet')
    
    if 'NA' in wb.sheetnames:
        na = wb['NA']
        # Check scenarios exist
        assumption_count = 0
        for r in range(1, na.max_row + 1):
            v = na.cell(r, 3).value
            if v and 'assumptions' in str(v).lower():
                assumption_count += 1
        if assumption_count < 3:
            issues.append(f'Expected 3 scenario assumption blocks, found {assumption_count}')
    
    if issues:
        print('Validation issues:')
        for i in issues:
            print(f'  - {i}')
        return False
    else:
        print('Validation passed!')
        return True
