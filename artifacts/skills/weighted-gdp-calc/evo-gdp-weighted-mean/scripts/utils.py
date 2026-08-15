import openpyxl
from openpyxl.utils import get_column_letter
import subprocess
import os
import zipfile
import xml.etree.ElementTree as ET
import io


def discover_task_layout(wb):
    """Discover the layout of the Task sheet at runtime."""
    ws = wb['Task']
    layout = {}
    
    # Find year row
    year_row = None
    for r in range(8, 12):
        for c in range(8, 13):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and 2000 <= v <= 2100:
                year_row = r
                break
        if year_row:
            break
    layout['year_row'] = year_row
    
    # Find data sections
    export_rows = []
    import_rows = []
    gdp_rows = []
    
    for r in range(10, 35):
        d_val = ws.cell(row=r, column=4).value
        if d_val and isinstance(d_val, str):
            if 'BXGS' in d_val:
                export_rows.append(r)
            elif 'BMGS' in d_val:
                import_rows.append(r)
            elif 'NGDPD' in d_val:
                gdp_rows.append(r)
    
    layout['export_rows'] = export_rows
    layout['import_rows'] = import_rows
    layout['gdp_rows'] = gdp_rows
    
    # Find net export rows
    net_export_rows = []
    for r in range(33, 50):
        e_val = ws.cell(row=r, column=5).value
        if e_val and isinstance(e_val, str) and 'Net Export' in e_val:
            f_val = ws.cell(row=r, column=6).value
            if f_val and isinstance(f_val, str) and 'Percent' in f_val:
                net_export_rows.append(r)
    layout['net_export_rows'] = net_export_rows
    
    # Find statistics rows
    stats = {}
    for r in range(40, 55):
        e_val = ws.cell(row=r, column=5).value
        if e_val and isinstance(e_val, str):
            e_lower = e_val.lower()
            if 'min' in e_lower and 'weighted' not in e_lower:
                stats['min'] = r
            elif 'max' in e_lower:
                stats['max'] = r
            elif 'median' in e_lower:
                stats['median'] = r
            elif 'simple mean' in e_lower:
                stats['mean'] = r
            elif '25' in e_lower and 'percentile' in e_lower:
                stats['p25'] = r
            elif '75' in e_lower and 'percentile' in e_lower:
                stats['p75'] = r
            elif 'weighted mean' in e_lower:
                stats['weighted_mean'] = r
    layout['stats'] = stats
    layout['first_data_col'] = 8
    layout['last_data_col'] = 12
    
    return layout


def discover_data_layout(wb):
    """Discover the layout of the Data sheet."""
    ws = wb['Data']
    layout = {}
    
    header_row = None
    series_code_col = None
    for r in range(1, 10):
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and v == 'Series_code':
                header_row = r
                series_code_col = c
                break
        if header_row:
            break
    layout['header_row'] = header_row
    layout['series_code_col'] = series_code_col
    
    year_cols = {}
    if header_row:
        for c in range(1, 20):
            v = ws.cell(row=header_row, column=c).value
            if isinstance(v, (int, float)) and 2000 <= v <= 2100:
                year_cols[int(v)] = c
    layout['year_cols'] = year_cols
    layout['data_start_row'] = 21
    layout['data_end_row'] = 40
    
    if year_cols:
        layout['first_year_col'] = min(year_cols.values())
        layout['last_year_col'] = max(year_cols.values())
    else:
        layout['first_year_col'] = 8
        layout['last_year_col'] = 13
    
    return layout


def write_lookup_formulas(wb, task_layout, data_layout):
    """Write INDEX/MATCH formulas for Step 1."""
    ws = wb['Task']
    year_row = task_layout['year_row']
    all_lookup_rows = (task_layout['export_rows'] + 
                       task_layout['import_rows'] + 
                       task_layout['gdp_rows'])
    
    d_start = data_layout['data_start_row']
    d_end = data_layout['data_end_row']
    sc_col = get_column_letter(data_layout['series_code_col'])
    fy_col = get_column_letter(data_layout['first_year_col'])
    ly_col = get_column_letter(data_layout['last_year_col'])
    hr = data_layout['header_row']
    
    data_range = f"Data!${fy_col}${d_start}:${ly_col}${d_end}"
    series_range = f"Data!${sc_col}${d_start}:${sc_col}${d_end}"
    year_header_range = f"Data!${fy_col}${hr}:${ly_col}${hr}"
    
    for row_num in all_lookup_rows:
        for col_num in range(task_layout['first_data_col'], task_layout['last_data_col'] + 1):
            col_letter = get_column_letter(col_num)
            formula = (f'=INDEX({data_range},'
                      f'MATCH($D{row_num},{series_range},0),'
                      f'MATCH({col_letter}${year_row},{year_header_range},0))')
            ws.cell(row=row_num, column=col_num).value = formula
    return wb


def write_net_export_formulas(wb, task_layout):
    """Write net export as % of GDP formulas for Step 2."""
    ws = wb['Task']
    export_rows = task_layout['export_rows']
    import_rows = task_layout['import_rows']
    gdp_rows = task_layout['gdp_rows']
    net_export_rows = task_layout['net_export_rows']
    
    for i, ne_row in enumerate(net_export_rows):
        if i >= len(export_rows) or i >= len(import_rows) or i >= len(gdp_rows):
            break
        exp_row = export_rows[i]
        imp_row = import_rows[i]
        gdp_row = gdp_rows[i]
        for col_num in range(task_layout['first_data_col'], task_layout['last_data_col'] + 1):
            col_letter = get_column_letter(col_num)
            formula = (f'=({col_letter}{exp_row}-{col_letter}{imp_row})'
                      f'/{col_letter}{gdp_row}*100')
            ws.cell(row=ne_row, column=col_num).value = formula
    return wb


def write_statistics_formulas(wb, task_layout):
    """Write descriptive statistics formulas for Step 2."""
    ws = wb['Task']
    ne_rows = task_layout['net_export_rows']
    stats = task_layout['stats']
    first_ne = ne_rows[0]
    last_ne = ne_rows[-1]
    
    for col_num in range(task_layout['first_data_col'], task_layout['last_data_col'] + 1):
        col_letter = get_column_letter(col_num)
        data_range = f'{col_letter}{first_ne}:{col_letter}{last_ne}'
        if 'min' in stats:
            ws.cell(row=stats['min'], column=col_num).value = f'=MIN({data_range})'
        if 'max' in stats:
            ws.cell(row=stats['max'], column=col_num).value = f'=MAX({data_range})'
        if 'median' in stats:
            ws.cell(row=stats['median'], column=col_num).value = f'=MEDIAN({data_range})'
        if 'mean' in stats:
            ws.cell(row=stats['mean'], column=col_num).value = f'=AVERAGE({data_range})'
        if 'p25' in stats:
            ws.cell(row=stats['p25'], column=col_num).value = f'=PERCENTILE({data_range},0.25)'
        if 'p75' in stats:
            ws.cell(row=stats['p75'], column=col_num).value = f'=PERCENTILE({data_range},0.75)'
    return wb


def write_weighted_mean_formula(wb, task_layout):
    """Write weighted mean formula for Step 3 using SUMPRODUCT."""
    ws = wb['Task']
    ne_rows = task_layout['net_export_rows']
    gdp_rows = task_layout['gdp_rows']
    stats = task_layout['stats']
    wm_row = stats.get('weighted_mean')
    if not wm_row:
        return wb
    first_ne = ne_rows[0]
    last_ne = ne_rows[-1]
    first_gdp = gdp_rows[0]
    last_gdp = gdp_rows[-1]
    
    for col_num in range(task_layout['first_data_col'], task_layout['last_data_col'] + 1):
        col_letter = get_column_letter(col_num)
        ne_range = f'{col_letter}{first_ne}:{col_letter}{last_ne}'
        gdp_range = f'{col_letter}{first_gdp}:{col_letter}{last_gdp}'
        formula = f'=SUMPRODUCT({ne_range},{gdp_range})/SUM({gdp_range})'
        ws.cell(row=wm_row, column=col_num).value = formula
    return wb


def compute_cached_values(wb):
    """Compute cached values for all formula cells."""
    ws_data = wb['Data']
    ws_task = wb['Task']
    
    # Build data lookup from Data sheet
    data = {}
    for r in range(21, 41):
        code = ws_data.cell(row=r, column=2).value
        if code:
            data[code] = {}
            for c in range(8, 14):
                year = ws_data.cell(row=4, column=c).value
                data[code][year] = ws_data.cell(row=r, column=c).value
    
    # Get series codes
    series_codes = {}
    for r in range(12, 32):
        v = ws_task.cell(row=r, column=4).value
        if v and isinstance(v, str) and not v.startswith('='):
            series_codes[r] = v
    
    years = {8: 2019, 9: 2020, 10: 2021, 11: 2022, 12: 2023}
    cached = {}
    
    # Lookup values
    for r in list(range(12, 18)) + list(range(19, 25)) + list(range(26, 32)):
        code = series_codes.get(r)
        if code:
            for c, year in years.items():
                val = data.get(code, {}).get(year)
                if val is not None:
                    cached[(r, c)] = val
    
    # Net exports % GDP
    export_rows = [12, 13, 14, 15, 16, 17]
    import_rows = [19, 20, 21, 22, 23, 24]
    gdp_rows = [26, 27, 28, 29, 30, 31]
    ne_rows = [35, 36, 37, 38, 39, 40]
    
    for i in range(6):
        for c, year in years.items():
            exp = cached.get((export_rows[i], c))
            imp = cached.get((import_rows[i], c))
            gdp = cached.get((gdp_rows[i], c))
            if all(v is not None for v in [exp, imp, gdp]) and gdp != 0:
                cached[(ne_rows[i], c)] = (exp - imp) / gdp * 100
    
    # Statistics
    for c in years:
        ne_vals = [cached.get((r, c)) for r in ne_rows if cached.get((r, c)) is not None]
        if ne_vals:
            sorted_vals = sorted(ne_vals)
            n = len(sorted_vals)
            cached[(42, c)] = min(ne_vals)
            cached[(43, c)] = max(ne_vals)
            cached[(44, c)] = (sorted_vals[n//2-1] + sorted_vals[n//2]) / 2 if n % 2 == 0 else sorted_vals[n//2]
            cached[(45, c)] = sum(ne_vals) / n
            for pct, sr in [(0.25, 46), (0.75, 47)]:
                k = pct * (n - 1)
                idx = int(k)
                frac = k - idx
                cached[(sr, c)] = sorted_vals[idx] + frac * (sorted_vals[idx+1] - sorted_vals[idx]) if idx+1 < n else sorted_vals[idx]
            gdp_vals = [cached.get((gdp_rows[i], c)) for i in range(6)]
            if all(g is not None for g in gdp_vals):
                cached[(50, c)] = sum(p*g for p,g in zip(ne_vals, gdp_vals)) / sum(gdp_vals)
    
    return cached


def write_cached_values_to_xml(filepath, cached_values):
    """Write cached values into the XML of the xlsx file."""
    ns_uri = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ET.register_namespace('', ns_uri)
    ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
    
    col_letters = {8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L'}
    
    with zipfile.ZipFile(filepath, 'r') as zin:
        file_contents = {name: zin.read(name) for name in zin.namelist()}
    
    tree = ET.parse(io.BytesIO(file_contents['xl/worksheets/sheet1.xml']))
    root = tree.getroot()
    
    for row_elem in root.findall(f'.//{{{ns_uri}}}row'):
        row_num = int(row_elem.get('r'))
        for cell_elem in row_elem.findall(f'{{{ns_uri}}}c'):
            ref = cell_elem.get('r')
            col_letter = ref.rstrip('0123456789')
            col_num = None
            for cn, cl in col_letters.items():
                if cl == col_letter:
                    col_num = cn
                    break
            if col_num is None:
                continue
            key = (row_num, col_num)
            if key in cached_values:
                formula_elem = cell_elem.find(f'{{{ns_uri}}}f')
                if formula_elem is not None:
                    v_elem = cell_elem.find(f'{{{ns_uri}}}v')
                    if v_elem is None:
                        v_elem = ET.SubElement(cell_elem, f'{{{ns_uri}}}v')
                    v_elem.text = str(cached_values[key])
                    if 't' in cell_elem.attrib:
                        del cell_elem.attrib['t']
    
    modified_xml = io.BytesIO()
    tree.write(modified_xml, xml_declaration=True, encoding='UTF-8')
    file_contents['xl/worksheets/sheet1.xml'] = modified_xml.getvalue()
    
    with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, content in file_contents.items():
            zout.writestr(name, content)


def run_end_to_end(input_path, output_path=None):
    """End-to-end entry point."""
    if output_path is None:
        output_path = input_path
    
    wb = openpyxl.load_workbook(input_path)
    task_layout = discover_task_layout(wb)
    data_layout = discover_data_layout(wb)
    
    print(f"Task layout: year_row={task_layout['year_row']}")
    print(f"  Export rows: {task_layout['export_rows']}")
    print(f"  Import rows: {task_layout['import_rows']}")
    print(f"  GDP rows: {task_layout['gdp_rows']}")
    print(f"  Net export rows: {task_layout['net_export_rows']}")
    print(f"  Stats: {task_layout['stats']}")
    
    print("\nStep 1: Writing lookup formulas...")
    wb = write_lookup_formulas(wb, task_layout, data_layout)
    print("Step 2: Writing net export and statistics formulas...")
    wb = write_net_export_formulas(wb, task_layout)
    wb = write_statistics_formulas(wb, task_layout)
    print("Step 3: Writing weighted mean formula...")
    wb = write_weighted_mean_formula(wb, task_layout)
    
    # Compute cached values
    print("Computing cached values...")
    cached = compute_cached_values(wb)
    
    wb.save(output_path)
    wb.close()
    
    # Write cached values into XML
    print("Writing cached values to XML...")
    write_cached_values_to_xml(output_path, cached)
    
    print(f"Saved to {output_path}")
    return output_path
