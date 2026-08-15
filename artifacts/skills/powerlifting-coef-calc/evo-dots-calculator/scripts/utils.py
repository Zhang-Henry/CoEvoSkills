import openpyxl
from copy import copy

def get_dots_columns():
    """Return the column names needed for Dots calculation."""
    return ['Name', 'Sex', 'BodyweightKg', 'Best3SquatKg', 'Best3BenchKg', 'Best3DeadliftKg']

def find_column_indices(header_row, column_names):
    """Find column indices (0-based) for the given column names in the header row.
    Returns dict mapping column_name -> index."""
    headers = [cell.value for cell in header_row]
    result = {}
    for name in column_names:
        if name in headers:
            result[name] = headers.index(name)
        else:
            raise ValueError(f"Column '{name}' not found in headers: {headers}")
    return result

def fmt_coeff(val):
    """Format a coefficient to avoid scientific notation in Excel formulas."""
    return f'{val:.15g}'

def copy_columns_to_dots(wb, source_sheet_name='Data', dest_sheet_name='Dots'):
    """Copy the required columns from Data sheet to Dots sheet.
    Returns the workbook and column mapping info."""
    ws_data = wb[source_sheet_name]
    ws_dots = wb[dest_sheet_name]
    
    # Clear existing content in Dots sheet
    for row in ws_dots.iter_rows():
        for cell in row:
            cell.value = None
    
    needed_cols = get_dots_columns()
    header_row = list(ws_data[1])
    col_indices = find_column_indices(header_row, needed_cols)
    
    # Sort by original column order (index in Data sheet)
    sorted_cols = sorted(col_indices.items(), key=lambda x: x[1])
    
    max_row = ws_data.max_row
    
    for row_idx in range(1, max_row + 1):
        source_row = list(ws_data[row_idx])
        for dest_col_idx, (col_name, src_col_idx) in enumerate(sorted_cols, start=1):
            src_cell = source_row[src_col_idx]
            dest_cell = ws_dots.cell(row=row_idx, column=dest_col_idx)
            dest_cell.value = src_cell.value
            if src_cell.number_format:
                dest_cell.number_format = src_cell.number_format
    
    return wb, sorted_cols, max_row

def add_totalkg_formula(ws_dots, sorted_cols, max_row):
    """Add TotalKg column with Excel formula after existing columns.
    TotalKg = Best3SquatKg + Best3BenchKg + Best3DeadliftKg"""
    num_existing_cols = len(sorted_cols)
    total_col = num_existing_cols + 1
    
    col_name_to_dest = {}
    for dest_idx, (col_name, _) in enumerate(sorted_cols, start=1):
        col_name_to_dest[col_name] = dest_idx
    
    sq_col_letter = openpyxl.utils.get_column_letter(col_name_to_dest['Best3SquatKg'])
    bn_col_letter = openpyxl.utils.get_column_letter(col_name_to_dest['Best3BenchKg'])
    dl_col_letter = openpyxl.utils.get_column_letter(col_name_to_dest['Best3DeadliftKg'])
    
    ws_dots.cell(row=1, column=total_col, value='TotalKg')
    
    for row_idx in range(2, max_row + 1):
        formula = f'=ROUND({sq_col_letter}{row_idx}+{bn_col_letter}{row_idx}+{dl_col_letter}{row_idx},3)'
        ws_dots.cell(row=row_idx, column=total_col, value=formula)
    
    return total_col

def add_dots_formula(ws_dots, sorted_cols, max_row, total_col):
    """Add Dots column with Excel formula after TotalKg.
    Dots = TotalKg * 500 / (A*bw^4 + B*bw^3 + C*bw^2 + D*bw + E)
    where coefficients depend on Sex.
    
    Male coefficients:
    A = -0.000001093
    B = 0.0007391293
    C = -0.1918759221
    D = 24.0900756
    E = -307.75076
    
    Female coefficients:
    A = -0.0000010706
    B = 0.0005158568
    C = -0.1126655495
    D = 13.6175032
    E = -57.96288
    """
    dots_col = total_col + 1
    
    col_name_to_dest = {}
    for dest_idx, (col_name, _) in enumerate(sorted_cols, start=1):
        col_name_to_dest[col_name] = dest_idx
    
    sex_col_letter = openpyxl.utils.get_column_letter(col_name_to_dest['Sex'])
    bw_col_letter = openpyxl.utils.get_column_letter(col_name_to_dest['BodyweightKg'])
    total_col_letter = openpyxl.utils.get_column_letter(total_col)
    
    ws_dots.cell(row=1, column=dots_col, value='Dots')
    
    # Use explicit decimal notation to avoid scientific notation issues
    mA = '-0.000001093'
    mB = '0.0007391293'
    mC = '-0.1918759221'
    mD = '24.0900756'
    mE = '-307.75076'
    
    fA = '-0.0000010706'
    fB = '0.0005158568'
    fC = '-0.1126655495'
    fD = '13.6175032'
    fE = '-57.96288'
    
    for row_idx in range(2, max_row + 1):
        bw = f'{bw_col_letter}{row_idx}'
        sex = f'{sex_col_letter}{row_idx}'
        tot = f'{total_col_letter}{row_idx}'
        
        formula = (
            f'=ROUND({tot}*500/('
            f'IF({sex}="M",{mA},{fA})*{bw}^4+'
            f'IF({sex}="M",{mB},{fB})*{bw}^3+'
            f'IF({sex}="M",{mC},{fC})*{bw}^2+'
            f'IF({sex}="M",{mD},{fD})*{bw}+'
            f'IF({sex}="M",{mE},{fE})'
            f'),3)'
        )
        ws_dots.cell(row=row_idx, column=dots_col, value=formula)
    
    return dots_col

def run_dots_pipeline(input_path, output_path=None):
    """End-to-end entry point: read workbook, populate Dots sheet, save."""
    if output_path is None:
        output_path = input_path
    
    wb = openpyxl.load_workbook(input_path)
    
    wb, sorted_cols, max_row = copy_columns_to_dots(wb)
    
    ws_dots = wb['Dots']
    
    total_col = add_totalkg_formula(ws_dots, sorted_cols, max_row)
    dots_col = add_dots_formula(ws_dots, sorted_cols, max_row, total_col)
    
    wb.save(output_path)
    print(f"Saved workbook to {output_path}")
    print(f"Dots sheet has {max_row} rows, {dots_col} columns")
    return output_path

def validate_dots_workbook(path):
    """Validate the output workbook has correct structure."""
    wb = openpyxl.load_workbook(path)
    
    assert 'Dots' in wb.sheetnames, "Missing 'Dots' sheet"
    ws = wb['Dots']
    
    headers = [cell.value for cell in ws[1]]
    print(f"Dots headers: {headers}")
    
    expected_cols = get_dots_columns() + ['TotalKg', 'Dots']
    assert headers == expected_cols, f"Headers mismatch: {headers} != {expected_cols}"
    
    max_row = ws.max_row
    assert max_row > 1, "No data rows in Dots sheet"
    
    for row_idx in range(2, max_row + 1):
        total_cell = ws.cell(row=row_idx, column=len(get_dots_columns()) + 1)
        dots_cell = ws.cell(row=row_idx, column=len(get_dots_columns()) + 2)
        
        assert total_cell.value is not None, f"TotalKg empty at row {row_idx}"
        assert dots_cell.value is not None, f"Dots empty at row {row_idx}"
        
        total_val = str(total_cell.value)
        dots_val = str(dots_cell.value)
        
        assert total_val.startswith('='), f"TotalKg at row {row_idx} is not a formula: {total_val}"
        assert dots_val.startswith('='), f"Dots at row {row_idx} is not a formula: {dots_val}"
    
    print(f"Validation passed: {max_row - 1} data rows with formulas")
    
    # Print sample formulas
    print(f"Sample TotalKg formula (row 2): {ws.cell(row=2, column=len(get_dots_columns()) + 1).value}")
    print(f"Sample Dots formula (row 2): {ws.cell(row=2, column=len(get_dots_columns()) + 2).value}")
    
    # Verify computation with Python for first row
    name = ws.cell(row=2, column=1).value
    sex = ws.cell(row=2, column=2).value
    bw = ws.cell(row=2, column=3).value
    sq = ws.cell(row=2, column=4).value
    bn = ws.cell(row=2, column=5).value
    dl = ws.cell(row=2, column=6).value
    
    total = round(sq + bn + dl, 3)
    
    if sex == 'M':
        A, B, C, D, E = -0.000001093, 0.0007391293, -0.1918759221, 24.0900756, -307.75076
    else:
        A, B, C, D, E = -0.0000010706, 0.0005158568, -0.1126655495, 13.6175032, -57.96288
    
    denom = A*bw**4 + B*bw**3 + C*bw**2 + D*bw + E
    dots = round(total * 500 / denom, 3)
    
    print(f"\nVerification for {name} (Sex={sex}, BW={bw}):")
    print(f"  Total: {sq} + {bn} + {dl} = {total}")
    print(f"  Denominator: {denom}")
    print(f"  Dots: {dots}")
    
    return True
