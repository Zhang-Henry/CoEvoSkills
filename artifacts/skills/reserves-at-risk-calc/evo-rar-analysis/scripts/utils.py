"""
Reserves-at-Risk workbook builder.
Uses .GEO codes for cross-sheet entity matching.
All structure discovered at runtime.
"""
import openpyxl, requests, os, subprocess, shutil, re


def download_imf_commodity_data(output_path):
    if os.path.exists(output_path):
        return output_path
    page = requests.get('https://www.imf.org/en/research/commodity-prices', timeout=30)
    page.raise_for_status()
    links = re.findall(r'href="([^"]*\.xls[x]?[^"]*?)"', page.text, re.I)
    if not links:
        raise ValueError("No Excel link found")
    url = links[0] if links[0].startswith('http') else 'https://www.imf.org' + links[0]
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(output_path, 'wb') as f:
        f.write(r.content)
    return output_path


def extract_prices(imf_path, header_text):
    wb = openpyxl.load_workbook(imf_path, data_only=True)
    col, fdr, ws = None, None, None
    for sn in wb.sheetnames:
        w = wb[sn]
        for c in range(1, w.max_column + 1):
            d = w.cell(row=2, column=c).value
            if d and header_text and header_text.lower()[:40] in str(d).lower():
                col, ws = c, w
                break
            h = w.cell(row=1, column=c).value
            if h and 'gold' in str(h).lower() and col is None:
                col, ws = c, w
        if col: break
    if not col:
        raise ValueError("Commodity column not found")
    for r in range(2, 20):
        v = ws.cell(row=r, column=1).value
        if v and re.match(r'\d{4}M\d{1,2}', str(v)):
            fdr = r; break
    prices = {}
    for r in range(fdr, ws.max_row + 1):
        dl = ws.cell(row=r, column=1).value
        p = ws.cell(row=r, column=col).value
        if dl and p is not None:
            try: prices[str(dl)] = float(p)
            except: pass
    wb.close()
    return prices


def _parse_dl(label):
    m = re.match(r'(\d{4})M(\d{1,2})', str(label))
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def _scan_dates(ws):
    dates = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and re.match(r'\d{4}M\d{1,2}', str(v)):
            dates.append((r, str(v)))
        elif v is None and dates: break
    return dates


def _sheet_roles(wb):
    roles = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        a1 = ws.cell(row=1, column=1).value
        b1 = ws.cell(row=1, column=2).value
        has_dates = any(ws.cell(row=r, column=1).value and
                        re.match(r'\d{4}M\d{1,2}', str(ws.cell(row=r, column=1).value))
                        for r in range(2, min(10, ws.max_row + 1)))
        has_steps = any(ws.cell(row=r, column=2).value and
                        re.match(r'(?i)step\s*\d', str(ws.cell(row=r, column=2).value))
                        for r in range(1, min(20, ws.max_row + 1)))
        is_desc = (a1 and str(a1).strip() == '.DESC')
        if has_dates and b1 and not is_desc: roles['price'] = sn
        elif has_steps: roles['answer'] = sn
        elif is_desc:
            sl = sn.lower()
            if 'total' in sl: roles['total_reserves'] = sn
            elif 'vol' in sl: roles['volume'] = sn
            elif 'value' not in roles: roles['value'] = sn
    return roles


def _latest_year(ws):
    best, best_r = None, None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        try:
            yr = int(str(v).strip())
            if yr > 1900 and (best is None or yr > best): best, best_r = yr, r
        except: pass
    return best, best_r


def _find_meta_row(ws, marker):
    """Find a metadata row by its column A marker (e.g. '.GEO', '.DATA_TYPE')."""
    for r in range(1, min(15, ws.max_row + 1)):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip() == marker:
            return r
    return None


def _countries(ws, year):
    """Find countries with data for the given year.
    Returns list of dicts including GEO code for cross-sheet matching."""
    dr = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None and \
           str(ws.cell(row=r, column=1).value).strip() == str(year):
            dr = r; break
    if dr is None: return []
    
    cur_row = _find_meta_row(ws, '.DATA_TYPE')
    geo_row = _find_meta_row(ws, '.GEO')
    
    out = []
    for c in range(3, ws.max_column + 1):
        desc = ws.cell(row=1, column=c).value
        val = ws.cell(row=dr, column=c).value
        cur = ws.cell(row=cur_row, column=c).value if cur_row else None
        geo = ws.cell(row=geo_row, column=c).value if geo_row else None
        if desc and val is not None:
            raw = str(desc).split(':')[0].strip()
            out.append({
                'raw': raw, 'desc': str(desc),
                'cl': openpyxl.utils.get_column_letter(c),
                'ci': c, 'dr': dr, 'val': val,
                'cur': str(cur).strip() if cur else None,
                'geo': str(geo).strip() if geo else None
            })
    return out


def _display_name_for_geo(geo_code, all_sheets_countries):
    """Find a display name for a GEO code that will substring-match
    all header variants across sheets. Uses the longest common word
    prefix of all names found for this GEO code."""
    names = []
    for countries in all_sheets_countries:
        for c in countries:
            if c.get('geo') == geo_code:
                names.append(c['raw'])
    if not names:
        return None
    if len(set(names)) == 1:
        return names[0]
    # Find longest common word prefix
    word_lists = [n.split() for n in names]
    min_len = min(len(wl) for wl in word_lists)
    common = []
    for i in range(min_len):
        if len({wl[i] for wl in word_lists}) == 1:
            common.append(word_lists[0][i])
        else:
            break
    if common:
        return ' '.join(common)
    return min(names, key=len)


def _vol_unit(desc):
    d = desc.lower()
    if 'thous' in d and ('troy' in d or 'ounce' in d):
        return 1.0 / 1000.0
    return 1.0


def _answer_rows(ws):
    rows = {}
    secs = []
    for r in range(1, ws.max_row + 1):
        lb = ws.cell(row=r, column=2).value
        if lb is None: continue
        ll = str(lb).strip().lower()
        sm = re.match(r'step\s*(\d+)', ll)
        if sm: secs.append((r, int(sm.group(1)))); continue
        if 'z-score' in ll or 'z score' in ll: rows['zs'] = r
        elif 'annualized' in ll: rows['v3a'] = r
        elif re.search(r'3.month', ll) and 'vol' in ll: rows['v3'] = r
        elif re.search(r'12.month', ll) and 'vol' in ll: rows['v12'] = r
    for i, (sr, sn) in enumerate(secs):
        nx = secs[i+1][0] if i+1 < len(secs) else ws.max_row + 1
        p = f's{sn}'
        for r in range(sr+1, nx):
            lb = ws.cell(row=r, column=2).value
            if lb is None: continue
            ll = str(lb).strip().lower()
            if 'country' in ll and f'{p}_c' not in rows: rows[f'{p}_c'] = r
            elif 'gold reserves' in ll and f'{p}_r' not in rows: rows[f'{p}_r'] = r
            elif 'exposure' in ll and f'{p}_e' not in rows: rows[f'{p}_e'] = r
            elif ll.startswith('rar'): rows[f'{p}_rar'] = r
            elif 'total reserve' in ll and f'{p}_tr' not in rows: rows[f'{p}_tr'] = r
    return rows


def _col_start(ws, row):
    lc = 2
    for c in range(1, 5):
        if ws.cell(row=row, column=c).value is not None: lc = c
    return lc + 1


def run_end_to_end(template_path, imf_path, output_path,
                   z_score, avg_price_max_month):
    """Build complete RaR workbook. All task parameters supplied by caller."""
    download_imf_commodity_data(imf_path)
    wb = openpyxl.load_workbook(template_path)
    roles = _sheet_roles(wb)
    ws_p = wb[roles['price']]
    ws_a = wb[roles['answer']]
    header = ws_p.cell(row=1, column=2).value
    prices = extract_prices(imf_path, header)
    dates = _scan_dates(ws_p)
    
    # Populate prices and formulas
    last = None
    for r, dl in dates:
        if dl in prices: ws_p.cell(row=r, column=2, value=prices[dl]); last = r
    if last is None: raise ValueError("No prices matched")
    first = dates[0][0]; fl = first + 1
    for r, _ in dates:
        if r == first or r > last: continue
        ws_p.cell(row=r, column=3, value=f'=LN(B{r}/B{r-1})*100')
    for r, _ in dates:
        if r < fl+2 or r > last: continue
        ws_p.cell(row=r, column=4, value=f'=STDEV(C{r-2}:C{r})')
    for r, _ in dates:
        if r < fl+11 or r > last: continue
        ws_p.cell(row=r, column=5, value=f'=STDEV(C{r-11}:C{r})')
    
    ar = _answer_rows(ws_a)
    ps = roles['price']
    
    # STEP 1
    ws_a.cell(row=ar['zs'], column=3, value=z_score)
    ws_a.cell(row=ar['v3'], column=3, value=f"='{ps}'!D{last}")
    ws_a.cell(row=ar['v3a'], column=3, value=f"=C{ar['v3']}*SQRT(12)")
    ws_a.cell(row=ar['v12'], column=3, value=f"='{ps}'!E{last}")
    
    # STEP 2 - discover countries
    ws_val = wb[roles['value']]
    ty, _ = _latest_year(ws_val)
    
    # Get all countries from all sheets for name resolution
    val_c = _countries(ws_val, ty)
    ws_vol = wb[roles['volume']]
    vol_c = _countries(ws_vol, ty)
    ws_tr = wb[roles['total_reserves']]
    tr_c = _countries(ws_tr, ty)
    all_sheets = [val_c, vol_c, tr_c]
    
    # Value countries with USD currency
    val_usd = [c for c in val_c if c['cur'] == 'US$']
    val_geos = {c['geo'] for c in val_usd if c.get('geo')}
    
    # Volume countries not in value (by GEO code)
    vol_extra = [c for c in vol_c if c.get('geo') and c['geo'] not in val_geos]
    
    # Average price rows
    avg_rows = [r for r, dl in dates if r <= last
                and _parse_dl(dl)[0] == ty
                and 1 <= _parse_dl(dl)[1] <= avg_price_max_month]
    if not avg_rows: raise ValueError(f"No avg price rows")
    ars, are = min(avg_rows), max(avg_rows)
    
    # Build step2 list with clean display names
    s2 = []
    for c in val_usd:
        display = _display_name_for_geo(c['geo'], all_sheets) or c['raw']
        s2.append({'nm': display, 'geo': c['geo'], 'src': 'v',
                   'sh': roles['value'], 'cl': c['cl'], 'dr': c['dr']})
    for c in vol_extra:
        display = _display_name_for_geo(c['geo'], all_sheets) or c['raw']
        s2.append({'nm': display, 'geo': c['geo'], 'src': 'o',
                   'sh': roles['volume'], 'cl': c['cl'], 'dr': c['dr'],
                   'desc': c['desc'], 'um': _vol_unit(c['desc'])})
    
    cr, rr, er = ar['s2_c'], ar['s2_r'], ar['s2_e']
    zr, v3r = ar['zs'], ar['v3']
    cs = _col_start(ws_a, cr)
    for i, co in enumerate(s2):
        col = cs + i
        cl = openpyxl.utils.get_column_letter(col)
        ws_a.cell(row=cr, column=col, value=co['nm'])
        if co['src'] == 'v':
            ws_a.cell(row=rr, column=col,
                      value=f"={co['sh']}!{co['cl']}{co['dr']}")
        else:
            ref = f"{co['sh']}!{co['cl']}{co['dr']}"
            um = co.get('um', 1.0)
            if um != 1.0:
                ws_a.cell(row=rr, column=col,
                    value=f"={ref}*{um}*AVERAGE('{ps}'!B{ars}:B{are})")
            else:
                ws_a.cell(row=rr, column=col,
                    value=f"={ref}*AVERAGE('{ps}'!B{ars}:B{are})")
        ws_a.cell(row=er, column=col,
                  value=f'={cl}{rr}*$C${zr}*$C${v3r}/100')
    
    # STEP 3 - countries with total reserves
    trs = roles['total_reserves']
    tr_usd_geos = {c['geo'] for c in tr_c if c['cur'] == 'US$' and c.get('geo')}
    
    if tr_c:
        cols_tr = [c['ci'] for c in tr_c]
        tsl = openpyxl.utils.get_column_letter(min(cols_tr))
        tel = openpyxl.utils.get_column_letter(max(cols_tr))
        tdr = tr_c[0]['dr']
        
        s3 = [c for c in s2 if c.get('geo') and c['geo'] in tr_usd_geos]
        
        cr3, rr3, er3 = ar['s3_c'], ar['s3_r'], ar['s3_e']
        tr3, rar3 = ar['s3_tr'], ar['s3_rar']
        c3s = _col_start(ws_a, cr3)
        for i, co in enumerate(s3):
            col = c3s + i
            cl = openpyxl.utils.get_column_letter(col)
            j = next(k for k, x in enumerate(s2) if x['geo'] == co['geo'])
            s2cl = openpyxl.utils.get_column_letter(cs + j)
            ws_a.cell(row=cr3, column=col, value=f'={s2cl}{cr}')
            ws_a.cell(row=rr3, column=col, value=f'={s2cl}{rr}')
            ws_a.cell(row=er3, column=col, value=f'={s2cl}{er}')
            ws_a.cell(row=tr3, column=col,
                value=f"=INDEX('{trs}'!{tsl}{tdr}:{tel}{tdr},"
                      f'MATCH("*"&{cl}{cr3}&"*",\'{trs}\'!{tsl}1:{tel}1,0))')
            ws_a.cell(row=rar3, column=col,
                      value=f'={cl}{er3}/{cl}{tr3}*100')
    
    # Save and recalculate
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    tmp = output_path + '.tmp.xlsx'
    wb.save(tmp); wb.close()
    lo_tmp = '/tmp/_rar_lo.xlsx'
    shutil.copy2(tmp, lo_tmp)
    if os.path.exists(output_path): os.remove(output_path)
    outdir = os.path.dirname(output_path)
    subprocess.run(['soffice', '--headless', '--calc', '--convert-to', 'xlsx',
                    '--outdir', outdir, lo_tmp],
                   capture_output=True, text=True, timeout=120)
    lo_out = os.path.join(outdir, '_rar_lo.xlsx')
    if os.path.exists(lo_out): shutil.move(lo_out, output_path)
    elif os.path.exists(tmp): shutil.copy2(tmp, output_path)
    for f in [tmp, lo_tmp]:
        if os.path.exists(f): os.remove(f)
    return output_path


def validate_workbook(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    roles = _sheet_roles(wb)
    issues = []
    if 'price' in roles:
        ws = wb[roles['price']]
        if sum(1 for r in range(2, ws.max_row+1) if ws.cell(row=r, column=2).value is not None) == 0:
            issues.append('No prices')
    if 'answer' in roles:
        ws = wb[roles['answer']]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if any(e in str(cell.value) for e in ['#REF!','#NAME?','#VALUE!','#DIV/0!','#N/A']):
                        issues.append(f'{cell.coordinate}: {cell.value}')
    wb.close()
    if issues:
        for i in issues: print(f'  - {i}')
        return False
    print('Validation passed.')
    return True
