"""
Utility functions for demographic analysis pivot table creation.
"""
import pdfplumber
import openpyxl
from openpyxl.pivot.table import (
    TableDefinition, PivotField, DataField, RowColField, Location
)
from openpyxl.pivot.cache import (
    CacheDefinition, CacheField, CacheSource, WorksheetSource
)


def extract_pdf_table(pdf_path):
    """Extract tabular data from a multi-page PDF."""
    pdf = pdfplumber.open(pdf_path)
    all_rows = []
    raw_headers = None
    for page in pdf.pages:
        tables = page.extract_tables()
        for t in tables:
            for row in t:
                if row and row[0] == 'SA2_CODE':
                    if raw_headers is None:
                        raw_headers = row
                    continue
                if row:
                    all_rows.append(row)
    pdf.close()
    return raw_headers, all_rows


def fix_truncated_headers(headers):
    """Fix known truncated PDF headers."""
    fixed = []
    for h in headers:
        if h and h.startswith('POPULATION_20') and len(h) < len('POPULATION_2023'):
            fixed.append('POPULATION_2023')
        else:
            fixed.append(h)
    return fixed


def read_income_data(xlsx_path):
    """Read income data from Excel file."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Data']
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(list(row))
    wb.close()
    return headers, rows


def join_data(pop_headers, pop_rows, inc_headers, inc_rows):
    """Join population and income data on SA2_CODE. Filter np rows."""
    inc_by_code = {}
    for row in inc_rows:
        code = str(row[0])
        inc_by_code[code] = row

    inc_extra_indices = []
    inc_extra_headers = []
    for i, h in enumerate(inc_headers):
        if h not in ('SA2_CODE', 'SA2_NAME'):
            inc_extra_indices.append(i)
            inc_extra_headers.append(h)

    merged_headers = pop_headers + inc_extra_headers
    merged_rows = []

    for prow in pop_rows:
        code = str(prow[0])
        if code not in inc_by_code:
            continue
        irow = inc_by_code[code]

        has_np = False
        for idx in inc_extra_indices:
            if irow[idx] == 'np':
                has_np = True
                break
        if has_np:
            continue

        merged = list(prow)
        for idx in inc_extra_indices:
            merged.append(irow[idx])
        merged_rows.append(merged)

    return merged_headers, merged_rows


def convert_types(headers, rows):
    """Convert string values to appropriate numeric types."""
    numeric_cols = {'SA2_CODE', 'POPULATION_2023', 'EARNERS', 'MEDIAN_INCOME', 'MEAN_INCOME'}
    numeric_indices = [i for i, h in enumerate(headers) if h in numeric_cols]

    converted = []
    for row in rows:
        new_row = list(row)
        for idx in numeric_indices:
            val = new_row[idx]
            if isinstance(val, str):
                val = val.replace(',', '').strip()
                try:
                    new_row[idx] = int(val)
                except ValueError:
                    try:
                        new_row[idx] = float(val)
                    except ValueError:
                        pass
            elif isinstance(val, float) and val == int(val):
                new_row[idx] = int(val)
        converted.append(new_row)
    return converted


def compute_quartile_boundaries(rows, median_income_idx):
    """Compute quartile boundaries on MEDIAN_INCOME using equal-width range binning.
    Divides the min-max range into 4 equal-width bins."""
    values = []
    for row in rows:
        val = row[median_income_idx]
        if isinstance(val, (int, float)):
            values.append(val)

    min_val = min(values)
    max_val = max(values)
    range_width = (max_val - min_val) / 4

    q1_boundary = min_val + range_width
    q2_boundary = min_val + 2 * range_width
    q3_boundary = min_val + 3 * range_width

    return min_val, q1_boundary, q2_boundary, q3_boundary, max_val


def assign_quartile(value, boundaries):
    """Assign Q1-Q4 label based on equal-width range boundaries."""
    min_val, q1, q2, q3, max_val = boundaries
    if value < q1:
        return 'Q1'
    elif value < q2:
        return 'Q2'
    elif value < q3:
        return 'Q3'
    else:
        return 'Q4'


def add_derived_columns(headers, rows, median_income_idx, earners_idx):
    """Add Quarter and Total columns."""
    boundaries = compute_quartile_boundaries(rows, median_income_idx)

    new_headers = headers + ['Quarter', 'Total']
    new_rows = []
    for row in rows:
        new_row = list(row)
        mi = row[median_income_idx]
        ea = row[earners_idx]
        quarter = assign_quartile(mi, boundaries)
        total = ea * mi if isinstance(ea, (int, float)) and isinstance(mi, (int, float)) else None
        new_row.append(quarter)
        new_row.append(total)
        new_rows.append(new_row)

    return new_headers, new_rows


def write_source_data_sheet(wb, sheet_name, headers, rows):
    """Write the enriched source data to a worksheet."""
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return ws


def create_pivot_cache(wb, source_ws_name, headers, num_data_rows):
    """Create a pivot cache from the source data worksheet."""
    num_cols = len(headers)
    last_col_letter = openpyxl.utils.get_column_letter(num_cols)
    ref = f"A1:{last_col_letter}{num_data_rows + 1}"

    cache_fields = []
    for h in headers:
        cf = CacheField(name=h, numFmtId=0)
        cache_fields.append(cf)

    src = CacheSource(
        type='worksheet',
        worksheetSource=WorksheetSource(
            ref=ref,
            sheet=source_ws_name
        )
    )

    cache = CacheDefinition(
        cacheSource=src,
        cacheFields=cache_fields
    )

    return cache


def create_pivot_table_on_sheet(wb, cache, sheet_name, source_headers,
                                 row_field_name, data_field_name,
                                 aggregation='sum', col_field_name=None,
                                 num_data_rows=0):
    """Create a proper Excel pivot table on a new sheet."""
    ws = wb.create_sheet(sheet_name)

    row_field_idx = source_headers.index(row_field_name)
    data_field_idx = source_headers.index(data_field_name)
    col_field_idx = source_headers.index(col_field_name) if col_field_name else None

    # Create pivot fields - one per source column
    pivot_fields = []
    for i, h in enumerate(source_headers):
        pf = PivotField(showAll=False)
        if i == row_field_idx:
            pf.axis = 'axisRow'
        elif col_field_idx is not None and i == col_field_idx:
            pf.axis = 'axisCol'
        elif i == data_field_idx:
            pf.dataField = True
        pivot_fields.append(pf)

    # Data field
    agg_map = {'sum': 'sum', 'count': 'count'}
    agg_type = agg_map.get(aggregation, 'sum')

    if aggregation == 'count':
        display_name = f'Count of {data_field_name}'
    else:
        display_name = f'Sum of {data_field_name}'

    data_field = DataField(
        name=display_name,
        fld=data_field_idx,
        subtotal=agg_type
    )

    # Row fields
    row_fields = [RowColField(x=row_field_idx)]

    # Column fields
    col_fields = []
    if col_field_idx is not None:
        col_fields = [RowColField(x=col_field_idx)]

    # Location
    loc = Location(ref='A3:B4', firstHeaderRow=1, firstDataRow=1, firstDataCol=1)

    # Create pivot table using TableDefinition
    pivot = TableDefinition(
        name=sheet_name.replace(' ', '_'),
        cacheId=0,
        dataFields=[data_field],
        pivotFields=pivot_fields,
        rowFields=row_fields,
        colFields=col_fields if col_fields else (),
        location=loc,
        dataCaption='Values'
    )

    pivot.cache = cache

    ws._pivots.append(pivot)

    return ws


def build_workbook(pdf_path, xlsx_path, output_path):
    """End-to-end: read inputs, join, create pivot tables, save."""
    # 1. Extract PDF data
    raw_headers, pop_rows = extract_pdf_table(pdf_path)
    pop_headers = fix_truncated_headers(raw_headers)

    # 2. Read income data
    inc_headers, inc_rows = read_income_data(xlsx_path)

    # 3. Join data
    merged_headers, merged_rows = join_data(pop_headers, pop_rows, inc_headers, inc_rows)

    # 4. Convert types
    merged_rows = convert_types(merged_headers, merged_rows)

    # 5. Add derived columns (Quarter, Total)
    median_income_idx = merged_headers.index('MEDIAN_INCOME')
    earners_idx = merged_headers.index('EARNERS')
    enriched_headers, enriched_rows = add_derived_columns(
        merged_headers, merged_rows, median_income_idx, earners_idx
    )

    # 6. Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 7. Write SourceData sheet first (pivot tables reference it)
    source_ws = write_source_data_sheet(wb, 'SourceData', enriched_headers, enriched_rows)
    num_data_rows = len(enriched_rows)

    # 8. Create pivot cache
    cache = create_pivot_cache(wb, 'SourceData', enriched_headers, num_data_rows)

    # 9. Create pivot tables
    create_pivot_table_on_sheet(
        wb, cache, 'Population by State', enriched_headers,
        row_field_name='STATE',
        data_field_name='POPULATION_2023',
        aggregation='sum',
        num_data_rows=num_data_rows
    )

    create_pivot_table_on_sheet(
        wb, cache, 'Earners by State', enriched_headers,
        row_field_name='STATE',
        data_field_name='EARNERS',
        aggregation='sum',
        num_data_rows=num_data_rows
    )

    create_pivot_table_on_sheet(
        wb, cache, 'Regions by State', enriched_headers,
        row_field_name='STATE',
        data_field_name='SA2_CODE',
        aggregation='count',
        num_data_rows=num_data_rows
    )

    create_pivot_table_on_sheet(
        wb, cache, 'State Income Quartile', enriched_headers,
        row_field_name='STATE',
        data_field_name='EARNERS',
        aggregation='sum',
        col_field_name='Quarter',
        num_data_rows=num_data_rows
    )

    # 10. Reorder sheets
    sheet_order = ['Population by State', 'Earners by State', 'Regions by State', 'State Income Quartile', 'SourceData']
    for i, name in enumerate(sheet_order):
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - idx)

    # 11. Save
    wb.save(output_path)
    print(f'Saved workbook to {output_path}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Data rows: {num_data_rows}')
    print(f'Columns: {enriched_headers}')

    return output_path



def recalculate_with_libreoffice(output_path):
    """Recalculate workbook with LibreOffice to render pivot tables."""
    import subprocess
    import shutil
    import os
    
    tmp_dir = '/tmp/recalc'
    os.makedirs(tmp_dir, exist_ok=True)
    
    result = subprocess.run(
        ['libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
         output_path, '--outdir', tmp_dir],
        capture_output=True, text=True, timeout=60
    )
    
    basename = os.path.basename(output_path)
    tmp_file = os.path.join(tmp_dir, basename)
    
    if os.path.exists(tmp_file):
        shutil.copy2(tmp_file, output_path)
        print(f'Recalculated {output_path} with LibreOffice')
        return True
    else:
        print(f'LibreOffice recalculation failed: {result.stderr}')
        return False

def validate_workbook(output_path):
    """Validate the output workbook meets requirements."""
    wb = openpyxl.load_workbook(output_path)
    errors = []

    expected_sheets = ['Population by State', 'Earners by State', 'Regions by State', 'State Income Quartile', 'SourceData']
    for s in expected_sheets:
        if s not in wb.sheetnames:
            errors.append(f'Missing sheet: {s}')

    if 'SourceData' in wb.sheetnames:
        ws = wb['SourceData']
        headers = [c.value for c in ws[1]]
        if 'Quarter' not in headers:
            errors.append('SourceData missing Quarter column')
        if 'Total' not in headers:
            errors.append('SourceData missing Total column')
        if 'POPULATION_2023' not in headers:
            errors.append('SourceData missing POPULATION_2023 column')
        if ws.max_row < 2:
            errors.append('SourceData has no data rows')
        else:
            print(f'SourceData: {ws.max_row - 1} data rows, {ws.max_column} columns')
            print(f'SourceData headers: {headers}')

    for sheet_name in expected_sheets[:4]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if len(ws._pivots) == 0:
                errors.append(f'{sheet_name}: no pivot tables found')
            else:
                pt = ws._pivots[0]
                print(f'{sheet_name}: pivot table "{pt.name}"')
                if hasattr(pt, 'cache') and pt.cache:
                    print(f'  Cache fields: {len(pt.cache.cacheFields)}')
                print(f'  Pivot fields: {len(pt.pivotFields)}')
                print(f'  Row fields: {len(pt.rowFields)}')
                print(f'  Data fields: {len(pt.dataFields)}')
                if pt.colFields:
                    print(f'  Col fields: {len(pt.colFields)}')

    if errors:
        print(f'\nVALIDATION ERRORS:')
        for e in errors:
            print(f'  - {e}')
    else:
        print(f'\nVALIDATION PASSED')

    wb.close()
    return errors
