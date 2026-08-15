"""Terminus-2 agent with host intervention and verification.

Extends HarborTerminus2WithSkills to add:
- Pre-task: inject meta-skills (skill-creator) into container
- Intra-task: host intervention loop (independent verifier, progress checklist, GT oracle)
- Post-task: export verifier scripts, read evolution summary, write detailed logs
"""

from __future__ import annotations

import asyncio
import ast
import base64
import hashlib
import importlib
import json
import logging
import math
import os
import re
import shlex
import shutil
import tempfile
import time
import tomllib
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from harbor.environments.base import BaseEnvironment
from harbor.environments.docker.docker import DockerEnvironment
from harbor.models.agent.context import AgentContext
from harbor.models.task.config import TaskConfig
from harbor.models.trial.paths import TrialPaths
from harbor.utils.env import resolve_env_vars

from libs.terminus_agent.evolution.independent_verifier import IndependentVerifier
from libs.terminus_agent.evolution.combinatorial_budget import (
    combinatorial_search_budget_issue,
    referenced_python_scripts,
)
from libs.terminus_agent.evolution.models import VerificationResult
from libs.terminus_agent.evolution.self_verifier import SelfVerifier
from libs.terminus_agent.evolution.skill_information_boundary import (
    audit_evolved_skill_directory,
)
from libs.terminus_agent.evolution import skill_schema
from libs.terminus_agent.evolution.skill_schema import validate_skill_directory
from libs.terminus_agent.agents.claude_code_vertex import ClaudeCodeProviderError

from .harbor_terminus_2_skills import Command, EpisodeExitResult, HarborTerminus2WithSkills
from .skill_docs import is_shell_warning

logger = logging.getLogger(__name__)

_TRACE_HEREDOC_RE = re.compile(
    r"<<-?\s*['\"]?([A-Za-z_][A-Za-z0-9_]*)['\"]?[^\n]*\n"
    r"(.*?)\n\1(?:\s*(?:\n|$))",
    re.DOTALL,
)


def _trace_python_sources(command: str) -> list[str]:
    """Extract inline Python owned by the fresh oracle from a Bash command.

    Claude commonly uses both heredocs and ``python -c``.  Attribution gates
    must inspect both forms; otherwise the same caller-owned instance search
    can evade the gate solely by changing its shell spelling.
    """

    sources = [match.group(2) for match in _TRACE_HEREDOC_RE.finditer(command)]
    try:
        tokens = shlex.split(command, posix=True)
    except ValueError:
        return sources

    for index, token in enumerate(tokens[:-2]):
        executable = Path(token).name
        if (
            re.fullmatch(r"python(?:\d+(?:\.\d+)?)?", executable)
            and tokens[index + 1] == "-c"
        ):
            sources.append(tokens[index + 2])
    return sources


_INSTANCE_TARGET_NAME_RE = re.compile(
    r"(?i)(?:extra|target|identity|identities|author|affiliat|venue|grant|"
    r"redact|entity|entities|name_?list|suspect|analyte|candidate|coord|"
    r"position|place(?:d|ment)?|district|city_?center|"
    r"(?:unit|conversion)_?(?:map|mapping|registry|ranges?)|"
    r"(?:feature|column)_?(?:map|mapping|registry|ranges?))"
)
_LOW_LEVEL_SKILL_INSPECTION_NAME_RE = re.compile(
    r"(?i)^(?:discover|extract|parse|read|scan|find|locate|collect|enumerate|"
    r"inspect|analy[sz]e|score|calc(?:ulate)?|compute|load|get|list|"
    r"can_?place)(?:_|$)"
)
_LOW_LEVEL_SKILL_MUTATION_NAME_RE = re.compile(
    r"(?i)^(?:add|insert|replace|remove|delete|exclude|set|update|write|"
    r"patch|migrate|fix|rewrite|modify|create)(?:_|$)"
)
_DIRECT_ENVIRONMENT_MUTATION_RE = re.compile(
    r"(?i)(?:^|[;&|]\s*)"
    r"(?:python\d*(?:\.\d+)?\s+-m\s+)?pip\s+install\b|"
    r"(?:^|[;&|]\s*)apt(?:-get)?\s+install\b|"
    r"(?:^|[;&|]\s*)(?:conda|mamba|micromamba)\s+install\b|"
    r"(?:^|[;&|]\s*)(?:npm|pnpm|yarn)\s+(?:install|add)\b|"
    r"(?:^|[;&|]\s*)gem\s+install\b|"
    r"(?:^|[;&|]\s*)cargo\s+add\b"
)
_DIRECT_PATCH_MUTATION_RE = re.compile(
    r"(?i)(?:^|[;&|]\s*)(?:apply_patch\b|git\s+apply\b|patch\s+(?:-[^\s]+\s+)*<|"
    r"sed\s+-i\b)"
)
_TRACE_EXACT_REPLACEMENT_ASSIGNMENT_RE = re.compile(
    r"(?is)\b(?P<side>old|new)(?:_[A-Za-z0-9_]+)?\s*=\s*"
    r"(?P<quote>['\"])(?P<value>[^'\"\n]{4,240})(?P=quote)"
)
_TRACE_REPLACEMENT_PRIMITIVE_RE = re.compile(
    r"(?i)\b(?:replace|rewrite|patch|migrate|relocat|transform|update|fix)"
    r"[A-Za-z0-9_]*\s*\("
)
_GIT_HISTORY_SUBCOMMAND_RE = re.compile(
    r"(?i)(?:^|[;&|]\s*|\s)git\s+"
    r"(?P<subcommand>show|log|reflog|rev-list|cat-file|format-patch)\b"
)
_GIT_DIFF_COMMAND_RE = re.compile(
    r"(?i)(?:^|[;&|]\s*|\s)git\s+diff\b(?P<arguments>[^\n;&|]*)"
)
_GIT_HISTORICAL_DIFF_OPERAND_RE = re.compile(
    r"(?i)(?:\.\.|[0-9a-f]{7,40}(?=[:.\s]|$)|HEAD(?:~\d*|\^+|@\{)|"
    r"refs/(?:heads|remotes|tags)/|(?:origin|upstream)/[A-Za-z0-9._/-]+)"
)
_HIDDEN_EVALUATOR_ACCESS_RE = re.compile(
    r"(?ix)(?:"
    r"/(?:root|app)/verifier(?:/|\b)|"
    r"(?<![A-Za-z0-9_./-])/tests(?:/|\b)|"
    r"(?:^|[/\s'\"=])test_outputs\.py(?:\b|$)|"
    r"(?:^|[/\s'\"=])(?:reference_solution|ground_truth|golden_answer)"
    r"(?:s)?(?:/|\b)"
    r")"
)

_CALLER_ARTIFACT_WRITE_METHODS = {
    "dump",
    "imwrite",
    "imsave",
    "save",
    "savefig",
    "to_csv",
    "to_excel",
    "to_feather",
    "to_json",
    "to_parquet",
    "to_pickle",
    "write_audio",
    "write_bytes",
    "write_text",
    "write_video",
}
_CALLER_ARTIFACT_SUFFIXES = {
    ".csv", ".docx", ".json", ".mp3", ".mp4", ".ods", ".parquet",
    ".patch", ".pdf", ".png", ".pptx", ".tsv", ".wav", ".webm",
    ".xlsx", ".xml", ".yaml", ".yml",
}
_GENERIC_IDENTITY_AUDIT_TERMS = {
    "acknowledg", "acknowledgement", "acknowledgment", "affiliation",
    "affiliations", "arxiv", "author", "authors", "conference", "doi",
    "email", "emails", "funding", "github", "grant", "grateful",
    "institution", "institutions", "metadata", "reference", "references",
    "support", "supported", "thank", "thanks", "university", "venue",
    "venues", "work", "workshop",
}


def _evo_print(msg: str) -> None:
    """Print evolution status to stdout, silently ignoring BrokenPipeError."""
    try:
        print(msg, flush=True)
    except BrokenPipeError:
        pass


class HarborTerminus2Evolution(HarborTerminus2WithSkills):
    """Terminus-2 agent with host intervention and verification."""

    _IDLE_EPISODE_LIMIT: int = 3  # consecutive empty-command episodes before forced intervention
    # Opus commonly needs more than ten JSON/terminal turns to inspect a large
    # workbook or dataset, form a reusable Skill, and execute it. Ten turns
    # caused actively progressing agents to be sent to the verifier before an
    # Skill or output existed. The pre-surrogate Skill gate is the structural
    # backstop; this limit is only for genuinely overlong active runs.
    _STALE_EPISODE_LIMIT: int = 30

    @staticmethod
    def _hidden_evaluator_access_issue(command: str) -> str | None:
        """Reject evolution-agent access to protected evaluator artifacts.

        The evolution container may expose task tests through an implementation
        detail such as ``/root/verifier``.  Those files are not part of the
        unchanged instruction, background document, or supplied inputs.  Reading them lets
        the candidate serialize evaluator expectations into the treatment and
        invalidates the background document/Skill comparison.
        """
        if _HIDDEN_EVALUATOR_ACCESS_RE.search(command):
            return "attempted to inspect or execute protected evaluator artifacts"
        return None

    async def _execute_commands(
        self,
        environment: BaseEnvironment,
        commands: list[Command],
    ) -> str:
        """Reject aggregate combinatorial runaways before they consume a worker.

        A per-command wall-clock timeout remains enabled, but it is too late for
        generated full-map Cartesian searches: the terminal can be occupied and
        lose a complete evolution turn.  Inspect both inline/heredoc Python and
        any directly invoked absolute ``.py`` file.  The rejection contains no
        task answer or evaluator information.
        """
        for command in commands:
            evaluator_issue = self._hidden_evaluator_access_issue(command.keystrokes)
            if evaluator_issue:
                logger.warning(
                    "Evolution command rejected by evaluator-boundary gate: %s",
                    evaluator_issue,
                )
                return (
                    "New Terminal Output:\n"
                    "COMMAND REJECTED BY EVALUATOR INFORMATION-BOUNDARY GATE: "
                    f"{evaluator_issue}. Work only from the unchanged task "
                    "instruction, public background document, supplied inputs, current "
                    "implementation, and ordinary runtime diagnostics."
                )

            public_contract = getattr(self, "_instruction", "")
            history_operation = self._manual_repository_history_oracle(
                command.keystrokes,
                "",
            )
            forbidden_history = self._manual_repository_history_oracle(
                command.keystrokes,
                public_contract,
            )
            if forbidden_history:
                logger.warning(
                    "Evolution command rejected by repository-history boundary gate"
                )
                return (
                    "New Terminal Output:\n"
                    "COMMAND REJECTED BY REPOSITORY-HISTORY INFORMATION-BOUNDARY "
                    "GATE: historical commits and revisions cannot be used as a "
                    "reference solution. Diagnose from the current checkout, public "
                    "background document, ordinary build/runtime output, and your own same-run "
                    "experiments."
                )
            if history_operation:
                # A public contract can explicitly prescribe a history operation,
                # so the command may execute, but that run no longer qualifies for
                # the narrow fresh-source-knowledge overlap exception.
                self._lineage_command_gate_clean = False

            referenced_sources: list[str] = []
            for path in referenced_python_scripts(command.keystrokes):
                result = await environment.exec(
                    command=f"head -c 1000000 -- {shlex.quote(path)} 2>/dev/null || true",
                    timeout_sec=15,
                )
                if result.stdout:
                    referenced_sources.append(result.stdout)

            issue = combinatorial_search_budget_issue(
                command.keystrokes,
                referenced_sources=tuple(referenced_sources),
            )
            if issue:
                # Search-shape findings are operational advice, not a validity
                # boundary.  The command timeout still provides a fail-closed
                # resource limit, while canonical GT decides whether the
                # resulting Skill solves the task.  Keep protected evaluator and
                # repository-history access above as hard rejections.
                logger.warning(
                    "Evolution command flagged by search-budget advisory; "
                    "executing under the command timeout: %s",
                    issue,
                )

        return await super()._execute_commands(environment, commands)

    @staticmethod
    def _safe_gt_failure_categories(gt_result: dict[str, Any] | None) -> list[str]:
        """Reduce failed test names to broad, answer-free diagnostic dimensions.

        Raw names, messages, expected values, thresholds, and reference outputs
        remain host-only.  The returned labels describe ordinary engineering
        concerns already inferable from task instructions and public background documents.
        """
        if not gt_result:
            return []

        mappings = (
            (("file", "schema", "format", "column", "field", "path", "output"),
             "artifact/interface compliance"),
            (("build", "compile", "install", "import", "runtime", "execute"),
             "build or runtime execution"),
            (("recall", "precision", "semantic", "content", "detect", "classif", "coverage", "pick",
              "color", "match", "object", "frame"),
             "semantic selection and coverage"),
            (("duration", "timestamp", "segment", "audio", "video", "waveform", "sync", "correspond"),
             "temporal or media consistency"),
            (("formula", "numeric", "unit", "total", "math", "value", "calculation"),
             "numeric, formula, or unit consistency"),
            (("security", "exploit", "vulnerab", "legitimate", "bypass"),
             "security and robustness behavior"),
            (("constraint", "quality", "valid", "invariant", "balance", "overlap", "depth"),
             "constraint and invariant compliance"),
        )
        categories: list[str] = []
        for detail in gt_result.get("test_details") or []:
            if str(detail.get("status", "")).upper() not in {"FAILED", "FAIL", "ERROR"}:
                continue
            normalized_name = re.sub(r"[^a-z0-9]+", " ", str(detail.get("name", "")).lower())
            for tokens, label in mappings:
                if any(token in normalized_name for token in tokens) and label not in categories:
                    categories.append(label)
        return categories

    @staticmethod
    def _safe_gt_failed_output_fields(
        gt_result: dict[str, Any] | None,
        instruction: str,
    ) -> list[str]:
        """Expose only instruction-declared output field names from GT failures.

        A field such as ``q2_answer`` is part of the public output contract, not
        a current-instance answer.  Returning that field lets evolution focus on
        the failing branch while expected/actual values, assertion text, test
        names, thresholds, paths, and evaluator code remain host-only.

        Restrict candidates to dictionary lookups in the raw verifier output and
        require an exact occurrence in the unchanged instruction.  This prevents
        row IDs, filenames, or hidden evaluator-only labels from becoming
        treatment input.
        """
        if not gt_result or not instruction:
            return []

        raw_output = str(gt_result.get("raw_output") or "")
        if not raw_output:
            return []
        failure_lines = "\n".join(
            line
            for line in raw_output.splitlines()
            if re.match(r"^\s*(?:>|E\s+)", line)
        )
        candidates = re.findall(
            r"(?:answers|answer|data|result|output)\s*\[\s*['\"]"
            r"([A-Za-z_][A-Za-z0-9_]{0,63})['\"]\s*\]",
            failure_lines,
        )
        safe: list[str] = []
        for field in candidates:
            if field in instruction and field not in safe:
                safe.append(field)
            if len(safe) >= 8:
                break
        return safe

    @staticmethod
    def _safe_gt_refinement_guidance(gt_result: dict[str, Any] | None) -> list[str]:
        """Return answer-free process guidance for GT attribution failures.

        Canonical test details remain host-only, but attribution failures are
        properties of the transfer protocol itself.  Telling the evolution
        agent that its immutable helper was bypassed does not reveal a task
        answer and prevents a futile loop in which it repairs only the current
        working tree while leaving the reusable Skill unchanged.
        """
        if not gt_result:
            return []

        reasons = {
            str(reason)
            for reason in (gt_result.get("failure_reasons") or [])
        }
        guidance: list[str] = []
        if "Clean evaluation bypassed the immutable evolved Skill" in reasons:
            guidance.append(
                "Attribution failed: after invoking the Skill, the fresh agent "
                "still needed a manual artifact or source mutation. Update the "
                "immutable evo-* Skill's end-to-end helper so it performs the "
                "diagnosis, treatment, and validation itself; regenerate the "
                "current result only by invoking that updated helper."
            )
        if "Claude Code did not invoke an injected evolved Skill" in reasons:
            guidance.append(
                "Transfer failed because the fresh agent did not invoke the "
                "evolved Skill. Make the Skill description, usage, and single "
                "end-to-end entry point unambiguous without adding any current-"
                "instance answer."
            )
        if "Clean evaluation modified the immutable evolved Skill" in reasons:
            guidance.append(
                "Transfer failed because the fresh evaluation attempted to "
                "change the immutable Skill. Move all adaptation into the "
                "evolution phase and expose a runtime-parameterized helper that "
                "does not rewrite its own Skill sources."
            )
        return guidance

    @staticmethod
    def _build_static_gate_feedback(
        gate: str,
        detailed_issues: list[Any] | tuple[Any, ...] | None = None,
    ) -> str:
        """Return fixed, answer-free guidance for programmatic preflight gates.

        ``detailed_issues`` is intentionally never rendered.  Static scanners
        can identify exact files, lines, current literals, paths, schema keys,
        or cutoffs; those findings are useful host-side audit evidence but are
        not legitimate evolution feedback.  Returning them would turn the gate
        into an instance-specific repair oracle and contaminate every later
        Skill in the lineage.
        """
        del detailed_issues
        templates = {
            "missing_frontmatter": (
                "SKILL DISCOVERY FAILURE: One or more evolved Skills are not "
                "discoverable because their metadata is invalid. No surrogate "
                "or GT evaluation was run.\n\n"
                "Review every evo-* SKILL.md against the public Skill metadata "
                "format. Ensure each begins with valid YAML frontmatter containing "
                "a reusable name and description. Validate discovery generically, "
                "then signal task_complete again."
            ),
            "transfer_api": (
                "SKILL TRANSFER API FAILURE: A fresh agent cannot execute the Skill "
                "exactly as documented. No surrogate or GT evaluation was run.\n\n"
                "Review the public transfer contract generically. Provide a thin "
                "end-to-end entry point plus an output validator, ensure the documented "
                "example actually imports and calls existing Skill code, and keep all "
                "reusable orchestration inside the Skill. The fresh caller may supply "
                "only runtime paths and parameters explicitly allowed by the instruction. "
                "Regenerate through that API and signal task_complete again."
            ),
            "hard_property": (
                "SKILL HARD-PROPERTY EVIDENCE FAILURE: The Skill can emit an "
                "affirmative property without proving it for the same runtime-selected "
                "entity. No surrogate or GT evaluation was run.\n\n"
                "Use a general positive/negative/unknown evidence model, require "
                "same-entity public runtime evidence before emitting an affirmative "
                "property, and fail closed on missing or unrelated evidence. Reopen and "
                "validate the generated artifact through the Skill, then signal "
                "task_complete again."
            ),
            "information_boundary": (
                "SKILL INFORMATION-BOUNDARY FAILURE: The evolved Skill is not "
                "transferable because it contains current-instance material or "
                "unjustified fixed policy. No surrogate or GT evaluation was run.\n\n"
                "Audit the Skill generically for embedded input-derived literals, "
                "instance manifests, layout assumptions, answer tables, fixed numeric "
                "sequences, and arbitrary decision policy. Discover or derive runtime "
                "facts from inputs, expose legitimate policy only through public "
                "configuration, and use unrelated synthetic examples. Do not weaken the "
                "output contract. Regenerate through the revised Skill and signal "
                "task_complete again."
            ),
            "spreadsheet_recalculation": (
                "SPREADSHEET RECALCULATION FAILURE: The declared spreadsheet output "
                "did not survive an independent structural recalculation check. No "
                "surrogate or GT evaluation was run.\n\n"
                "Review the reusable workbook-construction logic, formulas, dependency "
                "references, label discovery, and row/column orientation using public "
                "runtime or unrelated synthetic evidence. Regenerate through the "
                "documented end-to-end Skill entry point and strengthen its generic "
                "reopen validator before signaling task_complete again."
            ),
        }
        try:
            return templates[gate]
        except KeyError as exc:
            raise ValueError(f"unknown static gate: {gate}") from exc

    @staticmethod
    def _build_skill_schema_feedback(issues: list[str]) -> str:
        """Return precise schema-only feedback that is safe to expose.

        Unlike GT or information-boundary findings, Skill metadata errors do
        not reveal task answers.  Showing the exact field and rule is both safe
        and necessary for the evolution Agent to repair its own artifact.
        """

        rendered = "\n".join(f"- {issue}" for issue in issues)
        return (
            "SKILL SCHEMA VALIDATION FAILED: No surrogate or GT evaluation was "
            "run, and this attempt did not consume a GT iteration.\n\n"
            "Exact schema errors:\n"
            f"{rendered}\n\n"
            "Load `skill-creator`, repair the reported SKILL.md metadata yourself, "
            "run its schema validator, and signal task_complete again. Preserve the "
            "reusable Skill logic; do not add task-instance answers while repairing "
            "metadata."
        )

    @staticmethod
    def _log_static_gate_details(gate: str, detailed_issues: list[Any]) -> None:
        """Keep exact scanner evidence in host logs, never in agent feedback."""
        for issue in detailed_issues:
            logger.warning("Static gate detail [%s] (host-only): %s", gate, issue)

    @staticmethod
    def _declared_spreadsheet_outputs(instruction: str) -> list[str]:
        """Extract spreadsheet paths only from output/save clauses."""
        paths: list[str] = []
        for line in instruction.splitlines():
            if not re.search(r"\b(?:output|save|write|deliver)\b", line, re.I):
                continue
            for match in re.findall(
                r"(?:/root/[A-Za-z0-9_./-]+|[A-Za-z0-9_.-]+)\.xlsx\b",
                line,
                re.I,
            ):
                path = match if match.startswith("/root/") else f"/root/{match}"
                if path not in paths:
                    paths.append(path)
        return paths

    async def _spreadsheet_recalculation_issues(
        self,
        environment: BaseEnvironment,
    ) -> list[str]:
        """Recalculate declared spreadsheet outputs and reject cached formula errors."""
        output_paths = self._declared_spreadsheet_outputs(self._instruction)
        if not output_paths:
            return []

        checker = r'''
import json
import base64
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

source = Path(sys.argv[1])
instruction = base64.b64decode(sys.argv[2]).decode("utf-8", errors="replace")
if not source.is_file():
    print(json.dumps({"error": f"declared output is missing: {source}"}))
    raise SystemExit(2)

# A recalculation engine is relevant only when the deliverable actually contains
# worksheet formulas. Inspect the OOXML directly first: this avoids walking an
# inflated worksheet used range and lets formula-free pivot/data workbooks pass
# to semantic verification even in minimal images without LibreOffice.
try:
    with zipfile.ZipFile(source) as archive:
        formula_count = sum(
            archive.read(name).count(b"<f>")
            + archive.read(name).count(b"<f ")
            for name in archive.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        )
except (OSError, zipfile.BadZipFile) as exc:
    print(json.dumps({"error": f"invalid spreadsheet package: {exc}"}))
    raise SystemExit(2)

if formula_count == 0:
    print(json.dumps({"formula_count": 0, "errors": []}))
    raise SystemExit(0)

office = shutil.which("libreoffice") or shutil.which("soffice")
if not office:
    print(json.dumps({"error": "no compatible spreadsheet recalculation engine"}))
    raise SystemExit(3)

with tempfile.TemporaryDirectory(prefix="skill-sheet-gate-") as temporary:
    root = Path(temporary)
    input_dir = root / "input"
    output_dir = root / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    staged = input_dir / source.name
    shutil.copy2(source, staged)
    completed = subprocess.run(
        [office, "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(staged)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        timeout=300,
    )
    recalculated = output_dir / source.name
    if completed.returncode != 0 or not recalculated.is_file():
        print(json.dumps({
            "error": "spreadsheet recalculation failed",
            "engine_output": completed.stdout[-1000:],
        }))
        raise SystemExit(4)

    formulas = load_workbook(source, data_only=False, read_only=False)
    values = load_workbook(recalculated, data_only=True, read_only=False)
    excel_errors = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NUM!", "#N/A", "#NULL!")
    errors = []
    recalculated_formula_count = 0

    # Public, answer-free semantic checks.  A one-row INDEX array is horizontal:
    # MATCH therefore selects its column, not its row.  Some spreadsheet engines
    # accept INDEX(horizontal_range, MATCH(...)) as a one-dimensional extension,
    # while others correctly return #REF! for every match after the first.  Catch
    # the ambiguous form syntactically so the evolved Skill must emit portable
    # formulas without relying on one engine's cached values.
    horizontal_index_with_match_row = re.compile(
        r"INDEX\s*\(\s*"
        r"(?:(?:'[^']+'|[A-Za-z_][A-Za-z0-9_. ]*)!)?"
        r"\$?([A-Z]{1,3})\$?(\d+)\s*:\s*"
        r"(?:(?:'[^']+'|[A-Za-z_][A-Za-z0-9_. ]*)!)?"
        r"\$?([A-Z]{1,3})\$?(\d+)\s*,\s*MATCH\s*\(",
        re.I,
    )

    # When the public instruction calls the exposure a near-term swing, an
    # annualized display row is not an interchangeable risk horizon.  This gate
    # discovers row roles from workbook labels and checks dependencies only; it
    # contains no workbook coordinates, current values, expected entities, or
    # evaluator thresholds.
    near_term_exposure = bool(
        re.search(r"near[- ]term", instruction, re.I)
        and re.search(r"(?:exposure|valuation)", instruction, re.I)
    )
    explicitly_annual_exposure = bool(
        re.search(
            r"(?:annual(?:ized|ised)?\b(?:\W+\w+){0,6}\W+(?:exposure|valuation)"
            r"|(?:exposure|valuation)\b(?:\W+\w+){0,6}\W+annual(?:ized|ised)?)",
            instruction,
            re.I,
        )
    )

    for formula_sheet in formulas.worksheets:
        if formula_sheet.title not in values.sheetnames:
            errors.append(f"missing sheet after recalculation: {formula_sheet.title}")
            continue
        value_sheet = values[formula_sheet.title]

        row_labels = {}
        for row in formula_sheet.iter_rows():
            labels = [
                cell.value.strip()
                for cell in row
                if isinstance(cell.value, str)
                and cell.value.strip()
                and not cell.value.startswith("=")
            ]
            if labels:
                row_labels[row[0].row] = " | ".join(labels)
        annualized_rows = {
            row_number
            for row_number, label in row_labels.items()
            if re.search(r"annual(?:ized|ised)", label, re.I)
            and re.search(r"volatil|risk|sigma", label, re.I)
        }

        for row in formula_sheet.iter_rows():
            row_label = row_labels.get(row[0].row, "")
            is_labeled_output_row = bool(row_label)
            is_exposure_row = bool(
                re.search(r"(?:exposure|valuation)", row_label, re.I)
            )
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                recalculated_formula_count += 1
                formula = cell.value

                ambiguous = horizontal_index_with_match_row.search(formula)
                if ambiguous and ambiguous.group(2) == ambiguous.group(4):
                    issue = (
                        f"{formula_sheet.title}: a horizontal INDEX formula uses MATCH "
                        "as the row selector; use an explicit row selector and bind "
                        "MATCH to the column dimension (or use a label-derived XLOOKUP)"
                    )
                    if issue not in errors:
                        errors.append(issue)

                if (
                    near_term_exposure
                    and not explicitly_annual_exposure
                    and is_exposure_row
                    and annualized_rows
                ):
                    same_sheet_reference_rows = {
                        int(match.group(1))
                        for match in re.finditer(
                            r"(?<![!A-Za-z0-9_])\$?[A-Z]{1,3}\$?(\d+)",
                            formula,
                            re.I,
                        )
                    }
                    referenced_annualized = sorted(
                        same_sheet_reference_rows & annualized_rows
                    )
                    if referenced_annualized:
                        labels = ", ".join(
                            row_labels[row_number]
                            for row_number in referenced_annualized
                        )
                        issue = (
                            f"{formula_sheet.title}: a near-term exposure formula "
                            f"references an annualized display row ({labels}); bind "
                            "the risk dependency to the instruction-selected horizon "
                            "using runtime-discovered labels"
                        )
                        if issue not in errors:
                            errors.append(issue)

                cached = value_sheet[cell.coordinate].value
                cached_cell = value_sheet[cell.coordinate]
                if cached_cell.data_type == "e" or (
                    isinstance(cached, str)
                    and any(token in cached.upper() for token in excel_errors)
                ):
                    errors.append(f"{formula_sheet.title}!{cell.coordinate}: {cached}")
                elif cached is None and is_labeled_output_row:
                    errors.append(
                        f"{formula_sheet.title}: a formula in labeled output row "
                        f"'{row_label}' has no cached value after recalculation"
                    )
                if len(errors) >= 50:
                    break
            if len(errors) >= 50:
                break
        if len(errors) >= 50:
            break
    formulas.close()
    values.close()
    print(json.dumps({"formula_count": recalculated_formula_count, "errors": errors}))
    raise SystemExit(1 if errors else 0)
'''
        encoded = base64.b64encode(checker.encode("utf-8")).decode("ascii")
        instruction_encoded = base64.b64encode(
            self._instruction.encode("utf-8")
        ).decode("ascii")
        issues: list[str] = []
        for output_path in output_paths:
            result = await environment.exec(
                command=(
                    f"echo {shlex.quote(encoded)} | base64 -d > /tmp/skill_sheet_gate.py "
                    f"&& python3 /tmp/skill_sheet_gate.py {shlex.quote(output_path)} "
                    f"{shlex.quote(instruction_encoded)}"
                ),
                timeout_sec=420,
            )
            if result.return_code != 0:
                detail = (result.stdout or "spreadsheet gate failed").strip()
                issues.append(f"{output_path}: {detail[-3000:]}")
        return issues

    def __init__(
        self,
        logs_dir: Path,
        model_name: str | None = None,
        sequential_run: int = 1,
        gt_oracle_model: str | None = None,
        max_host_interventions: int = 5,
        independent_verifier_model: str | None = None,
        gt_oracle_agent: str | None = None,
        timeout_multiplier: float = 1.0,
        skip_surrogate_verifier: bool = False,
        **kwargs,
    ):
        super().__init__(logs_dir=logs_dir, model_name=model_name, timeout_multiplier=timeout_multiplier, **kwargs)
        self._skip_surrogate_verifier = skip_surrogate_verifier

        self._sequential_run = sequential_run
        self._gt_oracle_model = gt_oracle_model
        # The release-standard oracle is a fresh Agent with the background document
        # physically absent and only the evolved Skill injected. Paired
        # background document-plus-Skill runs remain available only as explicit ablations.
        self._gt_oracle_agent = gt_oracle_agent or "claude-code-skill-only"

        # Evolution components
        self._verifier = SelfVerifier()  # Used to read pytest results after independent verifier runs
        self._independent_verifier = IndependentVerifier(
            model_name=independent_verifier_model or model_name or "",
            temperature=0.3,
        )

        # Runtime state
        self._environment: BaseEnvironment | None = None
        self._instruction: str = ""
        self._meta_skills_injected: list[str] = []
        self._host_evolved_skills_injected: list[str] = []

        # Host intervention state (reset at start of each run())
        self._host_intervention_count: int = 0  # only surrogate_pass increments this
        self._max_host_interventions: int = max_host_interventions
        self._gt_infrastructure_retry_count: int = 0
        self._max_gt_infrastructure_retries: int = 5
        self._surrogate_retry_count: int = 0  # surrogate_fail / verifier errors
        self._max_surrogate_retries: int = 15  # fixed cap for surrogate retries
        self._exit_reason: str = ""
        self._intervention_history: list[dict] = []
        self._pre_existing_skills: set[str] = set()
        self._fresh_source_knowledge_eligible: bool = False
        self._lineage_command_gate_clean: bool = True
        self._information_boundary_validated: bool = False
        self._last_information_boundary_issues: list[dict[str, Any]] = []
        self._initial_container_source_identifiers: dict[str, str] | None = None
        self._run_id: str = ""
        self._best_gt_snapshot: dict | None = None  # best GT skills snapshot for rollback
        self._surrogate_tests_locked: bool = False  # lock surrogate tests after first valid generation
        self._checklist_fail_count: int = 0  # consecutive checklist failures (enforce once, then advisory)
        self._skill_schema_fix_count: int = 0  # consecutive schema-only repairs; never consume GT budget
        self._max_skill_schema_fixes: int = 2
        self._skill_schema_validation_history: list[dict[str, Any]] = []
        self._missing_skill_fix_count: int = 0  # pre-GT structural fixes; never consume GT budget
        self._skill_boundary_fix_count: int = 0  # hardcoding fixes; never consume GT budget
        self._adversarial_surrogate_recheck: bool = False  # set after a hidden-evaluator rejection

        # Idle detection state
        self._consecutive_idle_episodes: int = 0  # consecutive episodes with no commands
        self._episodes_since_last_complete: int = 0  # episodes since last task_complete signal

    @staticmethod
    def name() -> str:
        return "terminus-2-evolution"

    def version(self) -> str | None:
        return "2.3.0"

    @staticmethod
    def _fresh_run_source_knowledge_eligibility(
        pre_existing_skills: set[str],
        host_evolved_skills_injected: list[str],
    ) -> bool:
        """Attest that this independent run did not begin with a Skill seed."""
        return (
            not host_evolved_skills_injected
            and not any(name.startswith("evo-") for name in pre_existing_skills)
        )

    def _allow_fresh_run_source_knowledge(self) -> bool:
        """Return the frozen no-seed attestation while command lineage is clean."""
        return (
            self._fresh_source_knowledge_eligible
            and self._lineage_command_gate_clean
        )

    @staticmethod
    def _skill_file_hashes(skill_root: Path | None) -> dict[str, str]:
        if skill_root is None or not skill_root.exists():
            return {}
        hashes: dict[str, str] = {}
        for path in sorted(skill_root.rglob("*")):
            if path.is_file():
                hashes[str(path.relative_to(skill_root))] = hashlib.sha256(
                    path.read_bytes()
                ).hexdigest()
        return hashes

    def _write_source_knowledge_provenance(
        self,
        *,
        status: str,
        skill_root: Path | None = None,
        issues: list[Any] | None = None,
    ) -> None:
        """Persist the evidence that scopes same-run source knowledge safely."""
        issue_payload = [
            {
                "kind": issue.kind,
                "file": issue.file,
                "line": issue.line,
                "evidence": issue.evidence,
            }
            for issue in (issues or [])
        ]
        payload = {
            "run_id": self._run_id,
            "status": status,
            "initial_skills": sorted(self._pre_existing_skills),
            "initial_evolved_skills": sorted(
                name
                for name in self._pre_existing_skills
                if name.startswith("evo-")
            ),
            "host_evolved_skills_injected": sorted(
                self._host_evolved_skills_injected
            ),
            "fresh_no_seed_eligible": self._fresh_source_knowledge_eligible,
            "lineage_command_gate_clean": self._lineage_command_gate_clean,
            "allow_fresh_run_source_knowledge": (
                self._allow_fresh_run_source_knowledge()
            ),
            "repository_history_gate": "execute-time-default-deny",
            "boundary_issues": issue_payload,
            "skill_files_sha256": self._skill_file_hashes(skill_root),
        }
        try:
            self.logs_dir.mkdir(parents=True, exist_ok=True)
            (self.logs_dir / "source-knowledge-provenance.json").write_text(
                json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )
        except OSError as exc:
            logger.warning("Failed to write source-knowledge provenance: %s", exc)

    async def _audit_exported_evolved_skills(
        self,
        environment: BaseEnvironment,
        evolved_skills_dir: Path,
        *,
        stage: str,
    ) -> list[Any]:
        """Run one provenance-aware boundary audit and retain its evidence."""
        task_dir = Path(getattr(environment, "environment_dir", "")).parent
        # Source-identifier provenance must describe the environment presented
        # to the agent, not files the agent has just generated.  Re-scanning
        # /app after an episode previously indexed task outputs such as
        # /app/solution.py and then falsely rejected a Skill for sharing the
        # helper names it had itself used to create that output.
        container_source_identifiers = (
            dict(self._initial_container_source_identifiers)
            if self._initial_container_source_identifiers is not None
            else await self._container_current_source_identifiers(environment)
        )
        issues = audit_evolved_skill_directory(
            evolved_skills_dir,
            task_dir,
            self._instruction,
            current_source_identifiers=container_source_identifiers,
            allow_fresh_run_source_knowledge=(
                self._allow_fresh_run_source_knowledge()
            ),
        )
        self._information_boundary_validated = not issues
        self._last_information_boundary_issues = [
            {
                "kind": issue.kind,
                "file": issue.file,
                "line": issue.line,
                "evidence": issue.evidence,
            }
            for issue in issues
        ]
        self._write_source_knowledge_provenance(
            status=(f"{stage}_passed" if not issues else f"{stage}_blocked"),
            skill_root=evolved_skills_dir,
            issues=issues,
        )
        return issues

    def _get_prompt_template_path(self) -> Path:
        """Return the evolution-specific prompt template for JSON parser, else delegate to parent."""
        if self._parser_name != "json":
            return super()._get_prompt_template_path()
        root = os.environ.get("SKILLSBENCH_ROOT")
        candidates = [Path(root)] if root else []
        candidates.append(Path.cwd())
        for c in candidates:
            p = c / "libs/terminus_agent/agents/prompt-templates/terminus-evolution-json.txt"
            if p.exists():
                return p
        return Path(__file__).resolve().parent.parent / "prompt-templates" / "terminus-evolution-json.txt"

    # ------------------------------------------------------------------
    # Setup
    # ------------------------------------------------------------------

    def _resolve_evolution_dir(self) -> Path:
        """Return the ``environment/.evolution/`` directory for storing evolution artifacts."""
        env_dir = getattr(self._environment, "environment_dir", None)
        if env_dir:
            return Path(env_dir) / ".evolution"
        return self.logs_dir / ".evolution"

    async def setup(self, environment: BaseEnvironment) -> None:
        await super().setup(environment)
        self._environment = environment

        # Inject meta-skills into container
        self._meta_skills_injected = await self._inject_meta_skills(environment)

        # A continuation run stores its reusable evo-* packages in the task's
        # host environment directory.  Fresh Docker containers do not see that
        # directory unless we explicitly inject it.  Do this before run() calls
        # refresh_skill_index(), so the evolution agent can load and improve the
        # previous Skill instead of silently starting from scratch.
        self._host_evolved_skills_injected = await self._inject_host_evolved_skills(
            environment
        )

        # Inject previous verifier script into container (from prior runs)
        await self._inject_previous_verifier(environment)

    # ------------------------------------------------------------------
    # Meta-skill injection
    # ------------------------------------------------------------------

    def _find_meta_skills_dir(self) -> Path | None:
        """Locate the repository's public meta-skill directory on the host."""
        repo_root = Path(__file__).resolve().parents[4]
        candidates = [
            repo_root / "meta_skills",
            repo_root / ".claude" / "skills",  # backward compatibility
        ]
        # Also check SKILLSBENCH_ROOT for embedded benchmark checkouts.
        root = os.environ.get("SKILLSBENCH_ROOT")
        if root:
            candidates[0:0] = [
                Path(root) / "meta_skills",
                Path(root) / ".claude" / "skills",
            ]

        for candidate in candidates:
            if candidate.is_dir():
                return candidate
        return None

    async def _inject_meta_skills(self, environment: BaseEnvironment) -> list[str]:
        """Inject meta-skills into the container.

        Only skill-creator is injected. Verification is handled by the
        independent verifier agent (separate LLM session) — skill-verifier
        is no longer needed in the evolution agent's context.
        """
        meta_dir = self._find_meta_skills_dir()
        if not meta_dir:
            logger.warning("Meta-skills directory not found")
            return []

        name = "skill-creator"
        src = meta_dir / name
        if not src.exists():
            logger.info("Meta-skill %s not found at %s", name, src)
            return []

        success = await self._inject_skill_dir_to_container(name, src, environment)
        if success:
            logger.info("Injected meta-skill: %s", name)
            return [name]

        return []

    async def _inject_host_evolved_skills(
        self,
        environment: BaseEnvironment,
    ) -> list[str]:
        """Restore persisted ``evo-*`` packages into a fresh run container."""
        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir:
            return []

        host_skills_dir = Path(env_dir) / "skills"
        if not host_skills_dir.is_dir():
            return []

        injected: list[str] = []
        for skill_dir in sorted(host_skills_dir.glob("evo-*")):
            if not skill_dir.is_dir() or not (skill_dir / "SKILL.md").is_file():
                continue
            if await self._inject_skill_dir_to_container(
                skill_dir.name,
                skill_dir,
                environment,
            ):
                injected.append(skill_dir.name)

        if injected:
            logger.info(
                "Restored %d persisted evolved skills into fresh container: %s",
                len(injected),
                injected,
            )
        return injected

    @staticmethod
    def _select_evolved_skill_names(skill_names: set[str]) -> list[str]:
        """Return every task Skill, including packages restored for continuation."""
        return sorted(name for name in skill_names if name.startswith("evo-"))

    async def _inject_previous_verifier(self, environment: BaseEnvironment) -> None:
        """Inject the most recent exported verifier script into a new container.

        Looks for ``environment/verifier/generated_scripts/test_outputs_v*.py``
        on the host, takes the highest version, and injects it into the container
        at ``/root/verifier/test_outputs.py``.  This ensures manual reruns benefit
        from the verifier built during a prior run.
        """
        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir:
            return

        scripts_dir = Path(env_dir) / "verifier" / "generated_scripts"
        if not scripts_dir.is_dir():
            return

        candidates = sorted(scripts_dir.glob("test_outputs_v*.py"))
        if not candidates:
            return

        latest = candidates[-1]
        logger.info("Injecting previous verifier script: %s", latest.name)

        raw = latest.read_bytes()
        encoded = base64.b64encode(raw).decode("ascii")
        result = await environment.exec(
            command=(f"mkdir -p /root/verifier && echo '{encoded}' | base64 -d > /root/verifier/test_outputs.py"),
            timeout_sec=30,
        )
        if result.return_code != 0:
            logger.warning("Failed to inject previous verifier: %s", result.stderr)
        else:
            logger.info("Previous verifier injected into container at /root/verifier/test_outputs.py")

    async def _inject_skill_dir_to_container(
        self,
        name: str,
        skill_dir: Path,
        environment: BaseEnvironment,
    ) -> bool:
        """Copy a skill directory (SKILL.md, scripts/, references/) into the container.

        All file content is transferred via base64 encoding to avoid heredoc
        delimiter collisions with arbitrary skill content.
        """
        skill_md = skill_dir / "SKILL.md"
        if not skill_md.exists():
            return False

        container_dir = f"/app/environment/skills/{name}"

        # Write SKILL.md via base64
        encoded = base64.b64encode(skill_md.read_bytes()).decode("ascii")
        result = await environment.exec(
            command=f"mkdir -p '{container_dir}' && echo '{encoded}' | base64 -d > '{container_dir}/SKILL.md'",
            timeout_sec=30,
        )
        if result.return_code != 0:
            logger.warning("Failed to inject skill %s: %s", name, result.stderr)
            return False

        # Copy scripts/ via base64
        scripts_dir = skill_dir / "scripts"
        if scripts_dir.is_dir():
            for script_file in sorted(scripts_dir.iterdir()):
                if script_file.is_file():
                    encoded = base64.b64encode(script_file.read_bytes()).decode("ascii")
                    await environment.exec(
                        command=(
                            f"mkdir -p '{container_dir}/scripts' && "
                            f"echo '{encoded}' | base64 -d > '{container_dir}/scripts/{script_file.name}' && "
                            f"chmod +x '{container_dir}/scripts/{script_file.name}'"
                        ),
                        timeout_sec=15,
                    )

        # Copy references/ via base64
        refs_dir = skill_dir / "references"
        if refs_dir.is_dir():
            for ref_file in sorted(refs_dir.iterdir()):
                if ref_file.is_file():
                    encoded = base64.b64encode(ref_file.read_bytes()).decode("ascii")
                    await environment.exec(
                        command=(
                            f"mkdir -p '{container_dir}/references' && "
                            f"echo '{encoded}' | base64 -d > '{container_dir}/references/{ref_file.name}'"
                        ),
                        timeout_sec=15,
                    )

        # Restore app_files/ to /app/ — these are agent-created task files
        # (e.g. solution.py) that live outside /app/environment/ and would
        # otherwise be lost when a new container is created.
        app_files_dir = skill_dir / "app_files"
        if app_files_dir.is_dir():
            for app_file in sorted(app_files_dir.iterdir()):
                if app_file.is_file():
                    encoded = base64.b64encode(app_file.read_bytes()).decode("ascii")
                    await environment.exec(
                        command=f"echo '{encoded}' | base64 -d > '/app/{app_file.name}'",
                        timeout_sec=15,
                    )
            logger.info("Restored %d app_files to /app/ for skill %s", len(list(app_files_dir.iterdir())), name)

        return True

    async def _validate_evolved_skill_schema(
        self,
        environment: BaseEnvironment,
    ) -> list[str]:
        """Validate every direct ``evo-*`` Skill with the pinned strict schema.

        The same dependency-free validator runs on the host and in every task
        container.  Validator output contains only metadata rules and paths, so
        exact errors can safely be returned to the evolution Agent.
        """

        skill_paths = await self._find_evolved_skill_manifests(environment)
        if not skill_paths:
            return []

        validator_source = Path(skill_schema.__file__).read_bytes()
        encoded = base64.b64encode(validator_source).decode("ascii")
        paths_arg = " ".join(shlex.quote(path) for path in skill_paths)
        validate_result = await environment.exec(
            command=(
                f"echo {shlex.quote(encoded)} | base64 -d > /tmp/skill_schema.py "
                f"&& python3 /tmp/skill_schema.py {paths_arg}"
            ),
            timeout_sec=20,
        )
        if validate_result.return_code != 0:
            detail = (
                getattr(validate_result, "stderr", "")
                or validate_result.stdout
                or "unknown error"
            ).strip()
            logger.warning("Skill schema validator failed closed: %s", detail)
            return [
                "/app/environment/skills/evo-*/SKILL.md: "
                f"schema validator could not run ({detail[:180]})"
            ]

        issues: list[str] = []
        for line in validate_result.stdout.splitlines():
            stripped = line.strip()
            if not stripped.startswith(skill_schema.RESULT_MARKER):
                continue
            try:
                payload = json.loads(stripped[len(skill_schema.RESULT_MARKER) :])
                path = str(payload["path"])
                code = str(payload["code"])
                message = str(payload["message"])
            except (KeyError, TypeError, ValueError, json.JSONDecodeError):
                logger.warning("Malformed Skill schema validator record: %s", stripped[:500])
                issues.append(
                    "/app/environment/skills/evo-*/SKILL.md: "
                    "schema validator emitted a malformed result"
                )
                continue
            issues.append(f"{path}: {message} [{code}]")
        return issues

    async def _validate_skill_frontmatter(self, environment: BaseEnvironment) -> list[str]:
        """Backward-compatible alias for the strict full schema validator."""

        return await self._validate_evolved_skill_schema(environment)

    async def _evolved_skill_schema_preflight(
        self,
        environment: BaseEnvironment,
        *,
        stage: str,
    ) -> EpisodeExitResult | None:
        """Fail closed before verifier/GT and let the Agent repair exact errors."""

        manifests = await self._find_evolved_skill_manifests(environment)
        if not manifests:
            self._missing_skill_fix_count += 1
            trigger = f"missing_skill_{stage}"
            await self._record_intervention(environment, trigger=trigger)
            feedback = (
                "STRUCTURAL EVOLUTION REQUIREMENT: No evolved Skill exists at "
                "/app/environment/skills/evo-*/SKILL.md. No surrogate or GT "
                "evaluation was run, and no GT iteration was consumed.\n\n"
                "Load `skill-creator`, create an evo-* Skill yourself, validate its "
                "schema, regenerate the task output through its reusable procedure, "
                "and signal task_complete again."
            )
            if self._missing_skill_fix_count > self._max_skill_schema_fixes:
                self._exit_reason = "missing_skill"
                return EpisodeExitResult(should_exit=True, exit_reason=self._exit_reason)
            return EpisodeExitResult(should_exit=False, override_prompt=feedback)

        self._missing_skill_fix_count = 0
        issues = await self._validate_evolved_skill_schema(environment)
        validation = {
            "timestamp": datetime.now(UTC).isoformat(),
            "stage": stage,
            "passed": not issues,
            "manifests": manifests,
            "issues": issues,
        }
        self._skill_schema_validation_history.append(validation)
        if not issues:
            self._skill_schema_fix_count = 0
            return None

        self._skill_schema_fix_count += 1
        trigger = f"skill_schema_invalid_{stage}"
        intervention = await self._record_intervention(environment, trigger=trigger)
        intervention["skill_schema"] = validation
        logger.info(
            "Evolution schema preflight failed (stage=%s, issues=%d, repair=%d/%d)",
            stage,
            len(issues),
            self._skill_schema_fix_count,
            self._max_skill_schema_fixes,
        )
        if self._skill_schema_fix_count > self._max_skill_schema_fixes:
            self._exit_reason = "skill_schema_invalid"
            return EpisodeExitResult(should_exit=True, exit_reason=self._exit_reason)
        return EpisodeExitResult(
            should_exit=False,
            override_prompt=self._build_skill_schema_feedback(issues),
        )

    async def _find_evolved_skill_manifests(self, environment: BaseEnvironment) -> list[str]:
        """Return discoverable evo-* SKILL.md paths without reading their contents."""
        result = await environment.exec(
            command=(
                "find /app/environment/skills -mindepth 2 -maxdepth 2 "
                "-path '/app/environment/skills/evo-*/SKILL.md' -type f -print 2>/dev/null"
            ),
            timeout_sec=10,
        )
        if result.return_code != 0 or not result.stdout:
            return []
        return sorted(
            line.strip()
            for line in result.stdout.splitlines()
            if line.strip()
            and Path(line.strip()).parent.name.startswith("evo-")
            and Path(line.strip()).name == "SKILL.md"
        )

    # ------------------------------------------------------------------
    # Intra-run host intervention
    # ------------------------------------------------------------------

    async def _check_episode_exit(
        self,
        episode: int,
        result: Any,
        environment: BaseEnvironment,
        terminal_output: str,
    ) -> EpisodeExitResult:
        """Override: host-level verification between episodes.

        Cascade:
        1. Token budget > 70% → record intervention + force stop
        2. Idle/stale detection (when NOT task_complete):
           → consecutive empty episodes >= 3 OR episodes since last complete >= 30:
             force task_complete to trigger surrogate verification
           → otherwise: keep going
        3. Intervention cap exceeded → record intervention + accept and stop
        4. Independent surrogate verifier (separate LLM session):
           → fails or pass_rate < 1.0: record intervention, reset checklist (P2-P6), feed failures back
           → no script / script error: record intervention, feed back to agent
        5. Progress checklist → enforce once (first fail returns agent), then advisory (proceed to GT oracle)
        6. Record intervention (surrogate result + skills snapshot)
        7. Full GT oracle (always, default to same model):
           → pass: early stop + write oracle_stop.json
           → fail: feed back to agent
           → error/timeout: trust surrogate, accept
        """
        # ── 1. Token budget check (always) ──
        if self._chat is not None:
            max_tokens = self._get_max_context_tokens()
            if self._chat.last_input_tokens >= max_tokens * 0.7:
                self._exit_reason = "token_budget"
                _evo_print(f"[evolution] Exit: token_budget ({self._chat.last_input_tokens}/{max_tokens} tokens)")
                logger.info(
                    "Token budget exceeded: %d / %d (%.0f%%), force stopping (intervention_count=%d)",
                    self._chat.last_input_tokens,
                    max_tokens,
                    self._chat.last_input_tokens / max_tokens * 100,
                    self._host_intervention_count,
                )
                await self._record_intervention(environment, trigger="token_budget")
                return EpisodeExitResult(should_exit=True, exit_reason="token_budget")

        # ── 2. Idle / stale detection (only when agent did NOT signal task_complete) ──
        if not result.is_task_complete:
            has_commands = len(result.commands) > 0 and any(c.keystrokes.strip() for c in result.commands)
            if has_commands:
                self._consecutive_idle_episodes = 0
            else:
                self._consecutive_idle_episodes += 1
            self._episodes_since_last_complete += 1

            idle_triggered = self._consecutive_idle_episodes >= self._IDLE_EPISODE_LIMIT
            stale_triggered = self._episodes_since_last_complete >= self._STALE_EPISODE_LIMIT

            if idle_triggered or stale_triggered:
                trigger = "idle_agent" if idle_triggered else "stale_agent"
                logger.info(
                    "Agent appears stuck (%s): idle=%d, stale=%d — forcing surrogate verification",
                    trigger,
                    self._consecutive_idle_episodes,
                    self._episodes_since_last_complete,
                )
                # Force into the surrogate verification flow below
                result.is_task_complete = True
            else:
                return EpisodeExitResult(should_exit=False)

        # Agent signalled task_complete (or was forced by idle detection) — reset stale counters
        self._consecutive_idle_episodes = 0
        self._episodes_since_last_complete = 0

        # ── 3. Deterministic Skill schema preflight ──
        # This gate is independent of the optional surrogate verifier and runs
        # before every path that can spend a GT intervention. Exact schema
        # errors are safe to return because they contain no evaluator answers.
        schema_stage = "pre_gt" if self._skip_surrogate_verifier else "pre_surrogate"
        schema_exit = await self._evolved_skill_schema_preflight(
            environment,
            stage=schema_stage,
        )
        if schema_exit is not None:
            return schema_exit

        # ── 4. Intervention cap → prevent infinite loop ──
        cap_hit = self._host_intervention_count >= self._max_host_interventions
        retry_cap_hit = self._surrogate_retry_count >= self._max_surrogate_retries
        if cap_hit or retry_cap_hit:
            self._exit_reason = "max_interventions" if cap_hit else "max_surrogate_retries"
            _evo_print(
                f"[evolution] Intervention cap reached ({self._exit_reason}): "
                f"oracle={self._host_intervention_count}/{self._max_host_interventions}, "
                f"retries={self._surrogate_retry_count}/{self._max_surrogate_retries}"
            )
            logger.info(
                "Intervention cap reached (oracle=%d/%d, surrogate_retries=%d/%d), running final GT oracle before accepting",
                self._host_intervention_count,
                self._max_host_interventions,
                self._surrogate_retry_count,
                self._max_surrogate_retries,
            )
            intervention = await self._record_intervention(environment, trigger="max_interventions")

            # Every normal oracle intervention already ran the canonical GT
            # oracle and saved its best snapshot.  Re-running GT here after the
            # configured cap would turn a five-iteration experiment into a
            # hidden sixth evaluation.  Reuse the recorded outcome instead;
            # retain the final-oracle fallback below only for caps reached
            # without any completed GT result (for example infrastructure-only
            # failures or a surrogate retry cap).
            recorded_gt = self._recorded_best_gt_result() if cap_hit else None
            if recorded_gt is not None:
                intervention["gt_result"] = recorded_gt
                task_name = self._extract_task_name(environment)
                self._write_oracle_stop_signal(task_name, gt_result=recorded_gt)
                return EpisodeExitResult(
                    should_exit=True,
                    exit_reason=self._exit_reason,
                )

            # Final GT oracle — record result for analysis, then always stop
            oracle_model = self._gt_oracle_model or self._model_name
            evolved_skills_dir = await self._export_evolved_skills_to_host(environment)
            try:
                if evolved_skills_dir is None:
                    gt_oracle_result = {
                        "passed": False,
                        "error": "evolved Skill export unavailable before final GT",
                        "source": "information_boundary_gate",
                    }
                else:
                    boundary_issues = await self._audit_exported_evolved_skills(
                        environment,
                        evolved_skills_dir,
                        stage="max_interventions_final",
                    )
                    if boundary_issues:
                        logger.info(
                            "Final max-intervention boundary audit is advisory "
                            "(issues=%d); running the fresh canonical GT oracle",
                            len(boundary_issues),
                        )
                        self._log_static_gate_details(
                            "information_boundary", boundary_issues
                        )
                    # Static similarities to the current public artifact are
                    # portability diagnostics, not task-completion evidence.
                    # Protected evaluator/reference/history access remains
                    # default-denied by the command boundary.  Let the fresh
                    # canonical oracle decide whether the autonomous Skill
                    # actually solves the task.
                    gt_oracle_result = await self._run_gt_oracle_check(
                        self._instruction,
                        environment,
                        oracle_model=oracle_model,
                        evolved_skills_dir=evolved_skills_dir,
                        oracle_label=(
                            f"gt-oracle-max-{self._host_intervention_count}"
                        ),
                    )
            except Exception as oracle_exc:
                if isinstance(oracle_exc, ClaudeCodeProviderError):
                    raise
                logger.warning("Final GT oracle check failed: %s", oracle_exc)
                gt_oracle_result = {"passed": None, "error": str(oracle_exc)}
            finally:
                if evolved_skills_dir:
                    shutil.rmtree(evolved_skills_dir, ignore_errors=True)

            gt_error = gt_oracle_result.get("error") if gt_oracle_result else None
            if gt_error:
                intervention["gt_result"] = {
                    "source": gt_oracle_result.get("source", "gt_oracle"),
                    "passed": gt_oracle_result.get("passed"),
                    "error": gt_error,
                }
            elif gt_oracle_result is not None:
                gt_passed = gt_oracle_result.get("passed")
                raw = gt_oracle_result.get("raw_output") or ""
                intervention["gt_result"] = {
                    "source": "gt_oracle",
                    "passed": gt_passed,
                    "pass_rate": gt_oracle_result.get("pass_rate"),
                    "reward": gt_oracle_result.get("reward"),
                    "tests_passed": gt_oracle_result.get("tests_passed"),
                    "total_tests": gt_oracle_result.get("total_tests"),
                    "test_details": gt_oracle_result.get("test_details"),
                    "failure_reasons": gt_oracle_result.get("failure_reasons"),
                    "raw_output": raw[-5000:] if len(raw) > 5000 else raw,
                }
                await self._maybe_save_best_snapshot(environment, gt_oracle_result, self._host_intervention_count)
                if gt_passed is True:
                    self._exit_reason = "gt_oracle_pass"
                    task_name = self._extract_task_name(environment)
                    self._write_oracle_stop_signal(task_name, gt_result=gt_oracle_result)
                    return EpisodeExitResult(should_exit=True, exit_reason="gt_oracle_pass")

            return EpisodeExitResult(should_exit=True, exit_reason=self._exit_reason)

        # ── 5. Transferability and information-boundary preflight ──
        # The treatment being evaluated is an evolved Skill.  Enforce its
        # existence and discoverability before spending an independent-verifier
        # generation or a GT intervention.  In particular, a task whose current
        # output is still wrong must not be allowed to loop forever without ever
        # forming a transferable Skill.
        #
        # The existence check also applies to the GT-only ablation.  Previously
        # ``skip_surrogate_verifier`` skipped this entire preflight and invoked
        # the GT oracle before the first Skill had been created, consuming one
        # of the bounded GT interventions on a guaranteed ``no Skill`` result.
        if not self._skip_surrogate_verifier:
            # Reject a serialized current answer before it can influence either
            # surrogate or GT results. The audit never reads tests/solutions.
            evolved_skills_dir = await self._export_evolved_skills_to_host(environment)
            if evolved_skills_dir is None:
                logger.info(
                    "Evolution preflight failed (trigger=skill_export_missing_pre_surrogate)"
                )
                await self._record_intervention(
                    environment,
                    trigger="skill_export_missing_pre_surrogate",
                )
                feedback = (
                    "SKILL EXPORT FAILURE: A discoverable evo-* Skill was seen in the "
                    "container, but it could not be exported for information-boundary "
                    "validation. Check that the Skill directory and SKILL.md are readable, "
                    "repair the Skill without deleting its reusable implementation, "
                    "regenerate the task output through it, and signal task_complete again."
                )
                return EpisodeExitResult(should_exit=False, override_prompt=feedback)

            try:
                boundary_issues = await self._audit_exported_evolved_skills(
                    environment,
                    evolved_skills_dir,
                    stage="pre_surrogate",
                )
            finally:
                shutil.rmtree(evolved_skills_dir, ignore_errors=True)
            transfer_api_kinds = {
                "missing_runnable_example",
                "invalid_runnable_example",
                "documented_api_missing",
                "documented_api_not_invoked",
                "documented_example_unresolved_name",
            }
            transfer_api_issues = [
                issue for issue in boundary_issues if issue.kind in transfer_api_kinds
            ]
            if transfer_api_issues:
                logger.info(
                    "Evolution preflight advisory (trigger=skill_transfer_api, issues=%d)",
                    len(transfer_api_issues),
                )
                self._log_static_gate_details("transfer_api", transfer_api_issues)
            hard_property_issues = [
                issue
                for issue in boundary_issues
                if issue.kind == "hard_property_positive_evidence_missing"
            ]
            if hard_property_issues:
                logger.info(
                    "Evolution preflight advisory (trigger=hard_property_evidence_gate, issues=%d)",
                    len(hard_property_issues),
                )
                self._log_static_gate_details("hard_property", hard_property_issues)
            if boundary_issues:
                logger.info(
                    "Evolution preflight advisory (trigger=skill_information_boundary, issues=%d)",
                    len(boundary_issues),
                )
                self._log_static_gate_details("information_boundary", boundary_issues)
                # Static overlaps are portability diagnostics, not a completion
                # gate. Protected
                # evaluator/reference/history access is still rejected by the
                # command boundary above, and the fresh canonical GT remains
                # the sole success criterion.  Do not consume an intervention
                # or expose exact scanner evidence to the evolution agent.

            spreadsheet_issues = await self._spreadsheet_recalculation_issues(
                environment
            )
            if spreadsheet_issues:
                logger.info(
                    "Evolution preflight advisory "
                    "(trigger=spreadsheet_recalculation_gate, issues=%d)",
                    len(spreadsheet_issues),
                )
                self._log_static_gate_details(
                    "spreadsheet_recalculation", spreadsheet_issues
                )

        _evo_print(
            f"[evolution] Running surrogate verifier (oracle={self._host_intervention_count}/{self._max_host_interventions}, "
            f"retries={self._surrogate_retry_count}/{self._max_surrogate_retries}, locked={self._surrogate_tests_locked})"
        )
        logger.info(
            "Running independent verifier (oracle=%d/%d, surrogate_retries=%d/%d, locked=%s)",
            self._host_intervention_count,
            self._max_host_interventions,
            self._surrogate_retry_count,
            self._max_surrogate_retries,
            self._surrogate_tests_locked,
        )

        # ── 5. Independent surrogate verifier ──
        # When skip_surrogate_verifier is set, bypass entirely and go straight to GT oracle.
        if self._skip_surrogate_verifier:
            logger.info("Surrogate verifier skipped (--skip-surrogate-verifier flag)")
            surrogate_dict: dict = {"tests_passed": 0, "tests_failed": 0, "total_tests": 0, "pass_rate": 1.0}
            await self._record_intervention(environment, trigger="surrogate_skipped", surrogate_result=surrogate_dict)
        # When locked, re-run the existing pytest script without spawning a new LLM.
        # When unlocked, launch a separate LLM session to generate + run pytest.
        elif self._surrogate_tests_locked:
            # Re-run existing surrogate tests (no new LLM session)
            logger.info("Surrogate tests locked — re-running existing script via SelfVerifier")
            try:
                verification = await self._verifier.verify(environment=environment)
            except Exception as exc:
                logger.warning("Locked surrogate re-run failed: %s", exc)
                self._surrogate_retry_count += 1
                logger.info("Surrogate retry %d/%d (trigger=locked_rerun_exception)", self._surrogate_retry_count, self._max_surrogate_retries)
                await self._record_intervention(environment, trigger="locked_rerun_exception")
                feedback = (
                    "Verification could not be completed due to an infrastructure error. "
                    "Please verify your output manually — re-read the task instruction, "
                    "check all output files, and signal task_complete again."
                )
                return EpisodeExitResult(should_exit=False, override_prompt=feedback)

            if verification.source in ("no_script", "script_error"):
                self._surrogate_retry_count += 1
                error_detail = verification.error or verification.reasoning
                trigger = "locked_" + verification.source
                logger.warning("Locked surrogate issue (%s): %s", trigger, error_detail)
                logger.info("Surrogate retry %d/%d (trigger=%s)", self._surrogate_retry_count, self._max_surrogate_retries, trigger)
                await self._record_intervention(environment, trigger=trigger)
                feedback = (
                    f"Verification could not be completed ({error_detail}). "
                    "Please verify your output manually — re-read the task instruction, "
                    "check all output files, and signal task_complete again."
                )
                return EpisodeExitResult(should_exit=False, override_prompt=feedback)
        else:
            # Unlocked: launch independent verifier LLM to generate + run tests
            try:
                verification = await self._independent_verifier.generate_and_run(
                    environment=environment,
                    instruction=self._instruction,
                    logs_dir=self.logs_dir,
                    adversarial_recheck=self._adversarial_surrogate_recheck,
                )
                self._adversarial_surrogate_recheck = False
            except Exception as exc:
                logger.warning("Independent verification failed: %s", exc)
                self._surrogate_retry_count += 1
                logger.info("Surrogate retry %d/%d (trigger=verifier_exception)", self._surrogate_retry_count, self._max_surrogate_retries)
                await self._record_intervention(environment, trigger="verifier_exception")
                feedback = (
                    "Verification could not be completed due to an infrastructure error. "
                    "Please verify your output manually — re-read the task instruction, "
                    "check all output files, and signal task_complete again."
                )
                return EpisodeExitResult(should_exit=False, override_prompt=feedback)

            if verification.source in ("no_script", "script_error"):
                self._surrogate_retry_count += 1
                error_detail = verification.error or verification.reasoning
                trigger = {
                    "script_error": "verifier_script_error",
                    "no_script": "verifier_no_script",
                }.get(verification.source, "verifier_no_script")
                logger.warning("Verifier issue (%s): %s", trigger, error_detail)
                logger.info("Surrogate retry %d/%d (trigger=%s)", self._surrogate_retry_count, self._max_surrogate_retries, trigger)
                await self._record_intervention(environment, trigger=trigger)
                feedback = (
                    f"Verification could not be completed ({error_detail}). "
                    "Please verify your output manually — re-read the task instruction, "
                    "check all output files, and signal task_complete again."
                )
                return EpisodeExitResult(should_exit=False, override_prompt=feedback)

            # Valid script generated — lock surrogate tests
            self._surrogate_tests_locked = True
            logger.info("Surrogate tests locked after valid generation #%d", self._independent_verifier._generation_count)

        if not self._skip_surrogate_verifier:
            surrogate_dict = {
                "tests_passed": verification.tests_passed,
                "tests_failed": verification.tests_failed,
                "total_tests": verification.total_tests,
                "pass_rate": verification.tests_passed / verification.total_tests if verification.total_tests > 0 else 0.0,
            }

        surrogate_pass_rate = surrogate_dict["pass_rate"]
        if not self._skip_surrogate_verifier:
            _evo_print(
                f"[evolution] Surrogate: {verification.tests_passed}/{verification.total_tests} passed ({surrogate_pass_rate:.0%})"
            )
        if not self._skip_surrogate_verifier and (verification.tests_failed > 0 or surrogate_pass_rate < 1.0):
            self._surrogate_retry_count += 1
            trigger = "surrogate_fail" if verification.tests_failed > 0 else "surrogate_incomplete"
            logger.info("Surrogate retry %d/%d (trigger=%s)", self._surrogate_retry_count, self._max_surrogate_retries, trigger)

            failed_names = [d["name"] for d in verification.test_details if d.get("status") == "FAILED"]
            skipped_count = verification.total_tests - verification.tests_passed - verification.tests_failed
            logger.info(
                "Independent verifier: %d/%d passed, %d failed, %d skipped — failed: %s",
                verification.tests_passed,
                verification.total_tests,
                verification.tests_failed,
                skipped_count,
                ", ".join(failed_names[:5]) + ("..." if len(failed_names) > 5 else ""),
            )

            # Record intervention (includes cheap GT test)
            await self._record_intervention(environment, trigger=trigger, surrogate_result=surrogate_dict)

            # Reset progress checklist (P2-P6 unchecked)
            reset_msg = await self._reset_progress_checklist(environment)

            # In the locked path, run a lightweight diagnosis agent to provide
            # root-cause analysis (the unlocked path gets diagnosis from the full
            # generate_and_run agent, but the locked path skips that).
            if self._surrogate_tests_locked and verification.tests_failed > 0 and not verification.diagnosis:
                try:
                    diagnosis = await self._independent_verifier.diagnose_failures(
                        environment=environment,
                        instruction=self._instruction,
                        logs_dir=self.logs_dir,
                        verification=verification,
                    )
                    if diagnosis:
                        verification.diagnosis = diagnosis
                        logger.info("Diagnosis-only agent provided %d chars of analysis", len(diagnosis))
                except Exception as diag_exc:
                    logger.warning("Diagnosis-only agent failed: %s", diag_exc)

            # Feed surrogate failures + reset notice back to agent
            if verification.tests_failed > 0:
                feedback = self._build_surrogate_feedback(verification)
            else:
                # Tests were skipped/errored (pass_rate < 1.0 but tests_failed == 0)
                feedback = (
                    f"Independent verifier: only {verification.tests_passed}/{verification.total_tests} tests passed "
                    f"({skipped_count} tests were skipped or errored — likely due to missing dependencies or import errors "
                    f"in the verifier script). A pass rate of 100% is required.\n\n"
                    "Please check that your output is complete and correct, then signal task_complete again."
                )
            feedback = reset_msg + "\n\n" + feedback
            return EpisodeExitResult(should_exit=False, override_prompt=feedback)

        # ── 6. Progress checklist — enforce once, then advisory ──
        # Skip checklist and frontmatter validation when surrogate is disabled (clean ablation)
        if self._skip_surrogate_verifier:
            unchecked = []
            checklist_feedback = None
        else:
            try:
                progress_text = await self._read_progress_checklist(environment)
                unchecked = self._parse_unchecked_items(progress_text)
            except Exception:
                unchecked = []

        checklist_feedback: str | None = None
        if unchecked:
            self._checklist_fail_count += 1
            if self._checklist_fail_count <= 1:
                # First checklist fail — enforce: send agent back to complete items
                self._surrogate_retry_count += 1
                logger.info(
                    "Surrogate retry %d/%d (trigger=surrogate_pass_checklist_fail, attempt=%d)",
                    self._surrogate_retry_count,
                    self._max_surrogate_retries,
                    self._checklist_fail_count,
                )
                await self._record_intervention(environment, trigger="surrogate_pass_checklist_fail", surrogate_result=surrogate_dict)
                feedback = (
                    "Independent verifier tests passed, but you have unchecked phases in /root/progress.md:\n"
                    + "\n".join(f"- {item}" for item in unchecked)
                    + "\n\nComplete each unchecked phase, update /root/progress.md, "
                    "then signal task_complete."
                )
                logger.info(
                    "Surrogate passed but %d checklist items unchecked — returning agent (attempt %d)",
                    len(unchecked),
                    self._checklist_fail_count,
                )
                return EpisodeExitResult(should_exit=False, override_prompt=feedback)

            # 2nd+ checklist fail — proceed to GT oracle anyway
            logger.info(
                "Surrogate passed, %d checklist items unchecked (attempt %d) — proceeding to GT oracle",
                len(unchecked),
                self._checklist_fail_count,
            )
            checklist_feedback = (
                "Note: You have unchecked phases in /root/progress.md:\n"
                + "\n".join(f"- {item}" for item in unchecked)
                + "\n\nPlease complete each unchecked phase and update /root/progress.md."
            )
        else:
            self._checklist_fail_count = 0  # reset on success

        # ── 7. Surrogate passed — run GT oracle ──
        self._host_intervention_count += 1
        trigger = "surrogate_pass_checklist_incomplete" if unchecked else "surrogate_pass"
        logger.info(
            "Oracle intervention %d/%d (surrogate 100%% pass, checklist_incomplete=%s)",
            self._host_intervention_count,
            self._max_host_interventions,
            bool(unchecked),
        )
        intervention = await self._record_intervention(environment, trigger=trigger, surrogate_result=surrogate_dict)

        # ── 7. Full GT oracle (always, default to same model) ──
        oracle_model = self._gt_oracle_model or self._model_name
        _evo_print(
            f"[evolution] Running GT oracle (iteration {self._host_intervention_count}/{self._max_host_interventions})..."
        )
        logger.info("Surrogate passed all tests — running GT oracle mid-loop with %s", oracle_model)

        evolved_skills_dir = await self._export_evolved_skills_to_host(environment)
        try:
            gt_oracle_result = await self._run_gt_oracle_check(
                self._instruction,
                environment,
                oracle_model=oracle_model,
                evolved_skills_dir=evolved_skills_dir,
                oracle_label=(
                    f"gt-oracle-{self._host_intervention_count}"
                    f"-attempt-{self._gt_infrastructure_retry_count + 1}"
                ),
            )
        except Exception as oracle_exc:
            if isinstance(oracle_exc, ClaudeCodeProviderError):
                raise
            logger.warning("GT oracle check failed during intervention: %s", oracle_exc)
            gt_oracle_result = {"passed": None, "error": str(oracle_exc)}
        finally:
            if evolved_skills_dir:
                shutil.rmtree(evolved_skills_dir, ignore_errors=True)

        gt_passed = gt_oracle_result.get("passed") if gt_oracle_result else None
        gt_error = gt_oracle_result.get("error") if gt_oracle_result else None
        if gt_error:
            _evo_print(f"[evolution] GT Oracle: ERROR ({gt_error})")
        elif gt_oracle_result is not None:
            gt_pr = gt_oracle_result.get("pass_rate", 0.0)
            _evo_print(f"[evolution] GT Oracle: {'PASS' if gt_passed else 'FAIL'} ({gt_pr:.0%})")
        else:
            _evo_print("[evolution] GT Oracle: ERROR (no result)")

        # Back-fill GT oracle result into this intervention entry for analysis
        # (test_details/failure_reasons/raw_output are for logging only — never shown to agent)
        if gt_oracle_result is not None:
            if gt_error:
                intervention["gt_result"] = {
                    "source": "gt_oracle",
                    "passed": None,
                    "error": gt_error,
                }
            else:
                raw = gt_oracle_result.get("raw_output") or ""
                intervention["gt_result"] = {
                    "source": "gt_oracle",
                    "passed": gt_passed,
                    "pass_rate": gt_oracle_result.get("pass_rate"),
                    "reward": gt_oracle_result.get("reward"),
                    "tests_passed": gt_oracle_result.get("tests_passed"),
                    "total_tests": gt_oracle_result.get("total_tests"),
                    "test_details": gt_oracle_result.get("test_details"),
                    "failure_reasons": gt_oracle_result.get("failure_reasons"),
                    "raw_output": raw[-5000:] if len(raw) > 5000 else raw,
                }
                await self._maybe_save_best_snapshot(environment, gt_oracle_result, self._host_intervention_count)

        if gt_passed is True:
            self._gt_infrastructure_retry_count = 0
            self._exit_reason = "gt_oracle_pass"
            _evo_print("[evolution] Exit: gt_oracle_pass — stopping early")
            logger.info("GT oracle passed — stopping early")
            task_name = self._extract_task_name(environment)
            self._write_oracle_stop_signal(task_name, gt_result=gt_oracle_result)
            return EpisodeExitResult(should_exit=True, exit_reason="gt_oracle_pass")

        if gt_passed is False:
            self._gt_infrastructure_retry_count = 0
            self._surrogate_tests_locked = False
            self._adversarial_surrogate_recheck = True
            self._checklist_fail_count = 0  # reset so agent gets one enforced checklist pass per GT cycle
            _evo_print("[evolution] GT oracle failed — returning agent for refinement")
            logger.info("GT oracle failed — unlocking surrogate tests for refinement")
            feedback = (
                "The independent verifier tests all passed, but a ground truth evaluation "
                "found issues with your output. Your solution likely has gaps.\n\n"
                "Please:\n"
                "1. Re-read the FULL task instruction and every supplied background document carefully. "
                "Write a short invariant checklist from them, and reject any proposed Skill "
                "change that contradicts those public invariants\n"
                "2. Enumerate several instruction-compatible failure hypotheses before changing code; "
                "check extraction versus canonical representation, boundary/indexing cases, formula "
                "provenance, runtime discovery, and final artifact validation using only supplied inputs\n"
                "3. Test the hypotheses with local or synthetic diagnostics and change one reusable "
                "procedure at a time rather than guessing an evaluator-specific answer. When one "
                "failure exposes a family-level gap, audit and validate the entire public family "
                "instead of adding a one-case exception\n"
                "4. Fix your skill scripts and re-run to regenerate output\n"
                "5. Do NOT edit output files directly — fix the skill logic\n"
                "Then signal task_complete."
            )
            safe_categories = self._safe_gt_failure_categories(gt_oracle_result)
            if safe_categories:
                feedback += (
                    "\n\nBroad failure dimensions observed: "
                    + "; ".join(safe_categories)
                    + ". These labels intentionally omit test names, expected values, "
                    "thresholds, reference outputs, and current-instance answers."
                )
            safe_output_fields = self._safe_gt_failed_output_fields(
                gt_oracle_result,
                self._instruction,
            )
            if safe_output_fields:
                feedback += (
                    "\n\nInstruction-declared output fields implicated: `"
                    + "`, `".join(safe_output_fields)
                    + "`. Field names come from the public output schema; expected "
                    "and actual values remain host-only."
                )
            safe_process_guidance = self._safe_gt_refinement_guidance(
                gt_oracle_result
            )
            if safe_process_guidance:
                feedback += (
                    "\n\nTransfer-process guidance:\n- "
                    + "\n- ".join(safe_process_guidance)
                )
            if checklist_feedback:
                feedback = feedback + "\n\n" + checklist_feedback
            logger.info("GT oracle failed — returning agent to fix issues")
            return EpisodeExitResult(should_exit=False, override_prompt=feedback)

        # Oracle error (passed=None) — fail closed.  A surrogate pass is never a
        # substitute for an explicit full-score result from the fresh GT run.
        # Infrastructure failures also do not constitute a scientific
        # intervention: return the budget reserved above and track a separate
        # bounded retry count. Otherwise one closed output pipe can exhaust all
        # refinement cycles without running a single canonical test.
        self._host_intervention_count = max(0, self._host_intervention_count - 1)
        self._gt_infrastructure_retry_count += 1
        gt_err_msg = gt_error or "unknown error"
        _evo_print(
            f"[evolution] GT oracle unavailable ({gt_err_msg}) — "
            f"infrastructure retry {self._gt_infrastructure_retry_count}/"
            f"{self._max_gt_infrastructure_retries}; scientific budget unchanged"
        )
        logger.warning("GT oracle error (%s) — failing closed and requesting retry", gt_err_msg)
        if self._gt_infrastructure_retry_count >= self._max_gt_infrastructure_retries:
            self._exit_reason = "gt_infrastructure_unavailable"
            return EpisodeExitResult(
                should_exit=True,
                exit_reason=self._exit_reason,
            )
        return EpisodeExitResult(
            should_exit=False,
            override_prompt=(
                "The fresh ground-truth evaluation was unavailable because of an "
                "infrastructure error. This is not a pass, and no output or Skill change "
                "is requested from this event. Confirm the existing output still exists "
                "and signal task_complete again so the full-score evaluation can retry."
            ),
        )

    @staticmethod
    def _build_surrogate_feedback(verification: VerificationResult) -> str:
        """Return broad, answer-free surrogate feedback to the evolution agent.

        The independent verifier and diagnosis worker are allowed to inspect the
        current artifact so that the host can decide whether another evolution
        round is warranted.  Their test names, assertion messages, tracebacks,
        and diagnosis are *not* treatment inputs: they can contain current row or
        frame identifiers, expected outputs, exact measurements, and suggested
        cutoffs.  Feeding those details back lets the Skill overfit the canonical
        instance even when its final source happens to look generic.

        Keep all detailed evidence in the host logs and expose only labels from
        the same fixed answer-free taxonomy used for GT feedback.
        """
        categories = HarborTerminus2Evolution._safe_gt_failure_categories(
            {"test_details": verification.test_details}
        )
        lines = [
            "HOST VERIFICATION: An independent verifier detected failures in your output.",
            "The result is not full score; detailed test names, instance IDs, "
            "expected/actual values, measurements, thresholds, tracebacks, and "
            "diagnosis remain host-only.",
        ]
        if categories:
            lines.extend(
                [
                    "",
                    "Broad failure dimensions observed: " + "; ".join(categories) + ".",
                ]
            )

        lines.extend(
            [
                "",
                "Re-read the unchanged instruction and background document, enumerate several "
                "compatible failure hypotheses, and test them with runtime or synthetic "
                "evidence before changing one reusable procedure in the Skill.",
                "Fix the general failure family in the skill's scripts/ and re-run to "
                "regenerate output; do not add a current-instance exception or a fixed "
                "cutoff copied from this evaluation.",
                "Do NOT edit output files directly — fix the skill logic.",
                "Then signal task_complete when you believe output is correct.",
            ]
        )

        return "\n".join(lines)

    @staticmethod
    def _extract_failure_tracebacks(raw_output: str) -> str:
        """Extract FAILURES section from pytest --tb=short output.

        Returns the content between ``= FAILURES =`` and
        ``short test summary info`` headers, truncated to ~3000 chars
        to avoid blowing up the context window.
        """
        start = raw_output.find("= FAILURES =")
        if start < 0:
            start = raw_output.find("FAILURES")
        if start < 0:
            return ""
        end = raw_output.find("short test summary info", start)
        if end < 0:
            end = raw_output.find("====", start + 20)
        if end < 0:
            end = len(raw_output)
        section = raw_output[start:end].strip()
        if len(section) > 3000:
            # Cut at a line boundary to avoid garbled output
            cut = section.rfind("\n", 0, 3000)
            if cut > 0:
                section = section[:cut]
            else:
                section = section[:3000]
            section += "\n... (truncated)"
        return section

    async def _record_intervention(
        self,
        environment: BaseEnvironment,
        trigger: str,
        surrogate_result: dict | None = None,
    ) -> dict:
        """Record an intervention entry with surrogate result and skills snapshot.

        Returns the intervention entry dict.  The ``gt_result`` field is initially
        ``None`` and may be back-filled later by the full GT oracle.
        """
        # Snapshot current skills in container
        try:
            skills = sorted(await self._list_container_skills(environment))
        except Exception as exc:
            logger.warning("Failed to snapshot container skills during intervention recording: %s", exc)
            skills = []

        entry: dict = {
            "intervention_number": len(self._intervention_history) + 1,
            "timestamp": datetime.now(UTC).isoformat(),
            "trigger": trigger,
            "surrogate_result": surrogate_result,
            "surrogate_tests_locked": self._surrogate_tests_locked,
            "gt_result": None,
            "skills_snapshot": skills,
        }
        self._intervention_history.append(entry)
        logger.info(
            "Intervention #%d recorded (trigger=%s)",
            entry["intervention_number"],
            trigger,
        )
        return entry

    async def _maybe_save_best_snapshot(self, environment: BaseEnvironment, gt_result: dict, intervention_number: int) -> None:
        """Save a skills snapshot if this GT result is the best so far.

        Each time the GT oracle returns a valid result, compare its canonical
        reward to the current best. If better, export the container's evolved skills
        to a persistent temp directory and record the snapshot metadata.
        The previous best snapshot's temp dir is cleaned up immediately.
        """
        full_score, reward = self._gt_full_score_and_reward(gt_result)
        pass_rate = gt_result.get("pass_rate", 0.0)
        current_best = self._best_gt_snapshot.get("reward", -1) if self._best_gt_snapshot else -1
        if reward <= current_best:
            return

        # Clean up old snapshot
        if self._best_gt_snapshot and self._best_gt_snapshot.get("skills_dir"):
            shutil.rmtree(self._best_gt_snapshot["skills_dir"], ignore_errors=True)

        # Export current skills to a new temp dir
        skills_dir = await self._export_evolved_skills_to_host(environment)
        self._best_gt_snapshot = {
            "passed": full_score,
            "reward": reward,
            "pass_rate": pass_rate,
            "tests_passed": gt_result.get("tests_passed"),
            "total_tests": gt_result.get("total_tests"),
            "intervention_number": intervention_number,
            "skills_dir": skills_dir,
        }
        _evo_print(
            f"[evolution] New best GT reward: {reward:.0%} "
            f"at iteration #{intervention_number}"
        )
        logger.info(
            "Saved best GT snapshot: reward=%.1f%%, tests=%.1f%% at intervention #%d",
            reward * 100,
            pass_rate * 100,
            intervention_number,
        )

    def _recorded_best_gt_result(self) -> dict | None:
        """Return the best completed GT outcome when no skill was exportable."""
        best = self._best_gt_snapshot
        if not best:
            return None

        matching_gt: dict = {}
        for intervention in reversed(self._intervention_history):
            candidate = intervention.get("gt_result") or {}
            _candidate_full, candidate_reward = self._gt_full_score_and_reward(
                candidate
            )
            if (
                candidate_reward == best.get("reward")
                and candidate.get("tests_passed") == best.get("tests_passed")
                and candidate.get("total_tests") == best.get("total_tests")
            ):
                matching_gt = candidate
                break

        result = {
            "passed": best.get("reward") == 1.0,
            "tests_passed": best.get("tests_passed"),
            "total_tests": best.get("total_tests"),
            "pass_rate": best.get("pass_rate"),
            "reward": best.get("reward"),
            "source": "recorded_best_terminal",
        }
        for key in ("test_details", "failure_reasons", "raw_output"):
            if key in matching_gt:
                result[key] = matching_gt[key]
        logger.info(
            "Using recorded terminal GT result: reward=%.1f%%, tests=%s/%s",
            float(best.get("reward") or 0.0) * 100,
            best.get("tests_passed"),
            best.get("total_tests"),
        )
        return result

    async def _reset_progress_checklist(self, environment: BaseEnvironment) -> str:
        """Reset /root/progress.md with P1 checked and P2-P6 unchecked.

        Returns a message string explaining the reset to prepend to agent feedback.
        """
        checklist = (
            "# Progress\n"
            "- [x] P1: Discover environment files (ls /app/environment/, /root/)\n"
            "- [x] P1b: Discover installed tools and libraries\n"
            "- [ ] P2: Create/update task skill with utility function scripts\n"
            "- [ ] P3: Self-reflect (re-read FULL instruction, verify skill covers ALL requirements)\n"
            "- [ ] P4: Execute task (run skill scripts, produce ALL output files)\n"
            "- [ ] P5: Fix any failures from host verifier feedback, re-run until stable\n"
            "- [ ] P6: Write /root/evolution_summary.md\n"
        )
        encoded = base64.b64encode(checklist.encode("utf-8")).decode("ascii")
        result = await environment.exec(
            command=f"echo '{encoded}' | base64 -d > /root/progress.md",
            timeout_sec=15,
        )
        if result.return_code != 0:
            logger.warning("Failed to reset progress checklist: %s", result.stderr)

        return (
            "CHECKLIST RESET: /root/progress.md has been reset. P1 and P1b (environment and tool discovery) "
            "are already done. You MUST complete P2-P6 in order before signalling task_complete.\n"
            "Start with P2: review and update your task skill, then proceed through each phase."
        )

    # ------------------------------------------------------------------
    # Main run loop
    # ------------------------------------------------------------------

    async def run(
        self,
        instruction: str,
        environment: BaseEnvironment,
        context: AgentContext,
    ) -> None:
        """Run the evolution agent: execute task in-container with host intervention.

        The agent creates skills, executes the task, runs tests, and writes a summary — all in-container.
        Host intervention (independent verifier, progress checklist, GT oracle) gates task_complete
        during execution.  Post-execution, the host exports verifier scripts and writes logs.
        """
        self._instruction = instruction
        run_timestamp = datetime.now(UTC).isoformat()
        task_name = self._extract_task_name(environment)
        run_id = f"{self.name()}-{task_name}-{run_timestamp}"
        self._run_id = run_id

        # Reset host intervention state for this run
        self._host_intervention_count = 0
        self._gt_infrastructure_retry_count = 0
        self._surrogate_retry_count = 0
        self._exit_reason = ""
        self._intervention_history = []
        self._independent_verifier.reset()
        self._surrogate_tests_locked = False
        self._checklist_fail_count = 0
        self._skill_schema_fix_count = 0
        self._skill_schema_validation_history = []
        self._missing_skill_fix_count = 0
        self._skill_boundary_fix_count = 0
        self._adversarial_surrogate_recheck = False
        self._fresh_source_knowledge_eligible = False
        self._lineage_command_gate_clean = True
        self._information_boundary_validated = False
        self._last_information_boundary_issues = []
        self._initial_container_source_identifiers = None
        self._consecutive_idle_episodes = 0
        self._episodes_since_last_complete = 0
        # Clean up any leftover best snapshot from a previous run
        if self._best_gt_snapshot and self._best_gt_snapshot.get("skills_dir"):
            shutil.rmtree(self._best_gt_snapshot["skills_dir"], ignore_errors=True)
        self._best_gt_snapshot = None
        logger.info("Evolution state reset for run %s (task=%s)", run_id, task_name)

        # ── Timers ──
        total_start = time.monotonic()

        # Refresh skill index so parent sees meta skills
        await self.refresh_skill_index()

        # Initialize result variables
        execution_start = time.monotonic()
        execution_duration = 0.0
        skills_loaded: list[str] = []
        trajectory_data: list[dict] | None = None
        trajectory_summary = ""
        post_execution_start = time.monotonic()
        evolution_summary_text = ""
        interrupted_by_error: str | None = None
        gt_oracle_result: dict | None = None
        main_exit_reason: str = ""
        main_intervention_count: int = 0

        # Snapshot pre-existing skills so we can diff after execution
        self._pre_existing_skills = await self._list_container_skills(environment)
        self._fresh_source_knowledge_eligible = (
            self._fresh_run_source_knowledge_eligibility(
                self._pre_existing_skills,
                self._host_evolved_skills_injected,
            )
        )
        # Freeze the visible-source provenance before the agent can create or
        # edit task outputs.  Every later boundary audit in this run uses this
        # immutable baseline.
        self._initial_container_source_identifiers = (
            await self._container_current_source_identifiers(environment)
        )
        self._write_source_knowledge_provenance(status="initialized")

        agent_error: BaseException | None = None
        try:
            # ── Single agent execution — agent does everything in-container ──
            # Wrap in nested try/except so post-execution work (export verifier,
            # read summary) still runs even if the agent
            # times out (AgentTimeoutError / asyncio.CancelledError).
            try:
                await super().run(instruction, environment, context)
            except BaseException as run_exc:
                agent_error = run_exc
                interrupted_by_error = f"{type(run_exc).__name__}: {run_exc}"
                logger.warning("Agent execution interrupted: %s", interrupted_by_error)

                # Parent's run() writes full_conversation.json at the end of its
                # method body.  When it's interrupted (timeout, cancel, etc.) that
                # write is skipped, so we do it here to preserve the conversation.
                try:
                    if self._chat is not None and hasattr(self._chat, "messages"):
                        (self.logs_dir / "full_conversation.json").write_text(
                            json.dumps(self._chat.messages, indent=2, ensure_ascii=False),
                            encoding="utf-8",
                        )
                        logger.info("Saved full_conversation.json after interrupted run")
                except Exception as conv_save_exc:
                    logger.warning("Failed to save conversation after interrupt: %s", conv_save_exc)

            execution_duration = time.monotonic() - execution_start

            # Snapshot exit reason from main run before follow-ups overwrite it
            main_exit_reason = self._exit_reason
            main_intervention_count = self._host_intervention_count

            # Extract results (read trajectory once for both summary and execution info)
            skills_loaded = list((context.metadata or {}).get("skills_loaded", []))
            trajectory_data = self._read_trajectory()
            trajectory_summary = self._build_trajectory_summary(trajectory_data)

            # Read evolution_summary.md from container
            evolution_summary_text = await self._read_evolution_summary(environment)

            # Export agent-generated verifier script before container is destroyed
            try:
                exported = await self._export_verifier_script(environment)
                if exported:
                    logger.info("Verifier script exported to %s", exported)
            except Exception as export_exc:
                logger.warning("Failed to export verifier script: %s", export_exc)

            # Import agent-created skills from container to host
            try:
                imported = await self._import_agent_created_skills(environment, self._pre_existing_skills)
                if imported:
                    logger.info("Imported %d agent-created skills to host: %s", len(imported), imported)
            except Exception as import_exc:
                logger.warning("Failed to import agent-created skills: %s", import_exc)

            # Final post-execution oracle test with evolved skills.  When the
            # configured GT budget is exhausted, reuse its recorded best result
            # rather than silently launching an extra oracle evaluation.
            oracle_budget_exhausted = (
                main_intervention_count >= self._max_host_interventions
                and self._recorded_best_gt_result() is not None
            )
            final_manifests = await self._find_evolved_skill_manifests(environment)
            final_schema_issues = (
                await self._validate_evolved_skill_schema(environment)
                if final_manifests
                else [
                    "/app/environment/skills/evo-*/SKILL.md: no Agent-authored "
                    "evolved Skill manifest exists [missing_skill]"
                ]
            )
            self._skill_schema_validation_history.append(
                {
                    "timestamp": datetime.now(UTC).isoformat(),
                    "stage": "post_execution_final",
                    "passed": not final_schema_issues,
                    "manifests": final_manifests,
                    "issues": final_schema_issues,
                }
            )
            if final_schema_issues:
                self._exit_reason = "skill_schema_invalid"
                gt_oracle_result = {
                    "passed": False,
                    "error": "evolved Skill schema validation failed before final GT",
                    "source": "skill_schema_gate",
                    "skill_schema_issues": final_schema_issues,
                }
                logger.warning(
                    "Skipping final GT because evolved Skill schema is invalid: %s",
                    "; ".join(final_schema_issues),
                )
            elif oracle_budget_exhausted:
                gt_oracle_result = self._recorded_best_gt_result()
                best = self._best_gt_snapshot
                if best and best.get("skills_dir"):
                    self._rollback_host_skills(environment, best["skills_dir"])
                logger.info(
                    "GT budget exhausted at %d/%d; using recorded best without "
                    "an additional final oracle run",
                    main_intervention_count,
                    self._max_host_interventions,
                )
            elif not self._did_last_oracle_pass():
                logger.info("Running final post-execution oracle test with evolved skills")

                # Prefer best snapshot skills (already GT-validated during iteration)
                # over last skills — avoids a wasteful double oracle run.
                best = self._best_gt_snapshot
                if best and best.get("skills_dir"):
                    logger.info(
                        "Using best snapshot skills (reward=%.1f%%, tests=%.1f%% "
                        "from intervention #%d) for final oracle",
                        best["reward"] * 100,
                        best["pass_rate"] * 100,
                        best["intervention_number"],
                    )
                    oracle_skills_dir = best["skills_dir"]
                    is_best_snapshot = True
                else:
                    oracle_skills_dir = await self._export_evolved_skills_to_host(environment)
                    is_best_snapshot = False

                if oracle_skills_dir is not None:
                    try:
                        boundary_issues = await self._audit_exported_evolved_skills(
                            environment,
                            oracle_skills_dir,
                            stage="post_execution_final",
                        )
                        if boundary_issues:
                            logger.info(
                                "Final post-execution boundary audit is advisory "
                                "(issues=%d); running the fresh canonical GT oracle",
                                len(boundary_issues),
                            )
                            self._log_static_gate_details(
                                "information_boundary", boundary_issues
                            )
                        oracle_model = self._gt_oracle_model or self._model_name
                        gt_oracle_result = await self._run_gt_oracle_check(
                            self._instruction,
                            environment,
                            oracle_model=oracle_model,
                            evolved_skills_dir=oracle_skills_dir,
                            oracle_label="gt-oracle-final",
                        )
                        if gt_oracle_result is not None and not gt_oracle_result.get("error"):
                            if is_best_snapshot:
                                gt_oracle_result["source"] = "best_snapshot"
                            logger.info(
                                "Final oracle result: passed=%s, pass_rate=%s",
                                gt_oracle_result.get("passed"),
                                gt_oracle_result.get("pass_rate"),
                            )
                        elif gt_oracle_result is not None and gt_oracle_result.get("error"):
                            logger.warning("Final oracle returned error: %s", gt_oracle_result.get("error"))
                            gt_oracle_result = None
                    except Exception as oracle_exc:
                        if isinstance(oracle_exc, ClaudeCodeProviderError):
                            raise
                        logger.warning("Final post-execution oracle test failed: %s", oracle_exc)
                        gt_oracle_result = None
                        # If oracle failed but we have best snapshot values, use them directly
                        if is_best_snapshot:
                            gt_oracle_result = {
                                "passed": best["reward"] == 1.0,
                                "tests_passed": best["tests_passed"],
                                "total_tests": best["total_tests"],
                                "pass_rate": best["pass_rate"],
                                "reward": best["reward"],
                                "source": "best_snapshot",
                            }
                            logger.info(
                                "Using recorded best snapshot values: reward=%.1f%%, tests=%.1f%%",
                                best["reward"] * 100,
                                best["pass_rate"] * 100,
                            )
                    finally:
                        # Only clean up if we exported last skills (best snapshot dir is managed elsewhere)
                        if not is_best_snapshot:
                            shutil.rmtree(oracle_skills_dir, ignore_errors=True)

                    # If oracle non-determinism gave a worse result than best snapshot, use snapshot values
                    if (
                        is_best_snapshot
                        and gt_oracle_result is not None
                    ):
                        _oracle_full, oracle_reward = self._gt_full_score_and_reward(
                            gt_oracle_result
                        )
                        if oracle_reward < best["reward"]:
                            logger.info(
                                "Oracle reward (%.1f%%) worse than snapshot (%.1f%%) "
                                "due to non-determinism — using snapshot values",
                                oracle_reward * 100,
                                best["reward"] * 100,
                            )
                            gt_oracle_result = {
                                "passed": best["reward"] == 1.0,
                                "tests_passed": best["tests_passed"],
                                "total_tests": best["total_tests"],
                                "pass_rate": best["pass_rate"],
                                "reward": best["reward"],
                                "source": "best_snapshot",
                            }

                    # Overwrite host skills with best snapshot version
                    if is_best_snapshot:
                        self._rollback_host_skills(environment, best["skills_dir"])

                # A terminal algorithmic failure can have a valid recorded GT
                # result but no exported skill directory. The canonical case is
                # five oracle rejections because evolution produced no Skill.
                # Persist that real 0/1 outcome instead of discarding the run as
                # missing infrastructure.
                if gt_oracle_result is None:
                    gt_oracle_result = self._recorded_best_gt_result()

                if gt_oracle_result is not None:
                    task_name_for_oracle = self._extract_task_name(environment)
                    self._write_oracle_stop_signal(task_name_for_oracle, gt_result=gt_oracle_result)
            else:
                logger.info("Skipping final oracle — last intervention oracle already passed")
                # Extract GT result from last intervention so _write_evolution_log gets it
                last_gt = self._intervention_history[-1].get("gt_result") if self._intervention_history else None
                if last_gt is not None:
                    gt_oracle_result = {
                        "passed": last_gt.get("passed"),
                        "tests_passed": last_gt.get("tests_passed"),
                        "total_tests": last_gt.get("total_tests"),
                        "pass_rate": last_gt.get("pass_rate"),
                        "reward": last_gt.get("reward", last_gt.get("pass_rate")),
                    }

            # Re-raise agent error after all post-execution work is done
            if agent_error is not None:
                raise agent_error

        except BaseException as exc:
            if not interrupted_by_error:
                interrupted_by_error = f"{type(exc).__name__}: {exc}"
            logger.warning("Evolution run error: %s", interrupted_by_error)
            if not trajectory_summary:
                trajectory_summary = self._build_trajectory_summary()
            if not skills_loaded and context.metadata is not None:
                skills_loaded = list(context.metadata.get("skills_loaded", []))
            if not execution_duration:
                execution_duration = time.monotonic() - execution_start
            raise
        finally:
            post_execution_duration = time.monotonic() - post_execution_start
            total_duration = time.monotonic() - total_start

            # Clean up best GT snapshot temp dir
            if self._best_gt_snapshot and self._best_gt_snapshot.get("skills_dir"):
                shutil.rmtree(self._best_gt_snapshot["skills_dir"], ignore_errors=True)

            if context.metadata is not None:
                context.metadata["evolution_summary"] = evolution_summary_text

            try:
                self._write_evolution_log(
                    task_name=task_name,
                    run_timestamp=run_timestamp,
                    skills_loaded=skills_loaded,
                    trajectory_summary=trajectory_summary,
                    trajectory_data=trajectory_data,
                    context=context,
                    timing={
                        "total_duration_sec": round(total_duration, 1),
                        "execution_duration_sec": round(execution_duration, 1),
                        "post_execution_duration_sec": round(post_execution_duration, 1),
                        "interrupted_by": interrupted_by_error,
                        "exit_reason": main_exit_reason,
                        "host_intervention_count": main_intervention_count,
                        "surrogate_retry_count": self._surrogate_retry_count,
                    },
                    evolution_summary=evolution_summary_text,
                    gt_oracle_result=gt_oracle_result,
                    intervention_history=self._intervention_history,
                )
            except Exception as log_exc:
                logger.error("Failed to write evolution log in finally: %s", log_exc)

            # Re-generate full_conversation.md
            try:
                self._regenerate_conversation_md()
            except Exception as conv_exc:
                logger.warning("Failed to regenerate conversation md: %s", conv_exc)

    async def _run_ground_truth_evaluation(self, environment: BaseEnvironment) -> dict | None:
        """Run ground truth tests for measuring evolution effectiveness.

        Results are for analysis only — NOT fed back to the agent or written to history.
        Uploads the entire tests/ directory to /tests/ in the container (matching
        Harbor's normal verifier behaviour) so that co-located modules (test_utils.py,
        conftest.py) and data files (catalog.csv, labels.csv, …) are available.
        """
        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir:
            return None
        task_dir = Path(env_dir).parent
        task_toml_path = task_dir / "task.toml"
        verifier_env: dict[str, str] | None = None
        if task_toml_path.is_file():
            task_config = TaskConfig.model_validate(
                tomllib.loads(task_toml_path.read_text(encoding="utf-8"))
            )
            if task_config.verifier.env:
                # Match Harbor's normal verifier semantics.  The fresh GT
                # container previously ran test.sh without [verifier.env], so
                # tasks whose public verifier contract declares values such as
                # REPO_ID failed for infrastructure reasons before testing the
                # evolved Skill.
                verifier_env = resolve_env_vars(task_config.verifier.env)
        tests_dir = task_dir / "tests"
        test_file = tests_dir / "test_outputs.py"
        test_sh = tests_dir / "test.sh"
        if not test_file.exists() and not test_sh.exists():
            return None

        # Upload the whole tests/ directory (only once — skip if already there)
        check = await environment.exec(
            command="test -f /tests/test_outputs.py && echo exists || echo missing",
            timeout_sec=10,
        )
        if "missing" in (check.stdout or ""):
            try:
                await environment.upload_dir(
                    source_dir=str(tests_dir),
                    target_dir="/tests",
                )
            except Exception:
                logger.warning("GT evaluation: failed to upload tests/ directory, falling back to single-file inject")
                raw = test_file.read_bytes()
                encoded = base64.b64encode(raw).decode("ascii")
                inject = await environment.exec(
                    command=f"mkdir -p /tests && echo '{encoded}' | base64 -d > /tests/test_outputs.py",
                    timeout_sec=30,
                )
                if inject.return_code != 0:
                    return None
            # Ensure pytest + dependencies
            await environment.exec(
                command="pip install --break-system-packages -q pytest 2>/dev/null || pip install -q pytest",
                timeout_sec=60,
            )

        # Run GT test — prefer test.sh (handles dependencies + app startup) over bare pytest
        if test_sh.exists():
            # Upload test.sh and run it; it installs deps, starts app, runs pytest, writes reward.txt
            raw_sh = test_sh.read_bytes()
            encoded_sh = base64.b64encode(raw_sh).decode("ascii")
            await environment.exec(
                command=f"echo '{encoded_sh}' | base64 -d > /tests/test.sh && chmod +x /tests/test.sh",
                timeout_sec=30,
            )
            result = await environment.exec(
                command="bash /tests/test.sh 2>&1",
                env=verifier_env,
                timeout_sec=1800,
            )
        else:
            result = await environment.exec(
                command="python3 -m pytest /tests/test_outputs.py -v --tb=short 2>&1",
                env=verifier_env,
                timeout_sec=120,
            )
        # Reuse SelfVerifier's pytest parser
        vr = SelfVerifier._parse_pytest_output(result.stdout or "", result.return_code)

        # Fallback: some tasks redirect test output instead of leaving it on the
        # test.sh stdout. Harbor benchmark tasks use both output.log and
        # test-stdout.txt, so check both before reducing a successful test.sh to
        # an artificial 0/0 result.
        if vr.total_tests == 0:
            for verifier_log in (
                "/logs/verifier/output.log",
                "/logs/verifier/test-stdout.txt",
            ):
                log_result = await environment.exec(
                    command=f"cat {verifier_log} 2>/dev/null || true",
                    timeout_sec=10,
                )
                if not log_result.stdout or not log_result.stdout.strip():
                    continue
                vr_fallback = SelfVerifier._parse_pytest_output(
                    log_result.stdout,
                    result.return_code,
                )
                if vr_fallback.total_tests > 0:
                    logger.info(
                        "GT evaluation: parsed %d tests from %s (stdout had 0)",
                        vr_fallback.total_tests,
                        verifier_log,
                    )
                    vr = vr_fallback
                    break

        # Canonical benchmark reward is authoritative even when pytest reports
        # every structural test as passed. Some verifiers intentionally award a
        # fractional score after those tests, so reading reward.txt only for a
        # 0-test parse can incorrectly promote a non-full result as 1.0.
        reward_override = None
        reward_result = await environment.exec(
            command="cat /logs/verifier/reward.txt 2>/dev/null || true",
            timeout_sec=10,
        )
        if reward_result.stdout and reward_result.stdout.strip():
            # Docker exec can prefix stdout with shell/TTY warnings. Parse the
            # last standalone numeric line instead of requiring the complete
            # captured stream to be a float.
            numeric_lines = re.findall(
                r"(?m)^\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+))\s*$",
                reward_result.stdout,
            )
            if numeric_lines:
                reward_override = float(numeric_lines[-1])
                logger.info("GT evaluation: found reward.txt = %s", reward_override)

        gt_result = {
            "tests_passed": vr.tests_passed,
            "total_tests": vr.total_tests,
            "tests_failed": vr.tests_failed,
            "pass_rate": vr.tests_passed / vr.total_tests if vr.total_tests > 0 else 0.0,
            "estimated_success": vr.estimated_success,
            "test_details": vr.test_details,
            "failure_reasons": vr.failure_reasons,
            "raw_output": vr.raw_output,
        }
        if reward_override is not None:
            gt_result["reward"] = reward_override
            # Synthesize test counts from reward when pytest parsing found 0 tests
            # (e.g., tasks using custom test scripts that write reward.txt directly)
            if vr.total_tests == 0:
                gt_result["total_tests"] = 1
                gt_result["tests_passed"] = 1 if reward_override >= 1.0 else 0
                gt_result["tests_failed"] = 0 if reward_override >= 1.0 else 1
                gt_result["pass_rate"] = 1.0 if reward_override >= 1.0 else 0.0
        return gt_result

    @staticmethod
    def _gt_full_score_and_reward(gt_result: dict[str, Any]) -> tuple[bool, float]:
        """Resolve strict success from canonical reward, then parsed test counts.

        A finite numeric ``reward`` is authoritative. Missing or malformed
        reward data falls back to parsed tests so ordinary pytest-only tasks
        retain their established behavior.
        """

        reported_reward = gt_result.get("reward")
        if reported_reward is not None:
            try:
                numeric_reward = float(reported_reward)
            except (TypeError, ValueError):
                numeric_reward = float("nan")
            if math.isfinite(numeric_reward):
                return numeric_reward == 1.0, numeric_reward

        total = gt_result.get("total_tests", 0)
        passed = gt_result.get("tests_passed", 0)
        return (
            total > 0 and passed == total,
            passed / total if total > 0 else 0.0,
        )

    def _create_oracle_agent(self, agent_type: str, logs_dir: Path, model_name: str):
        """Create an oracle agent instance based on the configured agent type."""
        built_in_agents = {
            "terminus-2",
            "claude-code",
            "claude-code-skills",
            "claude-code-skill-only",
            "codex",
            "codex-skill-only",
            "codex-subscription",
            "gemini-cli",
        }
        if agent_type not in built_in_agents:
            raw_imports = os.environ.get(
                "COEVOSKILLS_CUSTOM_AGENT_IMPORT_PATHS", "{}"
            )
            try:
                custom_imports = json.loads(raw_imports)
            except json.JSONDecodeError as exc:
                raise ValueError(
                    "COEVOSKILLS_CUSTOM_AGENT_IMPORT_PATHS must contain a JSON object"
                ) from exc
            import_path = (
                custom_imports.get(agent_type)
                if isinstance(custom_imports, dict)
                else None
            )
            if not isinstance(import_path, str) or ":" not in import_path:
                raise ValueError(
                    f"Unknown oracle agent type: {agent_type!r}. Register custom "
                    "Agents with --agent-import-path NAME=MODULE:CLASS."
                )
            module_name, class_name = import_path.rsplit(":", 1)
            try:
                agent_class = getattr(importlib.import_module(module_name), class_name)
            except (ImportError, AttributeError) as exc:
                raise ValueError(
                    f"Cannot import custom oracle Agent {agent_type!r} from "
                    f"{import_path!r}"
                ) from exc
            return agent_class(logs_dir=logs_dir, model_name=model_name)

        if agent_type == "terminus-2":
            return HarborTerminus2WithSkills(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "claude-code":
            from libs.terminus_agent.agents.bedrock_claude_code import BedrockClaudeCode

            return BedrockClaudeCode(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "claude-code-skills":
            from libs.terminus_agent.agents.claude_code_skills import ClaudeCodeSkills

            return ClaudeCodeSkills(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "claude-code-skill-only":
            from libs.terminus_agent.agents.claude_code_skill_only import (
                ClaudeCodeSkillOnly,
            )

            return ClaudeCodeSkillOnly(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "codex":
            from harbor.agents.installed.codex import Codex

            return Codex(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "codex-skill-only":
            from libs.terminus_agent.agents.codex_skill_only import CodexSkillOnly

            return CodexSkillOnly(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "codex-subscription":
            from libs.terminus_agent.agents.codex_subscription import CodexSubscription

            return CodexSubscription(logs_dir=logs_dir, model_name=model_name)
        elif agent_type == "gemini-cli":
            from harbor.agents.installed.gemini_cli import GeminiCli

            return GeminiCli(logs_dir=logs_dir, model_name=model_name)
        else:
            raise ValueError(
                f"Unknown oracle agent type: {agent_type!r}. Expected "
                "'terminus-2', 'claude-code', 'claude-code-skills', "
                "'claude-code-skill-only', 'codex', 'codex-skill-only', "
                "'codex-subscription', or 'gemini-cli'."
            )

    @staticmethod
    def _claude_trace_used_evolved_skill(
        trace_path: Path,
        expected_skill_names: set[str],
    ) -> bool:
        """Return whether Claude Code loaded or executed an injected evo-* Skill.

        Claude Code can use an installed Skill through the native ``Skill`` tool,
        but it may also follow the transfer prompt by reading ``SKILL.md`` and
        importing or executing the package's helper scripts directly. Both are
        genuine treatment uptake and must be recorded as such.
        """
        if not trace_path.exists() or not expected_skill_names:
            return False
        try:
            with trace_path.open(encoding="utf-8") as handle:
                for line in handle:
                    try:
                        event = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    if event.get("type") != "assistant":
                        continue
                    message = event.get("message") or {}
                    content = message.get("content") or []
                    if not isinstance(content, list):
                        continue
                    for block in content:
                        if not isinstance(block, dict):
                            continue
                        if block.get("type") != "tool_use":
                            continue
                        tool_name = block.get("name")
                        tool_input = block.get("input") or {}
                        if tool_name == "Skill":
                            skill_name = tool_input.get("skill")
                            if skill_name in expected_skill_names:
                                return True
                            continue
                        if tool_name not in {"Read", "Glob", "Bash"}:
                            continue
                        serialized_input = json.dumps(
                            tool_input,
                            ensure_ascii=False,
                            sort_keys=True,
                        )
                        for skill_name in expected_skill_names:
                            skill_paths = (
                                f"/root/.claude/skills/{skill_name}",
                                f"/app/environment/skills/{skill_name}",
                            )
                            if any(path in serialized_input for path in skill_paths):
                                return True
        except OSError:
            return False
        return False

    @staticmethod
    def _claude_trace_execution_violation(
        trace_path: Path,
        expected_skill_names: set[str],
        public_contract: str = "",
    ) -> str | None:
        """Return why a fresh Claude run bypassed its immutable Skill, if any.

        Loading a Skill is not treatment compliance when its documented helper
        fails and the clean agent repairs the target manually.  The evaluation
        prompt permits runtime path/parameter adaptation but explicitly forbids
        ad-hoc reimplementation.  Require at least one successful Bash call into
        an injected ``evo-*`` package after the Skill tool invocation, and reject
        direct Edit/Write mutations of task files after that point.
        """
        if not trace_path.exists() or not expected_skill_names:
            return "Claude trace or expected evolved Skill is unavailable"

        skill_invoked = False
        helper_call_ids: set[str] = set()
        successful_helper = False
        failed_helper = False
        mutation_tools = {"Edit", "Write", "NotebookEdit"}

        try:
            with trace_path.open(encoding="utf-8") as handle:
                for line in handle:
                    try:
                        event = json.loads(line)
                    except json.JSONDecodeError:
                        continue

                    if event.get("type") == "assistant":
                        message = event.get("message") or {}
                        content = message.get("content") or []
                        if not isinstance(content, list):
                            continue
                        for block in content:
                            if not isinstance(block, dict) or block.get("type") != "tool_use":
                                continue
                            tool_name = block.get("name")
                            tool_input = block.get("input") or {}
                            if tool_name == "Skill":
                                if tool_input.get("skill") in expected_skill_names:
                                    skill_invoked = True
                                continue
                            if not skill_invoked:
                                continue
                            if tool_name in mutation_tools:
                                target = (
                                    tool_input.get("file_path")
                                    or tool_input.get("path")
                                    or tool_input.get("notebook_path")
                                    or "task artifact"
                                )
                                if "/environment/skills/evo-" not in str(target):
                                    return (
                                        f"fresh agent used {tool_name} to mutate {target} "
                                        "after invoking the Skill"
                                    )
                            if tool_name == "Bash":
                                command = str(tool_input.get("command") or "")
                                if HarborTerminus2Evolution._manual_repository_history_oracle(
                                    command,
                                    public_contract,
                                ):
                                    return (
                                        "manual_repository_history_oracle: fresh agent "
                                        "inspected repository history or a historical revision "
                                        "outside the immutable helper, although the public task "
                                        "contract did not prescribe that operation"
                                    )
                                if HarborTerminus2Evolution._manual_instance_target_fallback(command):
                                    return (
                                        "manual_instance_target_fallback: fresh agent "
                                        "supplied a hand-authored per-instance identity/target "
                                        "literal list instead of letting the immutable helper "
                                        "discover targets from runtime inputs"
                                    )
                                if HarborTerminus2Evolution._manual_instance_solution_search_fallback(
                                    command
                                ):
                                    return (
                                        "manual_instance_target_fallback: fresh agent "
                                        "performed current-instance combinatorial solution "
                                        "search outside the immutable helper instead of "
                                        "calling one helper-owned end-to-end optimizer"
                                    )
                                if HarborTerminus2Evolution._manual_low_level_skill_inspection(
                                    command
                                ):
                                    return (
                                        "manual_instance_inspection_fallback: fresh agent "
                                        "used low-level evolved-Skill parsing/discovery APIs "
                                        "to inventory the current input outside one "
                                        "helper-owned end-to-end generation and validation call"
                                    )
                                if HarborTerminus2Evolution._manual_low_level_skill_mutation(
                                    command,
                                    public_contract,
                                ):
                                    return (
                                        "manual_instance_mutation_fallback: fresh agent "
                                        "selected and composed low-level evolved-Skill "
                                        "mutation primitives for the current input instead "
                                        "of calling one helper-owned end-to-end repair and "
                                        "validation entry point"
                                    )
                                if HarborTerminus2Evolution._manual_shell_identity_target_fallback(
                                    command,
                                    public_contract,
                                ):
                                    return (
                                        "manual_instance_target_fallback: fresh agent "
                                        "constructed a current-instance identity/target search "
                                        "expression outside the immutable helper instead of "
                                        "using helper-owned runtime discovery and validation"
                                    )
                                if HarborTerminus2Evolution._manual_direct_artifact_construction(
                                    command
                                ):
                                    return (
                                        "manual_artifact_construction_fallback: fresh agent "
                                        "directly serialized a task artifact outside the "
                                        "immutable helper instead of invoking one end-to-end "
                                        "helper that owns generation and validation"
                                    )
                                if HarborTerminus2Evolution._manual_exact_replacement_fallback(
                                    command,
                                    public_contract,
                                ):
                                    return (
                                        "manual_exact_replacement_fallback: fresh agent "
                                        "supplied an undeclared fixed old/new source "
                                        "replacement to a lower-level mutation primitive "
                                        "instead of returning the incomplete end-to-end "
                                        "helper result for evolution"
                                    )
                                if failed_helper and _DIRECT_ENVIRONMENT_MUTATION_RE.search(command):
                                    return (
                                        "manual_environment_fallback: after the immutable helper "
                                        "failed, the fresh agent changed packages outside the "
                                        "helper instead of returning the failure for evolution"
                                    )
                                if failed_helper and _DIRECT_PATCH_MUTATION_RE.search(command):
                                    return (
                                        "manual_artifact_fallback: after the immutable helper "
                                        "failed, the fresh agent directly patched the task "
                                        "outside the helper"
                                    )
                                references_skill = "/skills/evo-" in command or any(
                                    skill_name in command for skill_name in expected_skill_names
                                )
                                tool_id = block.get("id")
                                if references_skill and isinstance(tool_id, str):
                                    helper_call_ids.add(tool_id)

                    elif event.get("type") == "user" and helper_call_ids:
                        message = event.get("message") or {}
                        content = message.get("content") or []
                        if not isinstance(content, list):
                            continue
                        for block in content:
                            if not isinstance(block, dict) or block.get("type") != "tool_result":
                                continue
                            tool_id = block.get("tool_use_id")
                            if tool_id not in helper_call_ids:
                                continue
                            if block.get("is_error") is True:
                                failed_helper = True
                                continue
                            successful_helper = True
        except OSError as exc:
            return f"could not read Claude trace: {exc}"

        if skill_invoked and not successful_helper:
            return "fresh agent invoked the Skill but no helper call completed successfully"
        return None

    @staticmethod
    def _manual_repository_history_oracle(
        command: str,
        public_contract: str = "",
    ) -> bool:
        """Detect caller-side use of repository history as a reference oracle.

        Ordinary inspection of the current checkout remains allowed, including
        ``git status``, ``git diff --check``, and a working-tree comparison with
        ``HEAD``.  An immutable helper may also inspect history internally: its
        implementation is not expanded into the Claude trace.  The exception
        here is only for a public task contract that explicitly names the same
        Git history operation.
        """
        matches: list[tuple[str, str]] = []
        for match in _GIT_HISTORY_SUBCOMMAND_RE.finditer(command):
            subcommand = match.group("subcommand").lower()
            matches.append((subcommand, f"git {subcommand}"))

        for match in _GIT_DIFF_COMMAND_RE.finditer(command):
            arguments = match.group("arguments") or ""
            if _GIT_HISTORICAL_DIFF_OPERAND_RE.search(arguments):
                matches.append(("diff", "git diff"))

        if not matches:
            return False

        contract = public_contract.lower()
        for _subcommand, operation in matches:
            operation_positions = [
                match.start()
                for match in re.finditer(re.escape(operation), contract)
            ]
            explicitly_prescribed = False
            for position in operation_positions:
                prefix = contract[max(0, position - 80):position]
                if re.search(
                    r"(?i)(?:do\s+not|don't|must\s+not|never|avoid|without|"
                    r"forbid(?:den)?(?:\s+to)?)\b[^.!?;\n]{0,40}$",
                    prefix,
                ):
                    continue
                explicitly_prescribed = True
                break
            if not explicitly_prescribed:
                return True
        return False

    @staticmethod
    def _manual_instance_target_fallback(command: str) -> bool:
        """Detect hand-authored current-instance target lists in Bash Python.

        Runtime input/output paths and ordinary configuration are permitted.
        This only rejects a high-confidence list-shaped target/identity variable
        (or target keyword) containing several literal instance values. A helper
        that derives the same list from its input at runtime remains valid.
        """
        sources = _trace_python_sources(command)
        if not sources:
            return False

        def literal_targets(node: ast.AST) -> set[str]:
            values = {
                child.value.strip()
                for child in ast.walk(node)
                if isinstance(child, ast.Constant)
                and isinstance(child.value, str)
                and len(child.value.strip()) >= 3
            }
            return {
                value
                for value in values
                if not value.startswith(("/", "./", "../", "-", "http://", "https://"))
                and value.lower() not in {
                    "text", "pages", "before_refs", "after_refs", "all", "auto",
                    "json", "csv", "pdf", "xlsx", "input", "output",
                }
                and not re.fullmatch(r"[A-Za-z0-9_.-]+\.(?:pdf|txt|csv|json|xlsx|py)", value)
            }

        def literal_coordinate_pairs(node: ast.AST) -> set[tuple[float, float]]:
            coordinates: set[tuple[float, float]] = set()
            for child in ast.walk(node):
                if not isinstance(child, (ast.Tuple, ast.List)) or len(child.elts) != 2:
                    continue
                values: list[float] = []
                for element in child.elts:
                    if (
                        isinstance(element, ast.Constant)
                        and isinstance(element.value, (int, float))
                        and not isinstance(element.value, bool)
                    ):
                        values.append(float(element.value))
                if len(values) == 2:
                    coordinates.add((values[0], values[1]))
            return coordinates

        for source in sources:
            try:
                tree = ast.parse(source)
            except SyntaxError:
                continue
            for node in ast.walk(tree):
                if isinstance(node, (ast.Assign, ast.AnnAssign)):
                    targets = node.targets if isinstance(node, ast.Assign) else [node.target]
                    names = [target.id for target in targets if isinstance(target, ast.Name)]
                    value_node = node.value
                    if (
                        value_node is not None
                        and any(_INSTANCE_TARGET_NAME_RE.search(name) for name in names)
                        and isinstance(value_node, (ast.List, ast.Tuple, ast.Set, ast.Dict))
                        and (
                            len(literal_targets(value_node)) >= 3
                            or len(literal_coordinate_pairs(value_node)) >= 3
                        )
                    ):
                        return True
                elif isinstance(node, ast.Call):
                    for keyword_argument in node.keywords:
                        if (
                            keyword_argument.arg
                            and _INSTANCE_TARGET_NAME_RE.search(keyword_argument.arg)
                            and isinstance(
                                keyword_argument.value,
                                (ast.List, ast.Tuple, ast.Set, ast.Dict),
                            )
                            and (
                                len(literal_targets(keyword_argument.value)) >= 3
                                or len(literal_coordinate_pairs(keyword_argument.value)) >= 3
                            )
                        ):
                            return True
        return False

    @staticmethod
    def _manual_instance_solution_search_fallback(command: str) -> bool:
        """Detect fresh-caller combinatorial search using low-level Skill APIs.

        The oracle may pass runtime paths to one end-to-end optimizer. It may
        not import lower-level candidate/scoring helpers and then run its own
        combinations, permutations, or Cartesian product over the current
        instance. That makes the caller—not the evolved Skill—the solver even
        if it never overwrites the final artifact.
        """

        if "/skills/evo-" not in command:
            return False
        combinators = {"combinations", "combinations_with_replacement", "permutations", "product"}
        for source in _trace_python_sources(command):
            try:
                tree = ast.parse(source)
            except SyntaxError:
                continue
            imported_combinators: set[str] = set()
            itertools_aliases: set[str] = {"itertools"}
            for node in ast.walk(tree):
                if isinstance(node, ast.Import):
                    for alias in node.names:
                        if alias.name == "itertools":
                            itertools_aliases.add(alias.asname or alias.name)
                elif isinstance(node, ast.ImportFrom) and node.module == "itertools":
                    for alias in node.names:
                        if alias.name in combinators:
                            imported_combinators.add(alias.asname or alias.name)
                elif isinstance(node, ast.Call):
                    if (
                        isinstance(node.func, ast.Name)
                        and node.func.id in imported_combinators
                    ):
                        return True
                    elif (
                        isinstance(node.func, ast.Attribute)
                        and isinstance(node.func.value, ast.Name)
                        and node.func.value.id in itertools_aliases
                        and node.func.attr in combinators
                    ):
                        return True
        return False

    @staticmethod
    def _manual_low_level_skill_inspection(command: str) -> bool:
        """Detect caller-owned semantic inventory via low-level Skill helpers.

        Runtime paths passed to a single end-to-end entry point are legitimate.
        A fresh oracle may not instead import parsing/discovery primitives from
        the evolved package and iterate over the current PDFs, map tiles, rows,
        pages, or other instance records to make its own target/acceptance
        decisions. This gate is intentionally high-confidence: it requires a
        Skill path, a called low-level imported API, a current artifact path,
        and either caller-side iteration or multiple such primitive calls.
        """

        if "/skills/evo-" not in command:
            return False

        path_suffixes = (
            ".csv", ".json", ".jsonl", ".pdf", ".xlsx", ".xlsm", ".xls",
            ".civ6map", ".db", ".sqlite", ".png", ".jpg", ".jpeg", ".mp4",
            ".wav", ".txt",
        )
        iteration_nodes = (
            ast.For,
            ast.AsyncFor,
            ast.While,
            ast.ListComp,
            ast.SetComp,
            ast.DictComp,
            ast.GeneratorExp,
        )

        for source in _trace_python_sources(command):
            try:
                tree = ast.parse(source)
            except SyntaxError:
                continue

            imported_functions: dict[str, str] = {}
            module_aliases: set[str] = set()
            for node in ast.walk(tree):
                if isinstance(node, ast.ImportFrom):
                    for alias in node.names:
                        imported_functions[alias.asname or alias.name] = alias.name
                elif isinstance(node, ast.Import):
                    module_aliases.update(alias.asname or alias.name for alias in node.names)

            low_level_calls: set[str] = set()
            for node in ast.walk(tree):
                if not isinstance(node, ast.Call):
                    continue
                original_name = ""
                if isinstance(node.func, ast.Name) and node.func.id in imported_functions:
                    original_name = imported_functions[node.func.id]
                elif (
                    isinstance(node.func, ast.Attribute)
                    and isinstance(node.func.value, ast.Name)
                    and node.func.value.id in module_aliases
                ):
                    original_name = node.func.attr
                if original_name and _LOW_LEVEL_SKILL_INSPECTION_NAME_RE.search(original_name):
                    low_level_calls.add(original_name.lower())

            if not low_level_calls:
                continue

            has_current_artifact_path = any(
                isinstance(node, ast.Constant)
                and isinstance(node.value, str)
                and (
                    node.value.lower().endswith(path_suffixes)
                    or node.value.startswith(("/data/", "/root/", "/input/"))
                )
                and "/skills/evo-" not in node.value
                for node in ast.walk(tree)
            )
            if not has_current_artifact_path:
                continue

            has_caller_iteration = any(
                isinstance(node, iteration_nodes) for node in ast.walk(tree)
            )
            if has_caller_iteration or len(low_level_calls) >= 2:
                return True
        return False

    @staticmethod
    def _manual_low_level_skill_mutation(
        command: str,
        public_contract: str = "",
    ) -> bool:
        """Detect caller-owned repair policy composed from low-level mutators.

        A fresh oracle may supply runtime paths to one end-to-end repair entry
        point. It may not choose current files, dependencies, versions, source
        replacements, or other repair policy and then orchestrate several
        lower-level evolved-Skill writers. This detector stays conservative:
        it requires an evolved-Skill path, a current artifact path, and either
        multiple mutation calls, caller iteration, or a mutation call supplied
        with several undeclared policy literals.
        """

        if "/skills/evo-" not in command:
            return False

        path_suffixes = (
            ".xml", ".pom", ".gradle", ".java", ".kt", ".scala", ".py",
            ".js", ".jsx", ".ts", ".tsx", ".json", ".yaml", ".yml",
            ".toml", ".ini", ".cfg", ".conf", ".properties", ".lock",
            ".csv", ".xlsx", ".pdf",
        )
        iteration_nodes = (
            ast.For,
            ast.AsyncFor,
            ast.While,
            ast.ListComp,
            ast.SetComp,
            ast.DictComp,
            ast.GeneratorExp,
        )
        normalized_contract = re.sub(
            r"[^a-z0-9_.:+-]+",
            " ",
            public_contract.lower(),
        )

        for source in _trace_python_sources(command):
            try:
                tree = ast.parse(source)
            except SyntaxError:
                continue

            imported_functions: dict[str, str] = {}
            module_aliases: set[str] = set()
            for node in ast.walk(tree):
                if isinstance(node, ast.ImportFrom):
                    for alias in node.names:
                        imported_functions[alias.asname or alias.name] = alias.name
                elif isinstance(node, ast.Import):
                    module_aliases.update(alias.asname or alias.name for alias in node.names)

            mutation_calls: list[ast.Call] = []
            for node in ast.walk(tree):
                if not isinstance(node, ast.Call):
                    continue
                original_name = ""
                if isinstance(node.func, ast.Name) and node.func.id in imported_functions:
                    original_name = imported_functions[node.func.id]
                elif (
                    isinstance(node.func, ast.Attribute)
                    and isinstance(node.func.value, ast.Name)
                    and node.func.value.id in module_aliases
                ):
                    original_name = node.func.attr
                if original_name and _LOW_LEVEL_SKILL_MUTATION_NAME_RE.search(original_name):
                    mutation_calls.append(node)

            if not mutation_calls:
                continue

            has_current_artifact_path = any(
                isinstance(node, ast.Constant)
                and isinstance(node.value, str)
                and (
                    node.value.lower().endswith(path_suffixes)
                    or node.value.startswith(("/data/", "/root/", "/input/", "/app/"))
                )
                and "/skills/evo-" not in node.value
                for node in ast.walk(tree)
            )
            if not has_current_artifact_path:
                continue

            if len(mutation_calls) >= 2 or any(
                isinstance(node, iteration_nodes) for node in ast.walk(tree)
            ):
                return True

            policy_literals = {
                child.value.strip()
                for call in mutation_calls
                for child in ast.walk(call)
                if isinstance(child, ast.Constant)
                and isinstance(child.value, str)
                and len(child.value.strip()) >= 2
                and not child.value.startswith(("/", "./", "../", "-"))
                and not child.value.lower().endswith(path_suffixes)
                and child.value.lower() not in {"utf-8", "true", "false", "none"}
            }
            undeclared_policy = {
                value
                for value in policy_literals
                if value.lower() not in normalized_contract
            }
            if len(undeclared_policy) >= 2:
                return True
        return False

    @staticmethod
    def _manual_shell_identity_target_fallback(
        command: str,
        public_contract: str = "",
    ) -> bool:
        """Detect caller-authored current-identity grep inventories.

        A clean oracle may inspect generic output properties, but it may not
        manually enumerate the current authors, institutions, grants, emails,
        or other instance identities and use that list as a substitute for the
        immutable helper's runtime discovery/validation. Publicly declared
        tokens and small generic audit expressions remain allowed.
        """

        normalized_contract = re.sub(r"[^a-z0-9@._]+", " ", public_contract.lower())

        for line in command.splitlines():
            try:
                tokens = shlex.split(line, posix=True)
            except ValueError:
                continue
            for index, token in enumerate(tokens):
                if Path(token).name != "grep":
                    continue

                pattern: str | None = None
                cursor = index + 1
                while cursor < len(tokens):
                    candidate = tokens[cursor]
                    if candidate in {"-e", "--regexp", "-E", "--extended-regexp"}:
                        cursor += 1
                        if cursor < len(tokens):
                            pattern = tokens[cursor]
                        break
                    if candidate.startswith("-"):
                        cursor += 1
                        continue
                    pattern = candidate
                    break

                if not pattern:
                    continue
                alternatives = re.split(r"(?<!\\)\|", pattern)
                # A compact generic grep such as error|warning|failed is normal
                # validation. The problematic live traces carry long, current-
                # instance identity inventories.
                if len(alternatives) < 4:
                    continue

                specific: list[str] = []
                for alternative in alternatives:
                    normalized = re.sub(
                        r"[^a-z0-9@._]+",
                        " ",
                        alternative.lower().replace(r"\|", " "),
                    ).strip()
                    if len(normalized) < 3:
                        continue
                    if normalized in normalized_contract:
                        continue
                    words = set(normalized.split())
                    if words and words <= _GENERIC_IDENTITY_AUDIT_TERMS:
                        continue
                    specific.append(normalized)

                if len(specific) >= 3:
                    return True
        return False

    @staticmethod
    def _manual_direct_artifact_construction(command: str) -> bool:
        """Detect caller-side serialization of the final task artifact.

        Importing low-level utilities from an evolved package is not an
        end-to-end treatment when the fresh oracle still selects policy,
        assembles results, and writes the requested artifact itself. Runtime
        input/output paths may be passed to an end-to-end helper; this detector
        targets direct Python serialization primitives in the caller script.
        """

        def literal_string(node: ast.AST | None, bindings: dict[str, str]) -> str | None:
            if isinstance(node, ast.Constant) and isinstance(node.value, str):
                return node.value
            if isinstance(node, ast.Name):
                return bindings.get(node.id)
            return None

        def is_task_artifact(path_value: str | None) -> bool:
            if not path_value:
                return False
            normalized = path_value.strip()
            if normalized.startswith(("/tmp/", "/app/environment/skills/evo-")):
                return False
            name = Path(normalized).name
            if name.startswith("_"):
                return False
            return Path(name).suffix.lower() in _CALLER_ARTIFACT_SUFFIXES

        for source in _trace_python_sources(command):
            try:
                tree = ast.parse(source)
            except SyntaxError:
                continue

            bindings: dict[str, str] = {}
            for node in ast.walk(tree):
                if isinstance(node, ast.Assign):
                    value = literal_string(node.value, bindings)
                    if value is not None:
                        for target in node.targets:
                            if isinstance(target, ast.Name):
                                bindings[target.id] = value
                elif isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
                    value = literal_string(node.value, bindings)
                    if value is not None:
                        bindings[node.target.id] = value

            for node in ast.walk(tree):
                if not isinstance(node, ast.Call):
                    continue
                if isinstance(node.func, ast.Attribute):
                    method = node.func.attr.lower()
                    if method not in _CALLER_ARTIFACT_WRITE_METHODS:
                        continue
                    path_node = node.args[0] if node.args else None
                    for keyword in node.keywords:
                        if keyword.arg in {"path", "path_or_buf", "filename", "fname"}:
                            path_node = keyword.value
                            break
                    if is_task_artifact(literal_string(path_node, bindings)):
                        return True
                elif isinstance(node.func, ast.Name) and node.func.id == "open" and node.args:
                    path_value = literal_string(node.args[0], bindings)
                    mode_node = node.args[1] if len(node.args) > 1 else None
                    mode = literal_string(mode_node, bindings) or "r"
                    if any(flag in mode for flag in ("w", "a", "x")) and is_task_artifact(path_value):
                        return True
        return False

    @staticmethod
    def _manual_exact_replacement_fallback(
        command: str,
        public_contract: str = "",
    ) -> bool:
        """Reject caller-authored exact old/new mutation pairs.

        Runtime paths and values derived by an immutable end-to-end helper are
        valid treatment use. A fresh caller may not compensate for an
        incomplete helper by importing a lower-level replacement primitive and
        supplying the current answer pair itself. Exact pairs explicitly
        prescribed by the public task contract remain allowed.
        """
        if not _TRACE_REPLACEMENT_PRIMITIVE_RE.search(command):
            return False

        values: dict[str, list[str]] = {"old": [], "new": []}
        for match in _TRACE_EXACT_REPLACEMENT_ASSIGNMENT_RE.finditer(command):
            value = match.group("value").strip()
            if value:
                values[match.group("side").lower()].append(value)

        if not values["old"] or not values["new"]:
            return False

        contract = public_contract.lower()
        for old_value in values["old"]:
            for new_value in values["new"]:
                if old_value == new_value:
                    continue
                if old_value.lower() in contract and new_value.lower() in contract:
                    continue
                return True
        return False

    @staticmethod
    async def _container_evolved_skill_digest(
        environment: BaseEnvironment,
    ) -> str | None:
        """Hash immutable evolved-Skill sources inside an oracle container."""
        result = await environment.exec(
            command=(
                "find /app/environment/skills/evo-* -type f "
                "! -path '*/__pycache__/*' ! -name '*.pyc' -print0 2>/dev/null "
                "| sort -z | xargs -0 sha256sum | sha256sum"
            ),
            timeout_sec=30,
        )
        if result.return_code != 0:
            return None
        matches = re.findall(r"(?m)^([0-9a-fA-F]{64})\b", result.stdout or "")
        return matches[-1].lower() if matches else None

    async def _run_gt_oracle_check(
        self,
        instruction: str,
        environment: BaseEnvironment,
        oracle_model: str | None = None,
        evolved_skills_dir: Path | None = None,
        oracle_label: str | None = None,
    ) -> dict | None:
        """Spawn a clean independent agent to re-execute the task, then run GT tests.

        A fresh Docker container is created from the same Dockerfile and a
        fresh agent executes the task with the newly evolved Skill. Depending
        on ``gt_oracle_agent``, the oracle either preserves the paired public
        background document (``claude-code-skills``) or physically removes it before the
        agent starts (``claude-code-skill-only``). Ground truth tests are then
        run inside that container.

        Design constraints:
        - ISOLATION: Fresh LLM session + clean container. No shared state with the
          evolution agent — prevents context contamination.
        - INFORMATION BARRIER: Only ``passed`` (True/False) is used by the evolution
          loop. Full test details are stored in ``oracle_stop.json`` for the experiment
          runner but never fed back to the agent.
        - COST: Uses ``oracle_model`` if provided, otherwise falls back to the
          evolution agent's own model (``self._model_name``).

        Returns:
            Dict with ``passed``, ``tests_passed``, ``total_tests``, ``pass_rate``,
            ``reward`` keys, or None on errors.
        """
        oracle_model = oracle_model or self._model_name
        if not oracle_model:
            logger.warning("GT oracle check: no model available")
            return None

        # Resolve task directory and load task config
        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir:
            logger.warning("GT oracle check: no environment_dir on environment")
            return None
        task_dir = Path(env_dir).parent
        task_toml_path = task_dir / "task.toml"
        if not task_toml_path.exists():
            logger.warning("GT oracle check: task.toml not found at %s", task_toml_path)
            return None

        task_config = TaskConfig.model_validate(tomllib.loads(task_toml_path.read_text(encoding="utf-8")))
        task_name = task_dir.name

        claude_skill_oracles = {
            "claude-code-skills",
            "claude-code-skill-only",
        }
        if self._gt_oracle_agent in claude_skill_oracles:
            evolved_skill_names = {
                path.name
                for path in (evolved_skills_dir or Path()).iterdir()
                if path.is_dir() and path.name.startswith("evo-")
            } if evolved_skills_dir is not None else set()
            if not evolved_skill_names:
                logger.warning("GT oracle rejected: evolution produced no Skill")
                return {
                    "passed": False,
                    "tests_passed": 0,
                    "total_tests": 1,
                    "pass_rate": 0.0,
                    "reward": 0.0,
                    "failure_reasons": ["Evolution produced no Skill"],
                    "raw_output": "No evolved Skill was available to the GT oracle.",
                }

        # Create a temporary directory for oracle trial paths
        temp_dir = Path(tempfile.mkdtemp(prefix=f"oracle-{task_name}-"))
        oracle_trial_paths = TrialPaths(trial_dir=temp_dir)
        oracle_trial_paths.mkdir()

        # Create a fresh DockerEnvironment (reuses the cached image from the same Dockerfile)
        oracle_session_id = f"oracle-{task_name}-{uuid4().hex[:8]}"
        oracle_env = DockerEnvironment(
            environment_dir=Path(env_dir),
            environment_name=task_name,
            session_id=oracle_session_id,
            trial_paths=oracle_trial_paths,
            task_env_config=task_config.environment,
        )

        try:
            await oracle_env.start(force_build=False)

            if self._gt_oracle_agent == "claude-code-skills":
                # Preserve the baked-in public background document for the paired
                # background document-plus-Skill treatment.
                doc_barrier = await oracle_env.exec(
                    command=(
                        "test -d /app/environment/doc "
                        "&& find /app/environment/doc -type f -name '*.md' "
                        "-print -quit | grep -q ."
                    ),
                    timeout_sec=15,
                )
                if doc_barrier.return_code != 0:
                    raise RuntimeError(
                        "GT oracle paired-treatment barrier failed: "
                        "the public background document is unavailable"
                    )
            else:
                # Every release-standard oracle is Skill-only, regardless of
                # whether the fresh executor is Claude Code, Codex, Gemini CLI,
                # or Terminus-2. Remove the background document from the short-lived
                # container and fail closed if it remains. The paired
                # claude-code-skills agent above is an explicit ablation only.
                doc_barrier = await oracle_env.exec(
                    command=(
                        "rm -rf -- /app/environment/doc "
                        "&& test ! -e /app/environment/doc"
                    ),
                    timeout_sec=15,
                )
                if doc_barrier.return_code != 0:
                    raise RuntimeError(
                        "GT oracle Skill-only barrier failed: "
                        "the background document path could not be removed"
                    )

            # Inject evolved skills into the oracle container so the oracle agent
            # can discover and use them via SkillDocLoader
            if evolved_skills_dir is not None:
                for skill_subdir in sorted(evolved_skills_dir.iterdir()):
                    if skill_subdir.is_dir():
                        success = await self._inject_skill_dir_to_container(
                            skill_subdir.name,
                            skill_subdir,
                            oracle_env,
                        )
                        if success:
                            logger.info("Injected evolved skill %s into oracle container", skill_subdir.name)
                        else:
                            logger.warning("Failed to inject evolved skill %s into oracle container", skill_subdir.name)

                # Symlink injected skills to all agent-specific discovery paths
                # so that any oracle agent type (Codex, Claude Code, etc.) can find them
                await oracle_env.exec(
                    command=(
                        "for d in /root/.codex/skills /root/.claude/skills /root/.terminus/skills "
                        "/root/.opencode/skills /root/.agents/skills /root/.gemini/skills "
                        "/root/.goose/skills /root/.factory/skills; do "
                        'mkdir -p "$d" && ln -sf /app/environment/skills/evo-* "$d/" 2>/dev/null; '
                        "done"
                    ),
                    timeout_sec=15,
                )

            # Instantiate a fresh oracle agent (new LLM session — complete isolation)
            oracle_agent = self._create_oracle_agent(
                self._gt_oracle_agent,
                logs_dir=oracle_trial_paths.agent_dir,
                model_name=oracle_model,
            )
            await oracle_agent.setup(oracle_env)

            skill_digest_before = await self._container_evolved_skill_digest(
                oracle_env
            )
            if skill_digest_before is None:
                raise RuntimeError(
                    "GT oracle immutable-Skill barrier could not hash evolved sources"
                )

            # Installed agents (claude-code, codex) need extra time for CLI setup.
            # Use task.toml agent timeout as the base (respects per-task tuning),
            # and apply the timeout multiplier for consistency.
            task_agent_timeout = int(task_config.agent.timeout_sec) if task_config.agent and task_config.agent.timeout_sec else 3600
            task_agent_timeout = int(task_agent_timeout * self._timeout_multiplier)
            oracle_timeout = task_agent_timeout

            # Augment instruction with skill discovery hint for ALL oracle agents.
            # Even agents with native skill discovery (Claude Code, Terminus-2)
            # benefit from an explicit hint to prioritize loading evolved skills.
            oracle_instruction = instruction
            if evolved_skills_dir is not None:
                skill_names = sorted(d.name for d in evolved_skills_dir.iterdir() if d.is_dir())
                if skill_names:
                    agent_skills_path = {
                        "codex": "/root/.codex/skills",
                        "codex-skill-only": "/logs/agent/skills",
                        "codex-subscription": "/logs/agent/skills",
                        "gemini-cli": "/root/.gemini/skills",
                        "claude-code": "/root/.claude/skills",
                        "claude-code-skills": "/root/.claude/skills",
                        "claude-code-skill-only": "/root/.claude/skills",
                        "terminus-2": "/root/.terminus/skills",
                    }.get(self._gt_oracle_agent, "/root/.agents/skills")

                    # Extract code examples from each skill's SKILL.md to give the
                    # oracle agent a concrete, copy-pasteable invocation pattern
                    code_examples = []
                    for sname in skill_names:
                        skill_md_path = evolved_skills_dir / sname / "SKILL.md"
                        if skill_md_path.exists():
                            md_text = skill_md_path.read_text(encoding="utf-8", errors="ignore")
                            # Extract first ```python ... ``` block as the usage example
                            import re as _re

                            m = _re.search(r"```python\n(.+?)```", md_text, _re.DOTALL)
                            if m:
                                example = m.group(1).strip()
                                # Rewrite /app/environment/skills/ paths to the agent-specific path
                                example = example.replace("/app/environment/skills/", f"{agent_skills_path}/")
                                code_examples.append(f"# Skill: {sname}\n{example}")

                    hint = (
                        f"IMPORTANT: Task-specific skills with ready-to-use utility scripts are available at {agent_skills_path}/. "
                        f"Skills: {', '.join(skill_names)}. "
                        "Inspect the relevant evolved Skill and treat it as reusable guidance and a starting implementation. "
                        "You may use its helpers, then inspect "
                        "the runtime, adapt parameters, install dependencies, test, and make any additional task-file changes "
                        "needed to satisfy the instruction. Do not modify the injected Skill itself. "
                        "If the documented Python example includes an end-to-end entry point or validator, prefer to execute it "
                        "with the supplied runtime paths, then diagnose and finish any remaining work. "
                        "The skill scripts encode domain-specific semantics (e.g., ordering rules, unit conversions) "
                        "that are critical for correctness and very difficult to reproduce from scratch.\n"
                    )
                    if code_examples:
                        hint += "\nReady-to-use code (adapt paths if needed):\n```python\n"
                        hint += "\n\n".join(code_examples)
                        hint += "\n```\n\n"
                    else:
                        hint += "\n"

                    oracle_instruction = hint + instruction

            # Run the oracle agent with a timeout
            await asyncio.wait_for(
                oracle_agent.run(oracle_instruction, oracle_env, AgentContext()),
                timeout=oracle_timeout,
            )

            skill_digest_after = await self._container_evolved_skill_digest(
                oracle_env
            )
            if skill_digest_after is None or skill_digest_after != skill_digest_before:
                logger.warning(
                    "GT oracle rejected: clean evaluation modified the evolved Skill"
                )
                return {
                    "passed": False,
                    "tests_passed": 0,
                    "total_tests": 1,
                    "pass_rate": 0.0,
                    "reward": 0.0,
                    "failure_reasons": [
                        "Clean evaluation modified the immutable evolved Skill"
                    ],
                    "raw_output": (
                        "The evolved Skill source digest changed during fresh evaluation."
                    ),
                }

            # Skill invocation is an attribution diagnostic, not a correctness
            # gate.  The treatment condition makes the evolved Skill available;
            # per the benchmark contract, only the canonical GT result determines
            # whether the fresh Agent completed the task.  Keep the observation in
            # the result so analyses can separate uptake from mere availability.
            skill_invoked: bool | None = None
            if self._gt_oracle_agent in claude_skill_oracles:
                expected_skill_names = {
                    path.name
                    for path in (evolved_skills_dir or Path()).iterdir()
                    if path.is_dir() and path.name.startswith("evo-")
                } if evolved_skills_dir is not None else set()
                trace_path = oracle_trial_paths.agent_dir / "claude-code.txt"
                skill_invoked = self._claude_trace_used_evolved_skill(
                    trace_path,
                    expected_skill_names,
                )
                if not skill_invoked:
                    logger.warning(
                        "GT oracle diagnostic: Claude Code did not read or invoke "
                        "an evolved Skill; canonical GT will still determine success"
                    )

            # Run ground truth tests inside the oracle container
            gt_result = await self._run_ground_truth_evaluation(oracle_env)
            if gt_result is None:
                logger.warning("GT oracle check: ground truth evaluation returned None")
                return {"passed": None, "error": "ground truth evaluation returned None"}

            total = gt_result.get("total_tests", 0)
            passed = gt_result.get("tests_passed", 0)
            full_score, oracle_reward = self._gt_full_score_and_reward(gt_result)
            return {
                "passed": full_score,
                "tests_passed": passed,
                "total_tests": total,
                "pass_rate": passed / total if total > 0 else 0.0,
                "reward": oracle_reward,
                "skill_invoked": skill_invoked,
                "test_details": gt_result.get("test_details"),
                "failure_reasons": gt_result.get("failure_reasons"),
                "raw_output": gt_result.get("raw_output"),
            }

        except TimeoutError:
            logger.warning("GT oracle check timed out after %ds", oracle_timeout)
            return {"passed": None, "error": f"timeout after {oracle_timeout}s"}
        except Exception as exc:
            if isinstance(exc, ClaudeCodeProviderError):
                raise
            # Preserve the traceback.  A message-only warning made unrelated
            # Docker, agent-log, and verifier failures all appear as the same
            # opaque "infrastructure unavailable" result during campaigns.
            logger.exception("GT oracle check failed: %s", exc)
            return {"passed": None, "error": str(exc)}
        finally:
            try:
                await oracle_env.stop(delete=True)
            except Exception as stop_exc:
                logger.warning("GT oracle check: failed to stop oracle container: %s", stop_exc)
            # Persist oracle logs for debugging before cleaning up
            label = oracle_label or f"gt-oracle-{self._host_intervention_count}"
            oracle_persist_dir = self.logs_dir / label
            try:
                if temp_dir.exists():
                    # Codex may populate CODEX_HOME with root-owned plugin caches
                    # and SQLite state. They are both noisy and unnecessary for
                    # debugging the oracle. Persist only the execution transcript,
                    # command/setup logs, and verifier outputs.
                    oracle_persist_dir.mkdir(parents=True, exist_ok=True)
                    agent_src = temp_dir / "agent"
                    agent_dst = oracle_persist_dir / "agent"
                    agent_dst.mkdir(parents=True, exist_ok=True)
                    # Keep the native transcript used by the invocation gate so
                    # offline analysis can distinguish a weak Skill from a Skill
                    # that the clean agent never actually loaded.  These files
                    # contain no subscription credentials.
                    for file_name in ("codex.txt", "claude-code.txt", "trajectory.json"):
                        source = agent_src / file_name
                        if source.is_file():
                            shutil.copy2(source, agent_dst / file_name)
                    for pattern in ("setup", "command-*"):
                        for source in agent_src.glob(pattern):
                            if source.is_dir():
                                shutil.copytree(
                                    source,
                                    agent_dst / source.name,
                                    dirs_exist_ok=True,
                                )
                    verifier_src = temp_dir / "verifier"
                    if verifier_src.is_dir():
                        shutil.copytree(
                            verifier_src,
                            oracle_persist_dir / "verifier",
                            dirs_exist_ok=True,
                        )
                    logger.info("Persisted GT oracle logs to %s", oracle_persist_dir)
            except Exception as copy_exc:
                logger.warning("Failed to persist oracle logs: %s", copy_exc)
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _write_oracle_stop_signal(self, task_name: str, gt_result: dict | None = None) -> None:
        """Write oracle_stop.json so the experiment runner can detect early-stop between sequential rounds.

        When ``gt_result`` is provided, the oracle's GT test results (reward, tests_passed,
        total_tests) are stored alongside the stop signal.  The experiment runner reads
        these to use the oracle reward instead of the (potentially stale) evolution
        container reward.
        """
        evo_dir = self._resolve_evolution_dir()
        evo_dir.mkdir(parents=True, exist_ok=True)
        stop_file = evo_dir / "oracle_stop.json"
        payload: dict = {
            "task_name": task_name,
            "timestamp": datetime.now(UTC).isoformat(),
            "sequential_run": self._sequential_run,
        }
        if gt_result is not None:
            payload["gt_result"] = gt_result
        stop_file.write_text(json.dumps(payload))
        logger.info("Oracle stop signal written to %s", stop_file)

    # ------------------------------------------------------------------
    # Evolution helpers
    # ------------------------------------------------------------------

    async def _read_progress_checklist(self, environment: BaseEnvironment) -> str:
        """Read /root/progress.md from container."""
        result = await environment.exec(
            command="cat /root/progress.md 2>/dev/null || true",
            timeout_sec=15,
        )
        return (result.stdout or "").strip()

    @staticmethod
    def _parse_unchecked_items(progress_text: str) -> list[str]:
        """Extract unchecked items from progress checklist."""
        unchecked = []
        for line in progress_text.splitlines():
            line = line.strip()
            if line.startswith("- [ ]"):
                unchecked.append(line[6:].strip())
        return unchecked

    async def _read_evolution_summary(self, environment: BaseEnvironment) -> str:
        """Read /root/evolution_summary.md from the container."""
        result = await environment.exec(
            command="cat /root/evolution_summary.md 2>/dev/null || true",
            timeout_sec=15,
        )
        raw = (result.stdout or "").strip()
        if not raw:
            logger.info("No evolution_summary.md found in container")
            return ""
        # Filter out bash shell warnings that precede actual content
        lines = [line for line in raw.splitlines() if not is_shell_warning(line)]
        content = "\n".join(lines).strip()
        logger.info("Read evolution_summary.md from container (%d chars)", len(content))
        return content

    async def _list_container_skills(self, environment: BaseEnvironment) -> set[str]:
        """List skill directory names under /app/environment/skills/ in the container."""
        result = await environment.exec(
            command="ls -1 /app/environment/skills/ 2>/dev/null || true",
            timeout_sec=15,
        )
        names = set()
        for line in (result.stdout or "").strip().splitlines():
            name = line.strip()
            if not name or is_shell_warning(line):
                continue
            names.add(name)

        # If exec returned nothing useful (service down), fall back to docker compose cp
        if not names and hasattr(environment, "download_dir"):
            try:
                import tempfile

                with tempfile.TemporaryDirectory() as tmp:
                    await environment.download_dir("/app/environment/skills/", tmp)
                    skills_tmp = Path(tmp) / "skills"
                    if skills_tmp.exists():
                        names = {d.name for d in skills_tmp.iterdir() if d.is_dir()}
                        logger.info("Listed %d skills via download_dir fallback", len(names))
            except Exception as dl_exc:
                logger.warning("download_dir fallback for skill listing failed: %s", dl_exc)

        return names

    @staticmethod
    def _strip_code_fences(text: str) -> str:
        """Strip markdown code fences (```json ... ```) from LLM responses."""
        cleaned = text.strip()
        if cleaned.startswith("```"):
            first_newline = cleaned.find("\n")
            if first_newline != -1:
                cleaned = cleaned[first_newline + 1 :]
            if cleaned.rstrip().endswith("```"):
                cleaned = cleaned.rstrip()[: -len("```")].rstrip()
        return cleaned

    @staticmethod
    def _clean_base64(raw: str) -> str:
        """Filter shell warnings from base64 output.

        Shell environments (especially Docker containers) may emit warnings
        to stdout (e.g., ``bash: cannot set terminal process group``).
        These corrupt the base64 string.  Filter to only valid base64 lines.
        """
        lines = []
        for line in raw.strip().splitlines():
            stripped = line.strip()
            if is_shell_warning(line):
                continue
            # Keep only lines that look like valid base64
            if re.fullmatch(r"[A-Za-z0-9+/=]+", stripped):
                lines.append(stripped)
        return "".join(lines)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_task_name(environment: BaseEnvironment) -> str:
        """Extract a task name from the environment."""
        env_dir = getattr(environment, "environment_dir", None)
        if env_dir:
            # environment_dir is typically .../tasks/task-name/environment
            task_dir = Path(env_dir).parent
            return task_dir.name
        return "unknown"

    def _read_trajectory(self) -> list[dict] | None:
        """Read and parse trajectory.json once. Returns None on failure."""
        trajectory_path = self.logs_dir / "trajectory.json"
        if not trajectory_path.exists():
            return None
        try:
            return json.loads(trajectory_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None

    def _build_trajectory_summary(self, trajectory: list[dict] | None = None) -> str:
        """Build a rich trajectory summary including command outputs and agent analysis."""
        if trajectory is None:
            trajectory = self._read_trajectory()
        if trajectory is None:
            return "No trajectory data available."

        lines: list[str] = []
        for i, entry in enumerate(trajectory):
            if not isinstance(entry, dict):
                lines.append(f"Episode {i}: Non-object trajectory entry")
                continue
            ep = entry.get("episode", "?")
            response = entry.get("response", "")
            prompt = entry.get("prompt", "")

            # Include terminal output from this episode's prompt (shows results of previous commands)
            # Skip the first episode's prompt (it's the full system prompt, not command output)
            if i > 0 and prompt:
                output_text = prompt.strip()
                for prefix in ["Current Terminal Screen:\n", "New Terminal Output:\n"]:
                    if output_text.startswith(prefix):
                        output_text = output_text[len(prefix) :]
                        break
                if output_text:
                    lines.append(f"Episode {ep} terminal output:\n{output_text[:1500]}")

            # Detect action type from response
            if "load_skill" in response:
                match = re.search(r'"load_skill"\s*:\s*"([^"]+)"', response)
                if match:
                    lines.append(f"Episode {ep}: Loaded skill {match.group(1)}")
                else:
                    lines.append(f"Episode {ep}: Skill load request")
            elif "task_complete" in response and "true" in response.lower():
                lines.append(f"Episode {ep}: Task completed")
            else:
                try:
                    data = json.loads(self._strip_code_fences(response))
                    analysis = data.get("analysis", "")
                    if analysis:
                        lines.append(f"Episode {ep} analysis: {analysis[:500]}")
                    cmds = data.get("commands", [])
                    if cmds:
                        cmd_strs = [
                            (
                                c.get("cmd", c.get("keystrokes", ""))
                                if isinstance(c, dict)
                                else str(c)
                            )[:200]
                            for c in cmds
                        ]
                        lines.append(f"Episode {ep} commands: {'; '.join(cmd_strs)}")
                    else:
                        lines.append(f"Episode {ep}: Response (no commands)")
                except (json.JSONDecodeError, TypeError, AttributeError):
                    lines.append(f"Episode {ep}: Response (unparseable)")

        return "\n".join(lines) if lines else "Empty trajectory."

    def _build_execution_info(
        self,
        skills_loaded: list[str],
        trajectory_summary: str,
        trajectory: list[dict] | None = None,
    ) -> dict:
        """Build the execution section of the log."""
        if trajectory is None:
            trajectory = self._read_trajectory()

        trajectory_entries: list[dict] = []
        total_episodes = 0

        if trajectory:
            total_episodes = len(trajectory)

            for i, entry in enumerate(trajectory):
                if not isinstance(entry, dict):
                    trajectory_entries.append(
                        {
                            "episode": i,
                            "action": "unknown",
                            "detail": "Non-object trajectory entry",
                        }
                    )
                    continue
                ep = entry.get("episode", "?")
                response = entry.get("response", "")

                if "load_skill" in response:
                    match = re.search(r'"load_skill"\s*:\s*"([^"]+)"', response)
                    trajectory_entries.append(
                        {
                            "episode": ep,
                            "action": "skill_load",
                            "skill_name": match.group(1) if match else "unknown",
                            "detail": "Agent requested to load skill",
                        }
                    )
                elif "task_complete" in response and "true" in response.lower():
                    trajectory_entries.append(
                        {
                            "episode": ep,
                            "action": "task_complete",
                            "detail": "Agent signaled task completion",
                        }
                    )
                else:
                    try:
                        data = json.loads(self._strip_code_fences(response))
                        cmds = data.get("commands", [])
                        cmd_strs = [
                            (
                                c.get("cmd", c.get("keystrokes", ""))
                                if isinstance(c, dict)
                                else str(c)
                            )[:100]
                            for c in cmds[:5]
                        ]
                        output_preview = ""
                        if i + 1 < len(trajectory):
                            next_entry = trajectory[i + 1]
                            next_prompt = (
                                next_entry.get("prompt", "")
                                if isinstance(next_entry, dict)
                                else ""
                            )
                            output_preview = next_prompt.strip()[:200]
                        trajectory_entries.append(
                            {
                                "episode": ep,
                                "action": "commands",
                                "commands": cmd_strs,
                                "output_preview": output_preview,
                            }
                        )
                    except (json.JSONDecodeError, TypeError):
                        trajectory_entries.append(
                            {
                                "episode": ep,
                                "action": "unknown",
                                "detail": "Unparseable response",
                            }
                        )

        task_completed = any(e.get("action") == "task_complete" for e in trajectory_entries)

        return {
            "total_episodes": total_episodes,
            "task_completed": task_completed,
            "skills_loaded_by_agent": sorted(skills_loaded),
            "trajectory": trajectory_entries,
            "trajectory_summary": trajectory_summary,
        }

    def _write_evolution_log(
        self,
        task_name: str,
        run_timestamp: str,
        skills_loaded: list[str],
        trajectory_summary: str,
        context: AgentContext,
        timing: dict,
        evolution_summary: str = "",
        gt_oracle_result: dict | None = None,
        intervention_history: list[dict] | None = None,
        trajectory_data: list[dict] | None = None,
    ) -> None:
        """Write the comprehensive evolution_run_log.json."""
        log = {
            # Basic info
            "task_name": task_name,
            "model": self._model_name,
            "timestamp": run_timestamp,
            "sequential_run": self._sequential_run,
            "evolution_summary": evolution_summary,
            # Execution
            "execution": self._build_execution_info(
                skills_loaded,
                trajectory_summary,
                trajectory_data,
            ),
            # GT Oracle check
            "gt_oracle_agent": self._gt_oracle_agent,
            "gt_oracle_result": gt_oracle_result,
            # Best GT snapshot (highest GT score seen during iterations)
            "best_gt_snapshot": {
                "passed": self._best_gt_snapshot["passed"],
                "reward": self._best_gt_snapshot["reward"],
                "pass_rate": self._best_gt_snapshot["pass_rate"],
                "tests_passed": self._best_gt_snapshot["tests_passed"],
                "total_tests": self._best_gt_snapshot["total_tests"],
                "intervention_number": self._best_gt_snapshot["intervention_number"],
            }
            if self._best_gt_snapshot
            else None,
            # Intervention history (per-intervention GT scores + surrogate results)
            "intervention_history": intervention_history or [],
            # Deterministic format checks performed before verifier/GT/export.
            "skill_schema_validation_history": self._skill_schema_validation_history,
            # Token usage
            "token_usage": {
                "agent_input_tokens": context.n_input_tokens or 0,
                "agent_output_tokens": context.n_output_tokens or 0,
            },
            # Timing
            "timing": timing,
        }

        log_path = self.logs_dir / "evolution_run_log.json"
        log_path.write_text(
            json.dumps(log, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        logger.info("Evolution run log written to %s", log_path)

        # Generate Markdown report alongside the JSON log
        from libs.terminus_agent.evolution.report_generator import generate_evolution_report

        md_content: str | None = None
        try:
            md_content = generate_evolution_report(log)
            md_path = self.logs_dir / "evolution_report.md"
            md_path.write_text(md_content, encoding="utf-8")
            logger.info("Evolution report written to %s", md_path)
        except Exception as exc:
            logger.warning("Failed to generate evolution report: %s", exc)

        # Phase 9: Sync verifier results to task environment directory
        self._sync_verifier_to_environment(log, md_content)

    def _sync_verifier_to_environment(self, log: dict, md_content: str | None = None) -> None:
        """Save verifier/evolution results to the task's environment directory.

        Writes to ``environment/verifier/`` so the data persists alongside
        skills and is available inside Docker on the next build.
        """
        env_dir = getattr(self._environment, "environment_dir", None)
        if not env_dir:
            logger.debug("No environment_dir available; skipping verifier sync")
            return

        verifier_dir = Path(env_dir) / "verifier"
        try:
            verifier_dir.mkdir(parents=True, exist_ok=True)

            # Full evolution run log
            (verifier_dir / "evolution_run_log.json").write_text(
                json.dumps(log, indent=2, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )

            # Extract a compact verification summary for quick inspection
            summary = {
                "task_name": log.get("task_name"),
                "timestamp": log.get("timestamp"),
                "gt_oracle_result": log.get("gt_oracle_result"),
                "intervention_history": log.get("intervention_history", []),
                "evolution_summary": log.get("evolution_summary", ""),
            }
            (verifier_dir / "verification_summary.json").write_text(
                json.dumps(summary, indent=2, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )

            # Markdown report (reuse pre-generated content)
            if md_content:
                (verifier_dir / "evolution_report.md").write_text(md_content, encoding="utf-8")

            logger.info("Verifier results synced to %s", verifier_dir)
        except Exception as exc:
            logger.warning("Failed to sync verifier results to environment: %s", exc)

    async def _export_verifier_script(self, environment: BaseEnvironment) -> str | None:
        """Export /root/verifier/test_outputs.py from container to host.

        Returns the host path where the script was saved, or None on failure.
        Versions previous scripts (test_outputs_generated_v1.py, v2.py, ...).
        """
        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir:
            return None

        result = await environment.exec(
            command="base64 /root/verifier/test_outputs.py 2>/dev/null || true",
            timeout_sec=15,
        )
        encoded = self._clean_base64(result.stdout or "")
        if not encoded:
            return None

        try:
            content = base64.b64decode(encoded).decode("utf-8")
        except Exception:
            logger.warning("Failed to decode verifier script from container")
            return None

        scripts_dir = Path(env_dir) / "verifier" / "generated_scripts"
        scripts_dir.mkdir(parents=True, exist_ok=True)

        # Version by counting existing files
        existing = sorted(scripts_dir.glob("test_outputs_v*.py"))
        next_version = len(existing) + 1
        target = scripts_dir / f"test_outputs_v{next_version}.py"

        target.write_text(content, encoding="utf-8")
        logger.info("Exported verifier script to %s", target)
        return str(target)

    async def _export_skills_from_container(
        self,
        environment: BaseEnvironment,
        skill_names: set[str],
        target_dir: Path,
        include_references: bool = False,
    ) -> list[str]:
        """Export skill directories from container to a host directory.

        Primary method: base64 via exec (works when terminal service is up).
        Fallback: docker compose cp via environment.download_dir (works even when
        the terminal service is down, e.g. after agent timeout or completion).

        Copies SKILL.md and scripts/ for each skill. Optionally copies references/.

        Returns:
            List of successfully exported skill names.
        """
        exported: list[str] = []

        for skill_name in sorted(skill_names):
            skill_host_dir = target_dir / skill_name
            skill_host_dir.mkdir(parents=True, exist_ok=True)

            # --- Primary path: base64 via exec ---
            encoded = await self._read_stable_container_text(
                environment,
                f"/app/environment/skills/{skill_name}/SKILL.md",
            )

            if encoded is not None:
                try:
                    content = base64.b64decode(encoded).decode("utf-8")
                    (skill_host_dir / "SKILL.md").write_text(content, encoding="utf-8")
                    # Export scripts/
                    await self._export_container_subdir(environment, skill_name, "scripts", skill_host_dir)
                    # Export references/ (only for oracle export)
                    if include_references:
                        await self._export_container_subdir(environment, skill_name, "references", skill_host_dir)
                    exported.append(skill_name)
                    logger.info("Exported skill %s to %s", skill_name, skill_host_dir)
                    continue
                except Exception:
                    logger.warning("Failed to decode SKILL.md for %s, trying download_dir fallback", skill_name)
                    # clean up partial dir so fallback can overwrite
                    import shutil

                    shutil.rmtree(skill_host_dir, ignore_errors=True)
            else:
                logger.warning("Skill %s: exec-based export returned no data, trying download_dir fallback", skill_name)

            # --- Fallback: docker compose cp (works even when terminal service is down) ---
            if not hasattr(environment, "download_dir"):
                logger.warning("Skill %s has no SKILL.md and download_dir not available, skipping", skill_name)
                continue
            try:
                import tempfile, shutil

                with tempfile.TemporaryDirectory() as tmp:
                    container_skill_path = f"/app/environment/skills/{skill_name}"
                    await environment.download_dir(container_skill_path, tmp)
                    # download_dir puts contents under tmp/<skill_name>/
                    src = Path(tmp) / skill_name
                    if not src.exists():
                        # some implementations put files directly under tmp/
                        src = Path(tmp)
                    skill_md = src / "SKILL.md"
                    if not skill_md.exists():
                        logger.warning("Skill %s has no SKILL.md even after download_dir, skipping", skill_name)
                        continue
                    skill_host_dir.mkdir(parents=True, exist_ok=True)
                    # Copy SKILL.md
                    shutil.copy2(skill_md, skill_host_dir / "SKILL.md")
                    # Copy scripts/
                    scripts_src = src / "scripts"
                    if scripts_src.exists():
                        shutil.copytree(scripts_src, skill_host_dir / "scripts", dirs_exist_ok=True)
                    # Copy references/ if requested
                    if include_references:
                        refs_src = src / "references"
                        if refs_src.exists():
                            shutil.copytree(refs_src, skill_host_dir / "references", dirs_exist_ok=True)
                exported.append(skill_name)
                logger.info("Exported skill %s via download_dir fallback to %s", skill_name, skill_host_dir)
            except Exception as dl_exc:
                logger.warning("download_dir fallback failed for skill %s: %s", skill_name, dl_exc)

        return exported

    async def _export_container_subdir(
        self,
        environment: BaseEnvironment,
        skill_name: str,
        subdir: str,
        skill_host_dir: Path,
    ) -> None:
        """Export a subdirectory (scripts/ or references/) of a skill from container."""
        container_subdir = f"/app/environment/skills/{skill_name}/{subdir}"
        ls_result = await environment.exec(
            command=(
                f"find {shlex.quote(container_subdir)} -type f "
                "! -path '*/__pycache__/*' ! -name '*.pyc' -printf '%P\\n' "
                "2>/dev/null || true"
            ),
            timeout_sec=15,
        )
        file_names = [
            s.strip()
            for s in (ls_result.stdout or "").strip().splitlines()
            if s.strip() and not is_shell_warning(s)
        ]
        if not file_names:
            return

        host_subdir = skill_host_dir / subdir
        host_subdir.mkdir(parents=True, exist_ok=True)
        for fname in file_names:
            relative = Path(fname)
            if relative.is_absolute() or ".." in relative.parts:
                logger.warning("Skipping unsafe exported Skill path: %s", fname)
                continue
            file_encoded = await self._read_stable_container_text(
                environment,
                f"{container_subdir}/{fname}",
            )
            if file_encoded is not None:
                try:
                    file_content = base64.b64decode(file_encoded).decode("utf-8")
                    target = host_subdir / relative
                    target.parent.mkdir(parents=True, exist_ok=True)
                    target.write_text(file_content, encoding="utf-8")
                except Exception as exc:
                    logger.warning("Failed to decode %s/%s/%s: %s", skill_name, subdir, fname, exc)

    async def _read_stable_container_text(
        self,
        environment: BaseEnvironment,
        container_path: str,
    ) -> str | None:
        """Read a Skill text file only after its bytes have stopped changing.

        The evolution agent can rewrite a helper in place just as a
        task-complete preflight snapshots the Skill. Reading that half-written
        file made the information-boundary audit silently omit it. Two matching
        checksums around a short quiet period provide a coherent snapshot;
        invalid UTF-8 is rejected and retried instead of being skipped.
        """
        quoted_path = shlex.quote(container_path)
        for attempt in range(4):
            result = await environment.exec(
                command=(
                    f"p={quoted_path}; "
                    "[ -f \"$p\" ] || exit 1; "
                    "h1=$(sha256sum \"$p\" | cut -d' ' -f1); "
                    "sleep 0.15; "
                    "h2=$(sha256sum \"$p\" | cut -d' ' -f1); "
                    "[ -n \"$h1\" ] && [ \"$h1\" = \"$h2\" ] || exit 2; "
                    "base64 \"$p\""
                ),
                timeout_sec=15,
            )
            encoded = self._clean_base64(result.stdout or "")
            if result.return_code == 0:
                try:
                    base64.b64decode(encoded).decode("utf-8")
                    return encoded
                except (ValueError, UnicodeDecodeError):
                    pass
            if attempt < 3:
                await asyncio.sleep(0.2 * (attempt + 1))
        logger.warning("Could not obtain a stable UTF-8 snapshot of %s", container_path)
        return None

    async def _import_agent_created_skills(
        self,
        environment: BaseEnvironment,
        pre_existing_skills: set[str],
    ) -> list[str]:
        """Import agent-created skills from container to host environment directory."""
        schema_issues = await self._validate_evolved_skill_schema(environment)
        if schema_issues:
            logger.warning(
                "Not importing invalid evolved Skills: %s",
                "; ".join(schema_issues),
            )
            return []

        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir:
            return []

        current_skills = await self._list_container_skills(environment)
        # Existing evo-* packages may have been modified in-place during a
        # continuation. Persist all of them; subtracting pre-existing names
        # silently discarded exactly the updates continuation is meant to keep.
        evolved_skill_names = self._select_evolved_skill_names(current_skills)

        if not evolved_skill_names:
            logger.info("No evolved skills to import")
            return []

        host_skills_dir = Path(env_dir) / "skills"
        host_skills_dir.mkdir(parents=True, exist_ok=True)

        return await self._export_skills_from_container(
            environment,
            set(evolved_skill_names),
            host_skills_dir,
            include_references=False,
        )

    def _rollback_host_skills(self, environment: BaseEnvironment, snapshot_skills_dir: Path) -> None:
        """Overwrite host environment skills with the best snapshot version.

        Replaces each evolved skill directory in ``environment/skills/`` with
        the corresponding directory from the snapshot, so that persisted skills
        match the GT-best version rather than the (potentially degraded) final
        container version.
        """
        env_dir = getattr(environment, "environment_dir", None)
        if not env_dir or not snapshot_skills_dir or not snapshot_skills_dir.is_dir():
            return

        host_skills_dir = Path(env_dir) / "skills"
        host_skills_dir.mkdir(parents=True, exist_ok=True)

        rolled_back: list[str] = []
        for skill_subdir in sorted(snapshot_skills_dir.iterdir()):
            if not skill_subdir.is_dir():
                continue
            target = host_skills_dir / skill_subdir.name
            try:
                if target.exists():
                    shutil.rmtree(target)
                shutil.copytree(skill_subdir, target)
                rolled_back.append(skill_subdir.name)
            except Exception as exc:
                logger.warning("Failed to rollback skill %s: %s", skill_subdir.name, exc)

        if rolled_back:
            logger.info("Rolled back %d host skills to best snapshot: %s", len(rolled_back), rolled_back)

    def _did_last_oracle_pass(self) -> bool:
        """Check whether the most recent intervention included a passing GT oracle."""
        if not self._intervention_history:
            return False
        gt_result = self._intervention_history[-1].get("gt_result")
        return gt_result is not None and gt_result.get("passed") is True

    async def _export_evolved_skills_to_host(self, environment: BaseEnvironment) -> Path | None:
        """Export evolved skills from the container to a temporary directory on the host.

        Returns:
            Path to the temp directory containing exported skills, or None if
            no evolved skills exist. Caller is responsible for cleanup.
        """
        current_skills = await self._list_container_skills(environment)
        evolved_skill_names = self._select_evolved_skill_names(current_skills)

        if not evolved_skill_names:
            logger.info("No evolved skills to export for oracle")
            return None

        temp_dir = Path(tempfile.mkdtemp(prefix="evolved-skills-"))

        exported = await self._export_skills_from_container(
            environment,
            set(evolved_skill_names),
            temp_dir,
            include_references=True,
        )

        if not exported:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None

        schema_issues = []
        for skill_name in exported:
            for issue in validate_skill_directory(temp_dir / skill_name):
                schema_issues.append(
                    f"{issue.path}: {issue.message} [{issue.code}]"
                )
        if schema_issues:
            self._skill_schema_validation_history.append(
                {
                    "timestamp": datetime.now(UTC).isoformat(),
                    "stage": "post_export",
                    "passed": False,
                    "manifests": [
                        str(temp_dir / name / "SKILL.md") for name in exported
                    ],
                    "issues": schema_issues,
                }
            )
            logger.error(
                "Exported evolved Skill failed schema validation: %s",
                "; ".join(schema_issues),
            )
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None

        logger.info("Exported %d evolved skills to %s", len(exported), temp_dir)
        return temp_dir

    async def _container_current_source_identifiers(
        self,
        environment: BaseEnvironment,
    ) -> dict[str, str]:
        """Extract a bounded rare-identifier index from visible task source.

        BugSwarm repositories are baked into the container and therefore are
        absent from ``environment_dir`` on the host.  This read-only snapshot
        deliberately excludes tests, verifier/solution assets, dependencies,
        generated build trees, background documents, and evolved Skills.  The host-side
        boundary audit applies the public-contract and generic-API filters.
        """
        scanner = r'''
import collections
import json
import os
import re

suffixes = {
    '.py', '.js', '.jsx', '.ts', '.tsx', '.java', '.scala', '.go', '.rs',
    '.c', '.cc', '.cpp', '.h', '.hpp', '.erl', '.ex', '.exs', '.sh',
    '.css', '.html', '.toml', '.yaml', '.yml', '.xml', '.gradle', '.kts',
}
excluded = {
    'doc', 'docs', 'skills', 'verifier', 'tests', 'test', 'testing',
    'solution', 'solutions', 'reference', 'references', 'fixtures',
    '.evolution', '.git', 'node_modules', '.next', 'dist', 'build', 'target',
    '__pycache__', '.tox', '.venv', 'vendor', 'output', 'outputs', 'cacher',
}
identifier_re = re.compile(r'(?<![A-Za-z0-9_])[A-Za-z_][A-Za-z0-9_]{7,}(?![A-Za-z0-9_])')
roots = set()
for base in ('/home/github/build/failed', '/home/travis/build/failed'):
    if not os.path.isdir(base):
        continue
    for current, dirs, _files in os.walk(base):
        if '.git' in dirs:
            roots.add(current)
            dirs[:] = []
            continue
        dirs[:] = [d for d in dirs if d.lower() not in excluded]
for base in ('/root',):
    if not os.path.isdir(base):
        continue
    try:
        children = os.listdir(base)
    except OSError:
        children = []
    for name in children:
        candidate = os.path.join(base, name)
        if os.path.isdir(os.path.join(candidate, '.git')):
            roots.add(candidate)
if os.path.isdir('/app'):
    roots.add('/app')

counts = collections.Counter()
paths = collections.defaultdict(set)
files_seen = 0
for root in sorted(roots):
    for current, dirs, files in os.walk(root):
        dirs[:] = [d for d in dirs if d.lower() not in excluded]
        for name in files:
            if files_seen >= 8000:
                break
            path = os.path.join(current, name)
            if os.path.splitext(name)[1].lower() not in suffixes:
                continue
            try:
                if os.path.getsize(path) > 1_000_000:
                    continue
                with open(path, encoding='utf-8', errors='replace') as handle:
                    text = handle.read()
            except OSError:
                continue
            files_seen += 1
            for identifier in identifier_re.findall(text):
                counts[identifier] += 1
                paths[identifier].add(os.path.relpath(path, root))
        if files_seen >= 8000:
            break

result = {
    identifier: sorted(identifier_paths)[0]
    for identifier, identifier_paths in paths.items()
    if counts[identifier] <= 12 and len(identifier_paths) <= 2
}
print(json.dumps(result, sort_keys=True))
'''
        encoded = base64.b64encode(scanner.encode("utf-8")).decode("ascii")
        result = await environment.exec(
            command=(
                "py=$(command -v python3 || command -v python || true); "
                '[ -n "$py" ] || exit 0; '
                f'"$py" -c "$(printf %s {shlex.quote(encoded)} | base64 -d)"'
            ),
            timeout_sec=60,
        )
        for line in reversed((result.stdout or "").splitlines()):
            line = line.strip()
            if not line.startswith("{"):
                continue
            try:
                parsed = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(parsed, dict):
                return {
                    str(identifier): str(origin)
                    for identifier, origin in parsed.items()
                    if isinstance(identifier, str) and isinstance(origin, str)
                }
        logger.warning("Could not obtain current-source identifier snapshot")
        return {}

    def _regenerate_conversation_md(self) -> None:
        """Re-generate full_conversation.md."""
        conv_path = self.logs_dir / "full_conversation.json"
        if not conv_path.exists():
            return
        try:
            messages = json.loads(conv_path.read_text(encoding="utf-8"))
            from .conversation_renderer import render_conversation_markdown

            md_content = render_conversation_markdown(messages)
            (self.logs_dir / "full_conversation.md").write_text(md_content, encoding="utf-8")
            logger.info("Re-generated full_conversation.md")
        except Exception as exc:
            logger.warning("Failed to re-generate conversation md: %s", exc)
